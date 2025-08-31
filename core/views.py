from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.utils.dateparse import parse_date
import json
from datetime import datetime

from .models import CustomUser, UserProfile, UMKMBusiness, WhatsAppBotConfig, SystemSettings, WebsiteSettings, ModuleSettings, APIEndpoint

User = get_user_model()


@login_required
def core_module_view(request):
    """Main core module view"""
    context = {
        'page_title': 'Manajemen Sistem',
        'page_subtitle': 'Kelola pengguna, UMKM, dan konfigurasi sistem'
    }
    return render(request, 'admin/modules/core.html', context)


# Core Statistics API
@login_required
def core_stats_api(request):
    """API for core module statistics"""
    try:
        stats = {
            'total_users': CustomUser.objects.count(),
            'active_users': CustomUser.objects.filter(is_active=True).count(),
            'village_staff': CustomUser.objects.filter(is_village_staff=True).count(),
            'total_umkm': UMKMBusiness.objects.count(),
            'active_umkm': UMKMBusiness.objects.filter(is_active=True).count(),
            'system_settings': SystemSettings.objects.filter(is_active=True).count(),
            'whatsapp_configs': WhatsAppBotConfig.objects.filter(is_active=True).count(),
            'recent_users': CustomUser.objects.filter(date_joined__gte=datetime.now().replace(day=1)).count()
        }
        
        return JsonResponse({
            'success': True,
            'data': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat statistik: {str(e)}'
        }, status=500)


# User Management API Views
@login_required
def users_list_api(request):
    """API for users list with pagination and search"""
    try:
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        search = request.GET.get('search', '')
        is_active = request.GET.get('is_active', '')
        is_village_staff = request.GET.get('is_village_staff', '')
        
        queryset = CustomUser.objects.all()
        
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(position__icontains=search)
            )
        
        if is_active:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
            
        if is_village_staff:
            queryset = queryset.filter(is_village_staff=is_village_staff.lower() == 'true')
        
        queryset = queryset.order_by('-date_joined')
        
        paginator = Paginator(queryset, per_page)
        page_obj = paginator.get_page(page)
        
        data = {
            'results': [
                {
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'email': user.email,
                    'phone_number': user.phone_number,
                    'position': user.position,
                    'is_active': user.is_active,
                    'is_village_staff': user.is_village_staff,
                    'is_staff': user.is_staff,
                    'is_superuser': user.is_superuser,
                    'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M'),
                    'last_login': user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Belum pernah login'
                }
                for user in page_obj
            ],
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat data pengguna: {str(e)}'
        }, status=500)


@login_required
def user_detail_api(request, user_id):
    """API for user detail"""
    try:
        user = get_object_or_404(CustomUser, id=user_id)
        
        data = {
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'phone_number': user.phone_number,
            'position': user.position,
            'is_active': user.is_active,
            'is_village_staff': user.is_village_staff,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
            'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M'),
            'last_login': user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else None,
            'created_at': user.created_at.strftime('%Y-%m-%d %H:%M'),
            'updated_at': user.updated_at.strftime('%Y-%m-%d %H:%M')
        }
        
        # Get profile if exists
        try:
            profile = user.userprofile
            data['profile'] = {
                'bio': profile.bio,
                'address': profile.address,
                'birth_date': profile.birth_date.strftime('%Y-%m-%d') if profile.birth_date else None,
                'avatar': profile.avatar.url if profile.avatar else None
            }
        except UserProfile.DoesNotExist:
            data['profile'] = None
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat detail pengguna: {str(e)}'
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def user_create_api(request):
    """API for creating user"""
    try:
        data = json.loads(request.body)
        
        user = CustomUser.objects.create_user(
            username=data['username'],
            email=data.get('email', ''),
            password=data['password'],
            first_name=data.get('first_name', ''),
            last_name=data.get('last_name', ''),
            phone_number=data.get('phone_number', ''),
            position=data.get('position', ''),
            is_village_staff=data.get('is_village_staff', False),
            is_staff=data.get('is_staff', False),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Pengguna berhasil ditambahkan',
            'data': {
                'id': user.id,
                'username': user.username,
                'email': user.email
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan pengguna: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def user_update_api(request, user_id):
    """API for updating user"""
    try:
        user = get_object_or_404(CustomUser, id=user_id)
        data = json.loads(request.body)
        
        user.first_name = data.get('first_name', user.first_name)
        user.last_name = data.get('last_name', user.last_name)
        user.email = data.get('email', user.email)
        user.phone_number = data.get('phone_number', user.phone_number)
        user.position = data.get('position', user.position)
        user.is_village_staff = data.get('is_village_staff', user.is_village_staff)
        user.is_staff = data.get('is_staff', user.is_staff)
        user.is_active = data.get('is_active', user.is_active)
        
        if 'password' in data and data['password']:
            user.set_password(data['password'])
        
        user.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Pengguna berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui pengguna: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def user_delete_api(request, user_id):
    """API for deleting user"""
    try:
        user = get_object_or_404(CustomUser, id=user_id)
        
        # Prevent deleting superuser or current user
        if user.is_superuser:
            return JsonResponse({
                'success': False,
                'message': 'Tidak dapat menghapus superuser'
            }, status=400)
            
        if user.id == request.user.id:
            return JsonResponse({
                'success': False,
                'message': 'Tidak dapat menghapus akun sendiri'
            }, status=400)
        
        user.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Pengguna berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus pengguna: {str(e)}'
        }, status=400)


# UMKM Business API Views
@login_required
def umkm_list_api(request):
    """API for UMKM businesses list"""
    try:
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        search = request.GET.get('search', '')
        business_type = request.GET.get('business_type', '')
        is_active = request.GET.get('is_active', '')
        
        queryset = UMKMBusiness.objects.all()
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(owner_name__icontains=search) |
                Q(business_type__icontains=search) |
                Q(description__icontains=search)
            )
        
        if business_type:
            queryset = queryset.filter(business_type__icontains=business_type)
            
        if is_active:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        queryset = queryset.order_by('-created_at')
        
        paginator = Paginator(queryset, per_page)
        page_obj = paginator.get_page(page)
        
        data = {
            'results': [
                {
                    'id': umkm.id,
                    'name': umkm.name,
                    'owner_name': umkm.owner_name,
                    'phone_number': umkm.phone_number,
                    'whatsapp_number': umkm.whatsapp_number,
                    'business_type': umkm.business_type,
                    'description': umkm.description[:100] + '...' if len(umkm.description) > 100 else umkm.description,
                    'address': umkm.address,
                    'is_active': umkm.is_active,
                    'whatsapp_link': umkm.whatsapp_link,
                    'created_at': umkm.created_at.strftime('%Y-%m-%d %H:%M')
                }
                for umkm in page_obj
            ],
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat data UMKM: {str(e)}'
        }, status=500)


@login_required
def umkm_detail_api(request, umkm_id):
    """API for UMKM business detail"""
    try:
        umkm = get_object_or_404(UMKMBusiness, id=umkm_id)
        
        data = {
            'id': umkm.id,
            'name': umkm.name,
            'owner_name': umkm.owner_name,
            'phone_number': umkm.phone_number,
            'whatsapp_number': umkm.whatsapp_number,
            'business_type': umkm.business_type,
            'description': umkm.description,
            'address': umkm.address,
            'is_active': umkm.is_active,
            'whatsapp_link': umkm.whatsapp_link,
            'created_at': umkm.created_at.strftime('%Y-%m-%d %H:%M'),
            'updated_at': umkm.updated_at.strftime('%Y-%m-%d %H:%M')
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat detail UMKM: {str(e)}'
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def umkm_create_api(request):
    """API for creating UMKM business"""
    try:
        data = json.loads(request.body)
        
        umkm = UMKMBusiness.objects.create(
            name=data['name'],
            owner_name=data['owner_name'],
            phone_number=data['phone_number'],
            whatsapp_number=data.get('whatsapp_number', ''),
            business_type=data['business_type'],
            description=data['description'],
            address=data['address'],
            is_active=data.get('is_active', True),
            whatsapp_link=data.get('whatsapp_link', '')
        )
        
        return JsonResponse({
            'success': True,
            'message': 'UMKM berhasil ditambahkan',
            'data': {
                'id': umkm.id,
                'name': umkm.name,
                'owner_name': umkm.owner_name
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan UMKM: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def umkm_update_api(request, umkm_id):
    """API for updating UMKM business"""
    try:
        umkm = get_object_or_404(UMKMBusiness, id=umkm_id)
        data = json.loads(request.body)
        
        umkm.name = data.get('name', umkm.name)
        umkm.owner_name = data.get('owner_name', umkm.owner_name)
        umkm.phone_number = data.get('phone_number', umkm.phone_number)
        umkm.whatsapp_number = data.get('whatsapp_number', umkm.whatsapp_number)
        umkm.business_type = data.get('business_type', umkm.business_type)
        umkm.description = data.get('description', umkm.description)
        umkm.address = data.get('address', umkm.address)
        umkm.is_active = data.get('is_active', umkm.is_active)
        umkm.whatsapp_link = data.get('whatsapp_link', umkm.whatsapp_link)
        umkm.save()
        
        return JsonResponse({
            'success': True,
            'message': 'UMKM berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui UMKM: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def umkm_delete_api(request, umkm_id):
    """API for deleting UMKM business"""
    try:
        umkm = get_object_or_404(UMKMBusiness, id=umkm_id)
        umkm.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'UMKM berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus UMKM: {str(e)}'
        }, status=400)


# System Settings API Views
@login_required
def system_settings_list_api(request):
    """API for system settings list"""
    try:
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        search = request.GET.get('search', '')
        is_active = request.GET.get('is_active', '')
        
        queryset = SystemSettings.objects.all()
        
        if search:
            queryset = queryset.filter(
                Q(setting_key__icontains=search) |
                Q(setting_value__icontains=search) |
                Q(description__icontains=search)
            )
            
        if is_active:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        queryset = queryset.order_by('setting_key')
        
        paginator = Paginator(queryset, per_page)
        page_obj = paginator.get_page(page)
        
        data = {
            'results': [
                {
                    'id': setting.id,
                    'setting_key': setting.setting_key,
                    'setting_value': setting.setting_value[:100] + '...' if len(setting.setting_value) > 100 else setting.setting_value,
                    'description': setting.description,
                    'is_active': setting.is_active,
                    'created_at': setting.created_at.strftime('%Y-%m-%d %H:%M'),
                    'updated_at': setting.updated_at.strftime('%Y-%m-%d %H:%M')
                }
                for setting in page_obj
            ],
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat pengaturan sistem: {str(e)}'
        }, status=500)


@login_required
def system_setting_detail_api(request, setting_id):
    """API for system setting detail"""
    try:
        setting = get_object_or_404(SystemSettings, id=setting_id)
        
        data = {
            'id': setting.id,
            'setting_key': setting.setting_key,
            'setting_value': setting.setting_value,
            'description': setting.description,
            'is_active': setting.is_active,
            'created_at': setting.created_at.strftime('%Y-%m-%d %H:%M'),
            'updated_at': setting.updated_at.strftime('%Y-%m-%d %H:%M')
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat detail pengaturan: {str(e)}'
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def system_setting_create_api(request):
    """API for creating system setting"""
    try:
        data = json.loads(request.body)
        
        setting = SystemSettings.objects.create(
            setting_key=data['setting_key'],
            setting_value=data['setting_value'],
            description=data.get('description', ''),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Pengaturan berhasil ditambahkan',
            'data': {
                'id': setting.id,
                'setting_key': setting.setting_key
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan pengaturan: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def system_setting_update_api(request, setting_id):
    """API for updating system setting"""
    try:
        setting = get_object_or_404(SystemSettings, id=setting_id)
        data = json.loads(request.body)
        
        setting.setting_key = data.get('setting_key', setting.setting_key)
        setting.setting_value = data.get('setting_value', setting.setting_value)
        setting.description = data.get('description', setting.description)
        setting.is_active = data.get('is_active', setting.is_active)
        setting.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Pengaturan berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui pengaturan: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def system_setting_delete_api(request, setting_id):
    """API for deleting system setting"""
    try:
        setting = get_object_or_404(SystemSettings, id=setting_id)
        setting.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Pengaturan berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus pengaturan: {str(e)}'
        }, status=400)


# WhatsApp Bot Config API Views
@login_required
def whatsapp_config_list_api(request):
    """API for WhatsApp bot configurations list"""
    try:
        configs = WhatsAppBotConfig.objects.all().order_by('-created_at')
        
        data = {
            'results': [
                {
                    'id': config.id,
                    'bot_name': config.bot_name,
                    'welcome_message': config.welcome_message[:100] + '...' if len(config.welcome_message) > 100 else config.welcome_message,
                    'business_hours_start': config.business_hours_start.strftime('%H:%M'),
                    'business_hours_end': config.business_hours_end.strftime('%H:%M'),
                    'auto_reply_enabled': config.auto_reply_enabled,
                    'is_active': config.is_active,
                    'created_at': config.created_at.strftime('%Y-%m-%d %H:%M')
                }
                for config in configs
            ]
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat konfigurasi WhatsApp: {str(e)}'
        }, status=500)


@login_required
def whatsapp_config_detail_api(request, config_id):
    """API for WhatsApp bot configuration detail"""
    try:
        config = get_object_or_404(WhatsAppBotConfig, id=config_id)
        
        data = {
            'id': config.id,
            'bot_name': config.bot_name,
            'welcome_message': config.welcome_message,
            'business_hours_start': config.business_hours_start.strftime('%H:%M'),
            'business_hours_end': config.business_hours_end.strftime('%H:%M'),
            'auto_reply_enabled': config.auto_reply_enabled,
            'is_active': config.is_active,
            'created_at': config.created_at.strftime('%Y-%m-%d %H:%M'),
            'updated_at': config.updated_at.strftime('%Y-%m-%d %H:%M')
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat detail konfigurasi: {str(e)}'
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def whatsapp_config_create_api(request):
    """API for creating WhatsApp bot configuration"""
    try:
        data = json.loads(request.body)
        
        config = WhatsAppBotConfig.objects.create(
            bot_name=data['bot_name'],
            welcome_message=data['welcome_message'],
            business_hours_start=data['business_hours_start'],
            business_hours_end=data['business_hours_end'],
            auto_reply_enabled=data.get('auto_reply_enabled', True),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Konfigurasi WhatsApp berhasil ditambahkan',
            'data': {
                'id': config.id,
                'bot_name': config.bot_name
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan konfigurasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def whatsapp_config_update_api(request, config_id):
    """API for updating WhatsApp bot configuration"""
    try:
        config = get_object_or_404(WhatsAppBotConfig, id=config_id)
        data = json.loads(request.body)
        
        config.bot_name = data.get('bot_name', config.bot_name)
        config.welcome_message = data.get('welcome_message', config.welcome_message)
        config.business_hours_start = data.get('business_hours_start', config.business_hours_start)
        config.business_hours_end = data.get('business_hours_end', config.business_hours_end)
        config.auto_reply_enabled = data.get('auto_reply_enabled', config.auto_reply_enabled)
        config.is_active = data.get('is_active', config.is_active)
        config.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Konfigurasi WhatsApp berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui konfigurasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def whatsapp_config_delete_api(request, config_id):
    """API for deleting WhatsApp bot configuration"""
    try:
        config = get_object_or_404(WhatsAppBotConfig, id=config_id)
        config.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Konfigurasi WhatsApp berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus konfigurasi: {str(e)}'
        }, status=400)


# Helper API Views
@login_required
def business_types_dropdown_api(request):
    """API for business types dropdown"""
    try:
        business_types = UMKMBusiness.objects.values_list('business_type', flat=True).distinct().order_by('business_type')
        
        data = {
            'business_types': [
                {'value': bt, 'label': bt}
                for bt in business_types if bt
            ]
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat jenis bisnis: {str(e)}'
        }, status=500)


# Website Settings API Views
@login_required
def website_settings_api(request):
    """API for website settings (GET/POST)"""
    if request.method == 'GET':
        try:
            settings, created = WebsiteSettings.objects.get_or_create(id=1)
            
            data = {
                'id': settings.id,
                'site_name': settings.site_name,
                'site_description': settings.site_description,
                'site_logo': settings.site_logo.url if settings.site_logo else None,
                'site_favicon': settings.site_favicon.url if settings.site_favicon else None,
                'contact_email': settings.contact_email,
                'contact_phone': settings.contact_phone,
                'contact_address': settings.contact_address,
                'facebook_url': settings.facebook_url,
                'instagram_url': settings.instagram_url,
                'twitter_url': settings.twitter_url,
                'youtube_url': settings.youtube_url,
                'theme': settings.theme,
                'primary_color': settings.primary_color,
                'secondary_color': settings.secondary_color,
                'default_language': settings.default_language,
                'timezone': settings.timezone,
                'meta_keywords': settings.meta_keywords,
                'meta_description': settings.meta_description,
                'google_analytics_id': settings.google_analytics_id,
                'maintenance_mode': settings.maintenance_mode,
                'maintenance_message': settings.maintenance_message,
                'allow_registration': settings.allow_registration,
                'max_file_upload_size': settings.max_file_upload_size,
                'email_notifications': settings.email_notifications,
                'sms_notifications': settings.sms_notifications,
                'whatsapp_notifications': settings.whatsapp_notifications,
                'updated_at': settings.updated_at.strftime('%Y-%m-%d %H:%M')
            }
            
            return JsonResponse({
                'success': True,
                'data': data
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Gagal memuat pengaturan website: {str(e)}'
            }, status=500)
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            settings, created = WebsiteSettings.objects.get_or_create(id=1)
            
            # Update fields
            for field, value in data.items():
                if hasattr(settings, field):
                    setattr(settings, field, value)
            
            settings.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Pengaturan website berhasil disimpan'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Gagal menyimpan pengaturan: {str(e)}'
            }, status=400)


# Module Settings API Views
@login_required
def module_settings_list_api(request):
    """API for module settings list"""
    try:
        modules = ModuleSettings.objects.all().order_by('menu_order', 'display_name')
        
        data = {
            'results': [
                {
                    'id': module.id,
                    'module_name': module.module_name,
                    'display_name': module.display_name,
                    'description': module.description,
                    'is_active': module.is_active,
                    'is_visible_in_menu': module.is_visible_in_menu,
                    'menu_order': module.menu_order,
                    'icon_class': module.icon_class,
                    'requires_permission': module.requires_permission,
                    'required_permission': module.required_permission,
                    'api_enabled': module.api_enabled,
                    'api_rate_limit': module.api_rate_limit,
                    'api_endpoints_count': module.api_endpoints.count(),
                    'created_at': module.created_at.strftime('%Y-%m-%d %H:%M')
                }
                for module in modules
            ]
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat pengaturan modul: {str(e)}'
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def module_settings_create_api(request):
    """API for creating module settings"""
    try:
        data = json.loads(request.body)
        
        module = ModuleSettings.objects.create(
            module_name=data['module_name'],
            display_name=data['display_name'],
            description=data.get('description', ''),
            is_active=data.get('is_active', True),
            is_visible_in_menu=data.get('is_visible_in_menu', True),
            menu_order=data.get('menu_order', 0),
            icon_class=data.get('icon_class', 'fas fa-cog'),
            requires_permission=data.get('requires_permission', False),
            required_permission=data.get('required_permission', ''),
            api_enabled=data.get('api_enabled', True),
            api_rate_limit=data.get('api_rate_limit', 100)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Pengaturan modul berhasil ditambahkan',
            'data': {
                'id': module.id,
                'module_name': module.module_name,
                'display_name': module.display_name
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan pengaturan modul: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def module_settings_update_api(request, module_id):
    """API for updating module settings"""
    try:
        module = get_object_or_404(ModuleSettings, id=module_id)
        data = json.loads(request.body)
        
        # Update fields
        for field, value in data.items():
            if hasattr(module, field):
                setattr(module, field, value)
        
        module.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Pengaturan modul berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui pengaturan modul: {str(e)}'
        }, status=400)


# API Endpoints Registry
@login_required
def api_endpoints_list_api(request):
    """API for listing all registered API endpoints"""
    try:
        module_name = request.GET.get('module', '')
        
        queryset = APIEndpoint.objects.select_related('module')
        
        if module_name:
            queryset = queryset.filter(module__module_name=module_name)
        
        endpoints = queryset.order_by('module__module_name', 'url_pattern')
        
        data = {
            'results': [
                {
                    'id': endpoint.id,
                    'module_name': endpoint.module.module_name,
                    'module_display_name': endpoint.module.display_name,
                    'name': endpoint.name,
                    'url_pattern': endpoint.url_pattern,
                    'method': endpoint.method,
                    'description': endpoint.description,
                    'is_active': endpoint.is_active,
                    'requires_auth': endpoint.requires_auth,
                    'created_at': endpoint.created_at.strftime('%Y-%m-%d %H:%M')
                }
                for endpoint in endpoints
            ]
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat daftar API endpoints: {str(e)}'
        }, status=500)


# Dashboard View
@login_required
def dashboard_view(request):
    """Main dashboard view"""
    return render(request, 'admin/modules/core/dashboard.html')


# Settings View
@login_required
def settings_view(request):
    """Website settings view"""
    return render(request, 'admin/modules/core/settings.html')


@login_required
def modules_view(request):
    """View untuk halaman manajemen modul"""
    return render(request, 'admin/modules/core/modules.html')


@login_required
def modules_list_api(request):
    """API untuk mendapatkan daftar semua modul"""
    try:
        modules = ModuleSettings.objects.all().order_by('menu_order', 'display_name')
        
        # Hitung jumlah API endpoints untuk setiap modul
        modules_data = []
        for module in modules:
            api_endpoints_count = APIEndpoint.objects.filter(module=module).count()
            module_data = {
                'id': module.id,
                'module_name': module.module_name,
                'display_name': module.display_name,
                'description': module.description,
                'icon_class': module.icon_class,
                'menu_order': module.menu_order,
                'is_active': module.is_active,
                'is_visible_in_menu': module.is_visible_in_menu,
                'requires_permission': module.requires_permission,
                'required_permission': module.required_permission,
                'api_enabled': module.api_enabled,
                'api_rate_limit': module.api_rate_limit,
                'api_endpoints_count': api_endpoints_count,
                'created_at': module.created_at.strftime('%Y-%m-%d %H:%M'),
                'updated_at': module.updated_at.strftime('%Y-%m-%d %H:%M')
            }
            modules_data.append(module_data)
        
        return JsonResponse({
            'success': True,
            'results': modules_data,
            'count': len(modules_data)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat data modul: {str(e)}'
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def modules_create_api(request):
    """API untuk membuat modul baru"""
    try:
        data = json.loads(request.body)
        
        # Validasi data required
        required_fields = ['module_name', 'display_name']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({
                    'success': False,
                    'message': f'Field {field} wajib diisi'
                }, status=400)
        
        # Cek apakah module_name sudah ada
        if ModuleSettings.objects.filter(module_name=data['module_name']).exists():
            return JsonResponse({
                'success': False,
                'message': 'Nama modul sudah digunakan'
            }, status=400)
        
        # Buat modul baru
        module = ModuleSettings.objects.create(
            module_name=data['module_name'],
            display_name=data['display_name'],
            description=data.get('description', ''),
            icon_class=data.get('icon_class', 'fas fa-cog'),
            menu_order=data.get('menu_order', 0),
            is_active=data.get('is_active', True),
            is_visible_in_menu=data.get('is_visible_in_menu', True),
            requires_permission=data.get('requires_permission', False),
            required_permission=data.get('required_permission', ''),
            api_enabled=data.get('api_enabled', True),
            api_rate_limit=data.get('api_rate_limit', 100)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Modul berhasil dibuat',
            'data': {
                'id': module.id,
                'module_name': module.module_name,
                'display_name': module.display_name
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal membuat modul: {str(e)}'
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def modules_update_api(request, module_id):
    """API untuk mengupdate modul"""
    try:
        module = get_object_or_404(ModuleSettings, id=module_id)
        data = json.loads(request.body)
        
        # Update fields yang diberikan
        if 'display_name' in data:
            module.display_name = data['display_name']
        if 'description' in data:
            module.description = data['description']
        if 'icon_class' in data:
            module.icon_class = data['icon_class']
        if 'menu_order' in data:
            module.menu_order = data['menu_order']
        if 'is_active' in data:
            module.is_active = data['is_active']
        if 'is_visible_in_menu' in data:
            module.is_visible_in_menu = data['is_visible_in_menu']
        if 'requires_permission' in data:
            module.requires_permission = data['requires_permission']
        if 'required_permission' in data:
            module.required_permission = data['required_permission']
        if 'api_enabled' in data:
            module.api_enabled = data['api_enabled']
        if 'api_rate_limit' in data:
            module.api_rate_limit = data['api_rate_limit']
        
        module.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Modul berhasil diupdate'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal mengupdate modul: {str(e)}'
        }, status=500)


@login_required
def reports_view(request):
    """View untuk halaman laporan gabungan"""
    return render(request, 'admin/modules/core/reports.html')


@csrf_exempt
@require_http_methods(["GET"])
def export_report_api(request):
    """API untuk export laporan dalam format PDF atau Excel"""
    try:
        format_type = request.GET.get('format', 'pdf')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        module_filter = request.GET.get('module')
        
        # Get aggregated data
        from django.utils import timezone
        from datetime import datetime, timedelta
        import io
        
        # Set default date range if not provided
        if not date_from:
            date_from = (timezone.now() - timedelta(days=30)).date()
        else:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            
        if not date_to:
            date_to = timezone.now().date()
        else:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        
        # Collect data from all modules
        report_data = {
            'date_from': date_from,
            'date_to': date_to,
            'generated_at': timezone.now(),
            'statistics': {},
            'module_data': {}
        }
        
        # Get basic statistics
        try:
            from references.models import Population
            report_data['statistics']['total_population'] = Population.objects.count()
        except:
            report_data['statistics']['total_population'] = 0
            
        try:
            from events.models import Event
            report_data['statistics']['total_events'] = Event.objects.filter(
                created_at__date__range=[date_from, date_to]
            ).count()
        except:
            report_data['statistics']['total_events'] = 0
            
        try:
            from documents.models import Document
            report_data['statistics']['total_documents'] = Document.objects.filter(
                created_at__date__range=[date_from, date_to]
            ).count()
        except:
            report_data['statistics']['total_documents'] = 0
            
        try:
            from business.models import UMKMBusiness
            report_data['statistics']['total_businesses'] = UMKMBusiness.objects.count()
        except:
            report_data['statistics']['total_businesses'] = 0
        
        if format_type == 'pdf':
            return generate_pdf_report(report_data)
        elif format_type == 'excel':
            return generate_excel_report(report_data)
        else:
            return JsonResponse({'success': False, 'message': 'Format tidak didukung'}, status=400)
            
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


def generate_pdf_report(data):
    """Generate PDF report"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from django.http import HttpResponse
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph('LAPORAN GABUNGAN SISTEM DESA', title_style))
        story.append(Spacer(1, 12))
        
        # Report info
        info_data = [
            ['Periode:', f"{data['date_from']} s/d {data['date_to']}"],
            ['Tanggal Generate:', data['generated_at'].strftime('%d/%m/%Y %H:%M:%S')]
        ]
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Statistics table
        stats_data = [
            ['Statistik', 'Jumlah'],
            ['Total Penduduk', str(data['statistics'].get('total_population', 0))],
            ['Total Events', str(data['statistics'].get('total_events', 0))],
            ['Total Dokumen', str(data['statistics'].get('total_documents', 0))],
            ['Total UMKM', str(data['statistics'].get('total_businesses', 0))]
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(Paragraph('RINGKASAN STATISTIK', styles['Heading2']))
        story.append(Spacer(1, 12))
        story.append(stats_table)
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="laporan-gabungan-{data["date_from"]}.pdf"'
        return response
        
    except ImportError:
        # Fallback if reportlab is not installed
        return JsonResponse({
            'success': False, 
            'message': 'PDF generation tidak tersedia. Install reportlab untuk menggunakan fitur ini.'
        }, status=500)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error generating PDF: {str(e)}'}, status=500)


def generate_excel_report(data):
    """Generate Excel report"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laporan Gabungan"
        
        # Title
        ws['A1'] = 'LAPORAN GABUNGAN SISTEM DESA'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Report info
        ws['A3'] = 'Periode:'
        ws['B3'] = f"{data['date_from']} s/d {data['date_to']}"
        ws['A4'] = 'Tanggal Generate:'
        ws['B4'] = data['generated_at'].strftime('%d/%m/%Y %H:%M:%S')
        
        # Statistics header
        ws['A6'] = 'RINGKASAN STATISTIK'
        ws['A6'].font = Font(size=14, bold=True)
        
        # Statistics data
        stats_headers = ['Statistik', 'Jumlah']
        for col, header in enumerate(stats_headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        stats_data = [
            ['Total Penduduk', data['statistics'].get('total_population', 0)],
            ['Total Events', data['statistics'].get('total_events', 0)],
            ['Total Dokumen', data['statistics'].get('total_documents', 0)],
            ['Total UMKM', data['statistics'].get('total_businesses', 0)]
        ]
        
        for row, (stat, value) in enumerate(stats_data, 8):
            ws.cell(row=row, column=1, value=stat)
            ws.cell(row=row, column=2, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="laporan-gabungan-{data["date_from"]}.xlsx"'
        return response
        
    except ImportError:
        # Fallback if openpyxl is not installed
        return JsonResponse({
            'success': False, 
            'message': 'Excel generation tidak tersedia. Install openpyxl untuk menggunakan fitur ini.'
        }, status=500)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error generating Excel: {str(e)}'}, status=500)


# Dashboard Aggregator API
@login_required
def dashboard_aggregator_api(request):
    """API for aggregating data from all modules for main dashboard"""
    try:
        # Import models from all modules
        from references.models import Penduduk, Dusun, Lorong
        from events.models import Event
        from documents.models import Document, DocumentRequest
        from business.models import Business
        from posyandu.models import PosyanduLocation
        from letters.models import Letter
        from beneficiaries.models import Beneficiary
        
        # Aggregate statistics
        stats = {
            'total_population': Penduduk.objects.count(),
            'total_dusun': Dusun.objects.count(),
            'total_lorong': Lorong.objects.count(),
            'total_events': Event.objects.count(),
            'active_events': Event.objects.filter(is_active=True).count(),
            'total_documents': Document.objects.count(),
            'pending_document_requests': DocumentRequest.objects.filter(status='pending').count(),
            'total_businesses': Business.objects.count(),
            'active_businesses': Business.objects.filter(is_active=True).count(),
            'total_posyandu': PosyanduLocation.objects.count(),
            'total_letters': Letter.objects.count(),
            'total_beneficiaries': Beneficiary.objects.count(),
            'total_users': CustomUser.objects.count(),
            'active_users': CustomUser.objects.filter(is_active=True).count(),
            'total_umkm': UMKMBusiness.objects.count(),
            'active_modules': ModuleSettings.objects.filter(is_active=True).count(),
            'total_api_endpoints': APIEndpoint.objects.filter(is_active=True).count()
        }
        
        # Recent activities (last 7 days)
        from datetime import datetime, timedelta
        week_ago = datetime.now() - timedelta(days=7)
        
        recent_stats = {
            'new_users_this_week': CustomUser.objects.filter(date_joined__gte=week_ago).count(),
            'new_documents_this_week': Document.objects.filter(created_at__gte=week_ago).count(),
            'new_letters_this_week': Letter.objects.filter(created_at__gte=week_ago).count(),
            'new_events_this_week': Event.objects.filter(created_at__gte=week_ago).count()
        }
        
        # Module status
        modules_status = [
            {
                'module_name': module.module_name,
                'display_name': module.display_name,
                'is_active': module.is_active,
                'api_enabled': module.api_enabled,
                'endpoints_count': module.api_endpoints.filter(is_active=True).count()
            }
            for module in ModuleSettings.objects.all().order_by('menu_order')
        ]
        
        return JsonResponse({
            'success': True,
            'data': {
                'statistics': stats,
                'recent_activities': recent_stats,
                'modules_status': modules_status,
                'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat data dashboard: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def api_aggregator(request):
    """API aggregator untuk mengumpulkan data dari semua modul dengan filter dan pagination"""
    try:
        module = request.GET.get('module', 'all')
        limit = int(request.GET.get('limit', 10))
        offset = int(request.GET.get('offset', 0))
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        result = {'success': True, 'data': {}}
        
        # Filter tanggal jika disediakan
        date_filter = {}
        if date_from:
            date_filter['created_at__gte'] = date_from
        if date_to:
            date_filter['created_at__lte'] = date_to
        
        if module == 'all' or module == 'population':
            from references.models import Penduduk
            population_qs = Penduduk.objects.all()
            if date_filter:
                population_qs = population_qs.filter(**date_filter)
            
            result['data']['population'] = {
                'total': population_qs.count(),
                'items': list(population_qs[offset:offset+limit].values(
                    'id', 'nama_lengkap', 'nik', 'jenis_kelamin', 'umur', 'status'
                ))
            }
        
        if module == 'all' or module == 'events':
            from events.models import Event
            events_qs = Event.objects.all()
            if date_from or date_to:
                if date_from:
                    events_qs = events_qs.filter(created_at__gte=date_from)
                if date_to:
                    events_qs = events_qs.filter(created_at__lte=date_to)
            
            result['data']['events'] = {
                'total': events_qs.count(),
                'items': list(events_qs[offset:offset+limit].values(
                    'id', 'nama_event', 'tanggal_mulai', 'tanggal_selesai', 'lokasi', 'is_active'
                ))
            }
        
        if module == 'all' or module == 'documents':
            from documents.models import Document
            documents_qs = Document.objects.all()
            if date_filter:
                documents_qs = documents_qs.filter(**date_filter)
            
            result['data']['documents'] = {
                'total': documents_qs.count(),
                'items': list(documents_qs[offset:offset+limit].values(
                    'id', 'nama_dokumen', 'jenis_dokumen', 'created_at', 'is_active'
                ))
            }
        
        if module == 'all' or module == 'business':
            from business.models import Business
            business_qs = Business.objects.all()
            if date_filter:
                business_qs = business_qs.filter(**date_filter)
            
            result['data']['business'] = {
                'total': business_qs.count(),
                'items': list(business_qs[offset:offset+limit].values(
                    'id', 'nama_usaha', 'pemilik', 'jenis_usaha', 'is_active'
                ))
            }
        
        if module == 'all' or module == 'letters':
            from letters.models import Letter
            letters_qs = Letter.objects.all()
            if date_filter:
                letters_qs = letters_qs.filter(**date_filter)
            
            result['data']['letters'] = {
                'total': letters_qs.count(),
                'items': list(letters_qs[offset:offset+limit].values(
                    'id', 'jenis_surat', 'pemohon', 'status', 'created_at'
                ))
            }
        
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Avg
from django.utils.dateparse import parse_date
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.middleware.csrf import get_token
import json
from datetime import datetime, date
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import (BusinessCategory, Business, BusinessOwner, BusinessProduct, 
                    BusinessFinance, Koperasi, BUMG, UKM, Aset, LayananJasa, JenisKoperasi)
from references.models import Penduduk


@login_required
def get_csrf_token(request):
    """Get CSRF token for AJAX requests"""
    return JsonResponse({'csrf_token': get_token(request)})


@login_required
def business_module_view(request):
    """Main business module view"""
    context = {
        'page_title': 'Bidang Usaha',
        'page_subtitle': 'Kelola data bidang usaha, UMKM, dan keuangan desa'
    }
    return render(request, 'admin/modules/business/index.html', context)


@login_required
def koperasi_view(request):
    """Koperasi Merah Putih view"""
    context = {
        'page_title': 'Koperasi Merah Putih',
        'page_subtitle': 'Kelola data koperasi merah putih'
    }
    return render(request, 'admin/modules/business/koperasi.html', context)


@login_required
def bumg_view(request):
    """BUMG view"""
    context = {
        'page_title': 'BUMG',
        'page_subtitle': 'Kelola data Badan Usaha Milik Gampong'
    }
    return render(request, 'admin/modules/business/bumg.html', context)


@login_required
def ukm_view(request):
    """Input Data UKM view"""
    context = {
        'page_title': 'Input Data UKM',
        'page_subtitle': 'Kelola data Usaha Kecil Menengah'
    }
    return render(request, 'admin/modules/business/ukm.html', context)


@login_required
def aset_view(request):
    """Input Data Aset view"""
    context = {
        'page_title': 'Input Data Aset',
        'page_subtitle': 'Kelola data aset bisnis dan usaha'
    }
    return render(request, 'admin/modules/business/aset.html', context)


@login_required
def jasa_view(request):
    """Layanan Jasa view"""
    context = {
        'page_title': 'Layanan Jasa',
        'page_subtitle': 'Kelola data layanan jasa'
    }
    return render(request, 'admin/modules/business/jasa.html', context)


# CRUD Operations for Koperasi
@login_required
@csrf_protect
def koperasi_api(request):
    if request.method == 'GET':
        koperasi_list = Koperasi.objects.all().order_by('-created_at')
        
        # Search functionality
        search = request.GET.get('search', '')
        if search:
            koperasi_list = koperasi_list.filter(
                Q(nama__icontains=search) |
                Q(ketua__icontains=search) |
                Q(nomor_badan_hukum__icontains=search)
            )
        
        # Filter by status
        status_filter = request.GET.get('status', '')
        if status_filter:
            koperasi_list = koperasi_list.filter(status=status_filter)
        
        # Pagination
        paginator = Paginator(koperasi_list, 10)
        page = request.GET.get('page', 1)
        koperasi_page = paginator.get_page(page)
        
        data = []
        for koperasi in koperasi_page:
            data.append({
                'id': koperasi.id,
                'nama': koperasi.nama,
                'nomor_badan_hukum': koperasi.nomor_badan_hukum,
                'tanggal_berdiri': koperasi.tanggal_berdiri.strftime('%Y-%m-%d'),
                'alamat': koperasi.alamat,
                'ketua': koperasi.ketua,
                'sekretaris': koperasi.sekretaris,
                'bendahara': koperasi.bendahara,
                'jumlah_anggota': koperasi.jumlah_anggota,
                'modal_awal': float(koperasi.modal_awal),
                'modal_sekarang': float(koperasi.modal_sekarang),
                'jenis_koperasi': koperasi.jenis_koperasi.id if koperasi.jenis_koperasi else None,
                'jenis_koperasi_nama': koperasi.jenis_koperasi.nama if koperasi.jenis_koperasi else '',
                'jenis_usaha': koperasi.jenis_usaha,
                'telepon': koperasi.telepon,
                'email': koperasi.email,
                'status': koperasi.status,
                'keterangan': koperasi.keterangan,
            })
        
        return JsonResponse({
            'success': True,
            'data': data,
            'total': paginator.count,
            'count': paginator.count,
            'current_page': koperasi_page.number,
            'num_pages': paginator.num_pages,
            'has_next': koperasi_page.has_next(),
            'has_previous': koperasi_page.has_previous(),
            'total_pages': paginator.num_pages,
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Handle jenis_koperasi
            jenis_koperasi = None
            if data.get('jenis_koperasi'):
                try:
                    jenis_koperasi = JenisKoperasi.objects.get(id=data['jenis_koperasi'])
                except JenisKoperasi.DoesNotExist:
                    pass
            
            koperasi = Koperasi.objects.create(
                nama=data['nama'],
                nomor_badan_hukum=data['nomor_badan_hukum'],
                tanggal_berdiri=parse_date(data['tanggal_berdiri']),
                alamat=data['alamat'],
                ketua=data['ketua'],
                sekretaris=data['sekretaris'],
                bendahara=data['bendahara'],
                jumlah_anggota=data.get('jumlah_anggota', 0),
                modal_awal=Decimal(str(data.get('modal_awal', 0))),
                modal_sekarang=Decimal(str(data.get('modal_sekarang', 0))),
                jenis_koperasi=jenis_koperasi,
                jenis_usaha=data['jenis_usaha'],
                telepon=data.get('telepon', ''),
                email=data.get('email', ''),
                status=data.get('status', 'aktif'),
                keterangan=data.get('keterangan', ''),
            )
            return JsonResponse({'success': True, 'id': koperasi.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            koperasi_id = data.get('id')
            koperasi = get_object_or_404(Koperasi, id=koperasi_id)
            
            # Handle jenis_koperasi
            jenis_koperasi = None
            if data.get('jenis_koperasi'):
                try:
                    jenis_koperasi = JenisKoperasi.objects.get(id=data['jenis_koperasi'])
                except JenisKoperasi.DoesNotExist:
                    pass
            
            koperasi.nama = data['nama']
            koperasi.nomor_badan_hukum = data['nomor_badan_hukum']
            koperasi.tanggal_berdiri = parse_date(data['tanggal_berdiri'])
            koperasi.alamat = data['alamat']
            koperasi.ketua = data['ketua']
            koperasi.sekretaris = data['sekretaris']
            koperasi.bendahara = data['bendahara']
            koperasi.jumlah_anggota = data.get('jumlah_anggota', 0)
            koperasi.modal_awal = Decimal(str(data.get('modal_awal', 0)))
            koperasi.modal_sekarang = Decimal(str(data.get('modal_sekarang', 0)))
            koperasi.jenis_koperasi = jenis_koperasi
            koperasi.jenis_usaha = data['jenis_usaha']
            koperasi.telepon = data.get('telepon', '')
            koperasi.email = data.get('email', '')
            koperasi.status = data.get('status', 'aktif')
            koperasi.keterangan = data.get('keterangan', '')
            koperasi.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e), 'message': str(e)})
    
    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            koperasi_id = data.get('id')
            koperasi = get_object_or_404(Koperasi, id=koperasi_id)
            koperasi.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


# CRUD Operations for BUMG
@login_required
@csrf_protect
def bumg_api(request):
    if request.method == 'GET':
        bumg_list = BUMG.objects.all().order_by('-created_at')
        
        # Search functionality
        search = request.GET.get('search', '')
        if search:
            bumg_list = bumg_list.filter(
                Q(nama__icontains=search) |
                Q(direktur__icontains=search) |
                Q(nomor_sk__icontains=search)
            )
        
        # Filter by status
        status_filter = request.GET.get('status', '')
        if status_filter:
            bumg_list = bumg_list.filter(status=status_filter)
        
        # Pagination
        paginator = Paginator(bumg_list, 10)
        page = request.GET.get('page', 1)
        bumg_page = paginator.get_page(page)
        
        data = []
        for bumg in bumg_page:
            data.append({
                'id': bumg.id,
                'nama': bumg.nama,
                'nomor_sk': bumg.nomor_sk,
                'tanggal_sk': bumg.tanggal_sk.strftime('%Y-%m-%d'),
                'alamat': bumg.alamat,
                'direktur': bumg.direktur,
                'komisaris': bumg.komisaris,
                'modal_dasar': float(bumg.modal_dasar),
                'modal_disetor': float(bumg.modal_disetor),
                'bidang_usaha': bumg.bidang_usaha,
                'telepon': bumg.telepon,
                'email': bumg.email,
                'website': bumg.website,
                'status': bumg.status,
                'keterangan': bumg.keterangan,
            })
        
        return JsonResponse({
            'results': data,
            'count': paginator.count,
            'page': bumg_page.number,
            'num_pages': paginator.num_pages,
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            bumg = BUMG.objects.create(
                nama=data['nama'],
                nomor_sk=data['nomor_sk'],
                tanggal_sk=parse_date(data['tanggal_sk']),
                alamat=data['alamat'],
                direktur=data['direktur'],
                komisaris=data['komisaris'],
                modal_dasar=Decimal(str(data.get('modal_dasar', 0))),
                modal_disetor=Decimal(str(data.get('modal_disetor', 0))),
                bidang_usaha=data['bidang_usaha'],
                telepon=data.get('telepon', ''),
                email=data.get('email', ''),
                website=data.get('website', ''),
                status=data.get('status', 'aktif'),
                keterangan=data.get('keterangan', ''),
            )
            return JsonResponse({'success': True, 'id': bumg.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            bumg_id = data.get('id')
            bumg = get_object_or_404(BUMG, id=bumg_id)
            
            bumg.nama = data['nama']
            bumg.nomor_sk = data['nomor_sk']
            bumg.tanggal_sk = parse_date(data['tanggal_sk'])
            bumg.alamat = data['alamat']
            bumg.direktur = data['direktur']
            bumg.komisaris = data['komisaris']
            bumg.modal_dasar = Decimal(str(data.get('modal_dasar', 0)))
            bumg.modal_disetor = Decimal(str(data.get('modal_disetor', 0)))
            bumg.bidang_usaha = data['bidang_usaha']
            bumg.telepon = data.get('telepon', '')
            bumg.email = data.get('email', '')
            bumg.website = data.get('website', '')
            bumg.status = data.get('status', 'aktif')
            bumg.keterangan = data.get('keterangan', '')
            bumg.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            bumg_id = data.get('id')
            bumg = get_object_or_404(BUMG, id=bumg_id)
            bumg.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


# CRUD Operations for UKM
@login_required
@csrf_protect
def ukm_api(request):
    if request.method == 'GET':
        ukm_list = UKM.objects.all().order_by('-created_at')
        
        # Search functionality
        search = request.GET.get('search', '')
        if search:
            ukm_list = ukm_list.filter(
                Q(nama_usaha__icontains=search) |
                Q(pemilik__icontains=search) |
                Q(jenis_usaha__icontains=search)
            )
        
        # Filter by status
        status_filter = request.GET.get('status', '')
        if status_filter:
            ukm_list = ukm_list.filter(status=status_filter)
        
        # Filter by skala
        skala_filter = request.GET.get('skala', '')
        if skala_filter:
            ukm_list = ukm_list.filter(skala_usaha=skala_filter)
        
        # Pagination
        paginator = Paginator(ukm_list, 10)
        page = request.GET.get('page', 1)
        ukm_page = paginator.get_page(page)
        
        data = []
        for ukm in ukm_page:
            data.append({
                'id': ukm.id,
                'nama_usaha': ukm.nama_usaha,
                'pemilik': ukm.pemilik,
                'nik_pemilik': ukm.nik_pemilik,
                'alamat_usaha': ukm.alamat_usaha,
                'alamat_pemilik': ukm.alamat_pemilik,
                'jenis_usaha': ukm.jenis_usaha,
                'skala_usaha': ukm.skala_usaha,
                'modal_awal': float(ukm.modal_awal),
                'omzet_bulanan': float(ukm.omzet_bulanan),
                'jumlah_karyawan': ukm.jumlah_karyawan,
                'tanggal_mulai': ukm.tanggal_mulai.strftime('%Y-%m-%d'),
                'nomor_izin': ukm.nomor_izin,
                'telepon': ukm.telepon,
                'email': ukm.email,
                'produk_utama': ukm.produk_utama,
                'target_pasar': ukm.target_pasar,
                'status': ukm.status,
                'keterangan': ukm.keterangan,
            })
        
        return JsonResponse({
            'success': True,
            'data': data,
            'total': paginator.count,
            'current_page': ukm_page.number,
            'total_pages': paginator.num_pages,
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            ukm = UKM.objects.create(
                nama_usaha=data['nama_usaha'],
                pemilik=data['pemilik'],
                nik_pemilik=data['nik_pemilik'],
                alamat_usaha=data['alamat_usaha'],
                alamat_pemilik=data['alamat_pemilik'],
                jenis_usaha=data['jenis_usaha'],
                skala_usaha=data['skala_usaha'],
                modal_awal=Decimal(str(data.get('modal_awal', 0))),
                omzet_bulanan=Decimal(str(data.get('omzet_bulanan', 0))),
                jumlah_karyawan=data.get('jumlah_karyawan', 1),
                tanggal_mulai=parse_date(data['tanggal_mulai']),
                nomor_izin=data.get('nomor_izin', ''),
                telepon=data['telepon'],
                email=data.get('email', ''),
                produk_utama=data['produk_utama'],
                target_pasar=data['target_pasar'],
                status=data.get('status', 'aktif'),
                keterangan=data.get('keterangan', ''),
            )
            return JsonResponse({'success': True, 'id': ukm.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            ukm_id = data.get('id')
            ukm = get_object_or_404(UKM, id=ukm_id)
            
            ukm.nama_usaha = data['nama_usaha']
            ukm.pemilik = data['pemilik']
            ukm.nik_pemilik = data['nik_pemilik']
            ukm.alamat_usaha = data['alamat_usaha']
            ukm.alamat_pemilik = data['alamat_pemilik']
            ukm.jenis_usaha = data['jenis_usaha']
            ukm.skala_usaha = data['skala_usaha']
            ukm.modal_awal = Decimal(str(data.get('modal_awal', 0)))
            ukm.omzet_bulanan = Decimal(str(data.get('omzet_bulanan', 0)))
            ukm.jumlah_karyawan = data.get('jumlah_karyawan', 1)
            ukm.tanggal_mulai = parse_date(data['tanggal_mulai'])
            ukm.nomor_izin = data.get('nomor_izin', '')
            ukm.telepon = data['telepon']
            ukm.email = data.get('email', '')
            ukm.produk_utama = data['produk_utama']
            ukm.target_pasar = data['target_pasar']
            ukm.status = data.get('status', 'aktif')
            ukm.keterangan = data.get('keterangan', '')
            ukm.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            ukm_id = data.get('id')
            ukm = get_object_or_404(UKM, id=ukm_id)
            ukm.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


# CRUD Operations for Aset
@login_required
@csrf_protect
def aset_api(request):
    if request.method == 'GET':
        aset_list = Aset.objects.all().order_by('-created_at')
        
        # Search functionality
        search = request.GET.get('search', '')
        if search:
            aset_list = aset_list.filter(
                Q(nama_aset__icontains=search) |
                Q(kode_aset__icontains=search) |
                Q(penanggung_jawab__icontains=search)
            )
        
        # Filter by kategori
        kategori_filter = request.GET.get('kategori', '')
        if kategori_filter:
            aset_list = aset_list.filter(kategori=kategori_filter)
        
        # Filter by kondisi
        kondisi_filter = request.GET.get('kondisi', '')
        if kondisi_filter:
            aset_list = aset_list.filter(kondisi=kondisi_filter)
        
        # Pagination
        paginator = Paginator(aset_list, 10)
        page = request.GET.get('page', 1)
        aset_page = paginator.get_page(page)
        
        data = []
        for aset in aset_page:
            data.append({
                'id': aset.id,
                'nama_aset': aset.nama_aset,
                'kategori': aset.kategori,
                'kode_aset': aset.kode_aset,
                'deskripsi': aset.deskripsi,
                'lokasi': aset.lokasi,
                'nilai_perolehan': float(aset.nilai_perolehan),
                'tanggal_perolehan': aset.tanggal_perolehan.strftime('%Y-%m-%d'),
                'kondisi': aset.kondisi,
                'masa_manfaat': aset.masa_manfaat,
                'penyusutan_per_tahun': float(aset.penyusutan_per_tahun),
                'nilai_buku': float(aset.nilai_buku),
                'penanggung_jawab': aset.penanggung_jawab,
                'nomor_sertifikat': aset.nomor_sertifikat,
                'keterangan': aset.keterangan,
            })
        
        return JsonResponse({
            'success': True,
            'data': data,
            'total': paginator.count,
            'current_page': aset_page.number,
            'total_pages': paginator.num_pages,
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            # Only use fields that exist in the Aset model
            aset = Aset.objects.create(
                nama_aset=data.get('nama_aset', ''),
                kategori=data.get('kategori', ''),
                kode_aset=data.get('kode_aset', ''),
                deskripsi=data.get('deskripsi', ''),
                lokasi=data.get('lokasi', ''),
                nilai_perolehan=Decimal(str(data.get('nilai_perolehan', 0))) if data.get('nilai_perolehan') else Decimal('0'),
                tanggal_perolehan=parse_date(data['tanggal_perolehan']) if data.get('tanggal_perolehan') else None,
                kondisi=data.get('kondisi', 'baik'),
                masa_manfaat=data.get('masa_manfaat', 0),
                penyusutan_per_tahun=Decimal(str(data.get('penyusutan_per_tahun', 0))),
                nilai_buku=Decimal(str(data.get('nilai_buku', 0))),
                penanggung_jawab=data.get('penanggung_jawab', ''),
                nomor_sertifikat=data.get('nomor_sertifikat', ''),
                keterangan=data.get('keterangan', ''),
                # Ignore non-existent fields: nilai_saat_ini, sumber_dana, merk, model, tahun_pembuatan, status_kepemilikan, tanggal_pemeliharaan, biaya_pemeliharaan
            )
            return JsonResponse({'success': True, 'id': aset.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e), 'message': str(e)})
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            aset_id = data.get('id')
            aset = get_object_or_404(Aset, id=aset_id)
            
            aset.nama_aset = data.get('nama_aset', '')
            aset.kategori = data.get('kategori', '')
            aset.kode_aset = data.get('kode_aset', '')
            aset.deskripsi = data.get('deskripsi', '')
            aset.lokasi = data.get('lokasi', '')
            aset.nilai_perolehan = Decimal(str(data.get('nilai_perolehan', 0))) if data.get('nilai_perolehan') else Decimal('0')
            aset.tanggal_perolehan = parse_date(data['tanggal_perolehan']) if data.get('tanggal_perolehan') else None
            aset.kondisi = data.get('kondisi', 'baik')
            aset.masa_manfaat = data.get('masa_manfaat', 0)
            aset.penyusutan_per_tahun = Decimal(str(data.get('penyusutan_per_tahun', 0)))
            aset.nilai_buku = Decimal(str(data.get('nilai_buku', 0)))
            aset.penanggung_jawab = data.get('penanggung_jawab', '')
            aset.nomor_sertifikat = data.get('nomor_sertifikat', '')
            aset.keterangan = data.get('keterangan', '')
            aset.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e), 'message': str(e)})
    
    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            aset_id = data.get('id')
            aset = get_object_or_404(Aset, id=aset_id)
            aset.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


@login_required
@csrf_protect
def aset_detail_api(request, aset_id):
    """API endpoint for individual Aset operations (PUT, DELETE)"""
    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
            aset = get_object_or_404(Aset, id=aset_id)
            
            aset.nama_aset = data.get('nama_aset', '')
            aset.kategori = data.get('kategori', '')
            aset.kode_aset = data.get('kode_aset', '')
            aset.deskripsi = data.get('deskripsi', '')
            aset.lokasi = data.get('lokasi', '')
            aset.nilai_perolehan = Decimal(str(data.get('nilai_perolehan', 0))) if data.get('nilai_perolehan') else Decimal('0')
            aset.tanggal_perolehan = parse_date(data['tanggal_perolehan']) if data.get('tanggal_perolehan') else None
            aset.kondisi = data.get('kondisi', 'baik')
            aset.masa_manfaat = data.get('masa_manfaat', 0)
            aset.penyusutan_per_tahun = Decimal(str(data.get('penyusutan_per_tahun', 0)))
            aset.nilai_buku = Decimal(str(data.get('nilai_buku', 0)))
            aset.penanggung_jawab = data.get('penanggung_jawab', '')
            aset.nomor_sertifikat = data.get('nomor_sertifikat', '')
            aset.keterangan = data.get('keterangan', '')
            
            # Note: Additional fields from form are ignored as they don't exist in the model
            # Fields like nilai_saat_ini, sumber_dana, merk, model, tahun_pembuatan, 
            # status_kepemilikan, tanggal_pemeliharaan, biaya_pemeliharaan are not in Aset model
            
            aset.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        try:
            aset = get_object_or_404(Aset, id=aset_id)
            aset.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# CRUD Operations for LayananJasa
@login_required
@csrf_protect
def jasa_api(request):
    if request.method == 'GET':
        jasa_list = LayananJasa.objects.all().order_by('-created_at')
        
        # Search functionality
        search = request.GET.get('search', '')
        if search:
            jasa_list = jasa_list.filter(
                Q(nama__icontains=search) |
                Q(penyedia__icontains=search) |
                Q(kategori__icontains=search)
            )
        
        # Filter by kategori
        kategori_filter = request.GET.get('kategori', '')
        if kategori_filter:
            jasa_list = jasa_list.filter(kategori=kategori_filter)
        
        # Filter by status
        status_filter = request.GET.get('status', '')
        if status_filter:
            jasa_list = jasa_list.filter(status=status_filter)
        
        # Pagination
        paginator = Paginator(jasa_list, 10)
        page = request.GET.get('page', 1)
        jasa_page = paginator.get_page(page)
        
        data = []
        for jasa in jasa_page:
            data.append({
                'id': jasa.id,
                'nama': jasa.nama,
                'kategori': jasa.kategori,
                'deskripsi': jasa.deskripsi,
                'penyedia': jasa.penyedia,
                'telepon': jasa.telepon,
                'email': jasa.email,
                'alamat': jasa.alamat,
                'pengalaman': jasa.pengalaman,
                'harga_min': float(jasa.harga_min) if jasa.harga_min else None,
                'harga_max': float(jasa.harga_max) if jasa.harga_max else None,
                'satuan_harga': jasa.satuan_harga,
                'waktu_layanan': jasa.waktu_layanan,
                'area_layanan': jasa.area_layanan,
                'status': jasa.status,
                'rating': float(jasa.rating) if jasa.rating else None,
                'website': jasa.website,
                'sertifikat': jasa.sertifikat,
                'keunggulan': jasa.keunggulan,
                'syarat_ketentuan': jasa.syarat_ketentuan,
            })
        
        return JsonResponse({
            'results': data,
            'count': paginator.count,
            'current_page': jasa_page.number,
            'num_pages': paginator.num_pages,
            'has_next': jasa_page.has_next(),
            'has_previous': jasa_page.has_previous(),
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            jasa = LayananJasa.objects.create(
                nama=data['nama'],
                kategori=data['kategori'],
                deskripsi=data['deskripsi'],
                penyedia=data['penyedia'],
                telepon=data['telepon'],
                email=data.get('email', ''),
                alamat=data['alamat'],
                pengalaman=data.get('pengalaman', 0),
                harga_min=Decimal(str(data['harga_min'])) if data.get('harga_min') else None,
                harga_max=Decimal(str(data['harga_max'])) if data.get('harga_max') else None,
                satuan_harga=data.get('satuan_harga', 'per_proyek'),
                waktu_layanan=data.get('waktu_layanan', ''),
                area_layanan=data.get('area_layanan', ''),
                status=data.get('status', 'aktif'),
                rating=Decimal(str(data['rating'])) if data.get('rating') else None,
                website=data.get('website', ''),
                sertifikat=data.get('sertifikat', ''),
                keunggulan=data.get('keunggulan', ''),
                syarat_ketentuan=data.get('syarat_ketentuan', ''),
            )
            return JsonResponse({'success': True, 'id': jasa.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            jasa_id = data.get('id')
            jasa = get_object_or_404(LayananJasa, id=jasa_id)
            
            jasa.nama = data['nama']
            jasa.kategori = data['kategori']
            jasa.deskripsi = data['deskripsi']
            jasa.penyedia = data['penyedia']
            jasa.telepon = data['telepon']
            jasa.email = data.get('email', '')
            jasa.alamat = data['alamat']
            jasa.pengalaman = data.get('pengalaman', 0)
            jasa.harga_min = Decimal(str(data['harga_min'])) if data.get('harga_min') else None
            jasa.harga_max = Decimal(str(data['harga_max'])) if data.get('harga_max') else None
            jasa.satuan_harga = data.get('satuan_harga', 'per_proyek')
            jasa.waktu_layanan = data.get('waktu_layanan', '')
            jasa.area_layanan = data.get('area_layanan', '')
            jasa.status = data.get('status', 'aktif')
            jasa.rating = Decimal(str(data['rating'])) if data.get('rating') else None
            jasa.website = data.get('website', '')
            jasa.sertifikat = data.get('sertifikat', '')
            jasa.keunggulan = data.get('keunggulan', '')
            jasa.syarat_ketentuan = data.get('syarat_ketentuan', '')
            jasa.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            jasa_id = data.get('id')
            jasa = get_object_or_404(LayananJasa, id=jasa_id)
            jasa.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


@login_required
@csrf_protect
def jasa_detail_api(request, jasa_id):
    """API for individual jasa operations (GET, PUT, DELETE)"""
    jasa = get_object_or_404(LayananJasa, id=jasa_id)
    
    if request.method == 'GET':
        data = {
            'id': jasa.id,
            'nama': jasa.nama,
            'kategori': jasa.kategori,
            'deskripsi': jasa.deskripsi,
            'penyedia': jasa.penyedia,
            'telepon': jasa.telepon,
            'email': jasa.email,
            'alamat': jasa.alamat,
            'pengalaman': jasa.pengalaman,
            'harga_min': float(jasa.harga_min) if jasa.harga_min else None,
            'harga_max': float(jasa.harga_max) if jasa.harga_max else None,
            'satuan_harga': jasa.satuan_harga,
            'waktu_layanan': jasa.waktu_layanan,
            'area_layanan': jasa.area_layanan,
            'status': jasa.status,
            'rating': float(jasa.rating) if jasa.rating else None,
            'website': jasa.website,
            'sertifikat': jasa.sertifikat,
            'keunggulan': jasa.keunggulan,
            'syarat_ketentuan': jasa.syarat_ketentuan,
        }
        return JsonResponse({'success': True, 'data': data})
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            
            jasa.nama = data['nama']
            jasa.kategori = data['kategori']
            jasa.deskripsi = data['deskripsi']
            jasa.penyedia = data['penyedia']
            jasa.telepon = data['telepon']
            jasa.email = data.get('email', '')
            jasa.alamat = data['alamat']
            jasa.pengalaman = data.get('pengalaman', 0)
            jasa.harga_min = Decimal(str(data['harga_min'])) if data.get('harga_min') else None
            jasa.harga_max = Decimal(str(data['harga_max'])) if data.get('harga_max') else None
            jasa.satuan_harga = data.get('satuan_harga', 'per_proyek')
            jasa.waktu_layanan = data.get('waktu_layanan', '')
            jasa.area_layanan = data.get('area_layanan', '')
            jasa.status = data.get('status', 'aktif')
            jasa.rating = Decimal(str(data['rating'])) if data.get('rating') else None
            jasa.website = data.get('website', '')
            jasa.sertifikat = data.get('sertifikat', '')
            jasa.keunggulan = data.get('keunggulan', '')
            jasa.syarat_ketentuan = data.get('syarat_ketentuan', '')
            jasa.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        try:
            jasa.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


# View functions for business sub-modules
@login_required
def business_module_view(request):
    """Main business module dashboard"""
    return render(request, 'admin/modules/business/index.html')

@login_required
def business_view(request):
    """Business management view with modern UI"""
    context = {
        'page_title': 'Manajemen Bisnis',
        'page_subtitle': 'Kelola data bisnis dan UMKM'
    }
    return render(request, 'admin/modules/business/business.html', context)

@login_required
def koperasi_view(request):
    """Koperasi Merah Putih management page"""
    return render(request, 'admin/modules/business/koperasi.html')

@login_required
def ukm_view(request):
    """UKM management page"""
    return render(request, 'admin/modules/business/ukm.html')

@login_required
def aset_view(request):
    """Asset management page"""
    return render(request, 'admin/modules/business/aset.html')

@login_required
def jasa_view(request):
    """Service management page"""
    return render(request, 'admin/modules/business/jasa.html')


# Business Category Views
@login_required
def business_category_list(request):
    """List business categories with pagination and search"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    is_active = request.GET.get('is_active', '')
    
    categories = BusinessCategory.objects.all()
    
    if search:
        categories = categories.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    
    if is_active:
        categories = categories.filter(is_active=is_active.lower() == 'true')
    
    categories = categories.order_by('name')
    
    paginator = Paginator(categories, per_page)
    page_obj = paginator.get_page(page)
    
    data = []
    for category in page_obj:
        data.append({
            'id': category.id,
            'name': category.name,
            'description': category.description,
            'is_active': category.is_active,
            'businesses_count': category.business_set.count(),
            'created_at': category.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    return JsonResponse({
        'results': data,
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })


@csrf_protect
@require_http_methods(["POST"])
@login_required
def business_category_create(request):
    """Create new business category"""
    try:
        data = json.loads(request.body)
        
        category = BusinessCategory.objects.create(
            name=data['name'],
            description=data.get('description', ''),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Kategori bisnis berhasil ditambahkan',
            'data': {
                'id': category.id,
                'name': category.name,
                'description': category.description,
                'is_active': category.is_active
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan kategori bisnis: {str(e)}'
        }, status=400)


@login_required
def business_category_detail(request, category_id):
    """Get business category detail"""
    category = get_object_or_404(BusinessCategory, id=category_id)
    
    return JsonResponse({
        'id': category.id,
        'name': category.name,
        'description': category.description,
        'is_active': category.is_active,
        'businesses_count': category.business_set.count(),
        'created_at': category.created_at.strftime('%Y-%m-%d %H:%M'),
        'updated_at': category.updated_at.strftime('%Y-%m-%d %H:%M')
    })


@csrf_protect
@require_http_methods(["PUT"])
@login_required
def business_category_update(request, category_id):
    """Update business category"""
    try:
        category = get_object_or_404(BusinessCategory, id=category_id)
        data = json.loads(request.body)
        
        category.name = data.get('name', category.name)
        category.description = data.get('description', category.description)
        category.is_active = data.get('is_active', category.is_active)
        category.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Kategori bisnis berhasil diperbarui'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui kategori bisnis: {str(e)}'
        }, status=400)


@csrf_protect
@require_http_methods(["DELETE"])
@login_required
def business_category_delete(request, category_id):
    """Delete business category"""
    try:
        category = get_object_or_404(BusinessCategory, id=category_id)
        
        if category.business_set.exists():
            return JsonResponse({
                'success': False,
                'message': 'Tidak dapat menghapus kategori yang masih memiliki bisnis'
            }, status=400)
        
        category.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Kategori bisnis berhasil dihapus'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus kategori bisnis: {str(e)}'
        }, status=400)


# Business Views
@login_required
def businesses_list(request):
    """List businesses with pagination and search"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    category_id = request.GET.get('category_id', '')
    business_type = request.GET.get('business_type', '')
    status = request.GET.get('status', '')
    
    businesses = Business.objects.select_related('category').all()
    
    if search:
        businesses = businesses.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(address__icontains=search)
        )
    
    if category_id:
        businesses = businesses.filter(category_id=category_id)
    
    if business_type:
        businesses = businesses.filter(business_type=business_type)
    
    if status:
        businesses = businesses.filter(status=status)
    
    businesses = businesses.order_by('-created_at')
    
    paginator = Paginator(businesses, per_page)
    page_obj = paginator.get_page(page)
    
    data = []
    for business in page_obj:
        # Get primary owner
        primary_owner = business.owners.filter(is_active=True).first()
        
        data.append({
            'id': business.id,
            'name': business.name,
            'category': business.category.name,
            'business_type': business.get_business_type_display(),
            'address': business.address,
            'phone': business.phone,
            'status': business.get_status_display(),
            'employee_count': business.employee_count,
            'owner_name': primary_owner.owner.nama if primary_owner else '-',
            'established_date': business.established_date.strftime('%Y-%m-%d') if business.established_date else '-',
            'created_at': business.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    return JsonResponse({
        'results': data,
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })


@csrf_protect
@require_http_methods(["POST"])
@login_required
def business_create(request):
    """Create new business"""
    try:
        data = json.loads(request.body)
        
        business = Business.objects.create(
            name=data['name'],
            category_id=data['category_id'],
            business_type=data['business_type'],
            description=data.get('description', ''),
            address=data['address'],
            phone=data.get('phone', ''),
            email=data.get('email', ''),
            website=data.get('website', ''),
            established_date=parse_date(data['established_date']) if data.get('established_date') else None,
            license_number=data.get('license_number', ''),
            status=data.get('status', 'aktif'),
            employee_count=data.get('employee_count', 1)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Bisnis berhasil ditambahkan',
            'data': {
                'id': business.id,
                'name': business.name
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan bisnis: {str(e)}'
        }, status=400)


@login_required
def business_detail(request, business_id):
    """Get business detail"""
    business = get_object_or_404(Business, id=business_id)
    
    # Get owners
    owners = []
    for owner in business.owners.filter(is_active=True):
        owners.append({
            'id': owner.id,
            'name': owner.owner.nama,
            'nik': owner.owner.nik,
            'ownership_percentage': float(owner.ownership_percentage),
            'role': owner.role
        })
    
    # Get products count
    products_count = business.products.filter(is_available=True).count()
    
    # Get financial summary
    finances = business.finances.all()
    total_income = finances.filter(transaction_type='income').aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense = finances.filter(transaction_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
    
    return JsonResponse({
        'id': business.id,
        'name': business.name,
        'category': business.category.name,
        'category_id': business.category.id,
        'business_type': business.business_type,
        'business_type_display': business.get_business_type_display(),
        'description': business.description,
        'address': business.address,
        'phone': business.phone,
        'email': business.email,
        'website': business.website,
        'established_date': business.established_date.strftime('%Y-%m-%d') if business.established_date else '',
        'license_number': business.license_number,
        'status': business.status,
        'status_display': business.get_status_display(),
        'employee_count': business.employee_count,
        'owners': owners,
        'products_count': products_count,
        'total_income': float(total_income),
        'total_expense': float(total_expense),
        'net_profit': float(total_income - total_expense),
        'created_at': business.created_at.strftime('%Y-%m-%d %H:%M'),
        'updated_at': business.updated_at.strftime('%Y-%m-%d %H:%M')
    })


@csrf_protect
@require_http_methods(["PUT"])
@login_required
def business_update(request, business_id):
    """Update business"""
    try:
        business = get_object_or_404(Business, id=business_id)
        data = json.loads(request.body)
        
        business.name = data.get('name', business.name)
        business.category_id = data.get('category_id', business.category_id)
        business.business_type = data.get('business_type', business.business_type)
        business.description = data.get('description', business.description)
        business.address = data.get('address', business.address)
        business.phone = data.get('phone', business.phone)
        business.email = data.get('email', business.email)
        business.website = data.get('website', business.website)
        
        if data.get('established_date'):
            business.established_date = parse_date(data['established_date'])
        
        business.license_number = data.get('license_number', business.license_number)
        business.status = data.get('status', business.status)
        business.employee_count = data.get('employee_count', business.employee_count)
        business.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Bisnis berhasil diperbarui'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui bisnis: {str(e)}'
        }, status=400)


@csrf_protect
@require_http_methods(["DELETE"])
@login_required
def business_delete(request, business_id):
    """Delete business"""
    try:
        business = get_object_or_404(Business, id=business_id)
        business.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Bisnis berhasil dihapus'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus bisnis: {str(e)}'
        }, status=400)


# Business Owner Views
@login_required
def business_owners_list(request):
    """List business owners with pagination and search"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    business_id = request.GET.get('business_id', '')
    is_active = request.GET.get('is_active', '')
    
    owners = BusinessOwner.objects.select_related('business', 'owner').all()
    
    if search:
        owners = owners.filter(
            Q(owner__nama__icontains=search) |
            Q(owner__nik__icontains=search) |
            Q(business__name__icontains=search) |
            Q(role__icontains=search)
        )
    
    if business_id:
        owners = owners.filter(business_id=business_id)
    
    if is_active:
        owners = owners.filter(is_active=is_active.lower() == 'true')
    
    owners = owners.order_by('-created_at')
    
    paginator = Paginator(owners, per_page)
    page_obj = paginator.get_page(page)
    
    data = []
    for owner in page_obj:
        data.append({
            'id': owner.id,
            'owner_name': owner.owner.nama,
            'owner_nik': owner.owner.nik,
            'business_name': owner.business.name,
            'ownership_percentage': float(owner.ownership_percentage),
            'role': owner.role,
            'start_date': owner.start_date.strftime('%Y-%m-%d'),
            'end_date': owner.end_date.strftime('%Y-%m-%d') if owner.end_date else '-',
            'is_active': owner.is_active,
            'created_at': owner.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    return JsonResponse({
        'results': data,
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })


# Business Product Views
@login_required
def business_products_list(request):
    """List business products with pagination and search"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    business_id = request.GET.get('business_id', '')
    is_available = request.GET.get('is_available', '')
    
    products = BusinessProduct.objects.select_related('business').all()
    
    if search:
        products = products.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(business__name__icontains=search)
        )
    
    if business_id:
        products = products.filter(business_id=business_id)
    
    if is_available:
        products = products.filter(is_available=is_available.lower() == 'true')
    
    products = products.order_by('-created_at')
    
    paginator = Paginator(products, per_page)
    page_obj = paginator.get_page(page)
    
    data = []
    for product in page_obj:
        data.append({
            'id': product.id,
            'name': product.name,
            'business_name': product.business.name,
            'description': product.description,
            'price': float(product.price),
            'unit': product.unit,
            'stock_quantity': product.stock_quantity,
            'min_stock': product.min_stock,
            'is_available': product.is_available,
            'stock_status': 'Stok Rendah' if product.stock_quantity <= product.min_stock else 'Normal',
            'created_at': product.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    return JsonResponse({
        'results': data,
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })


# Business Finance Views
@login_required
def business_finances_list(request):
    """List business finances with pagination and search"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    business_id = request.GET.get('business_id', '')
    transaction_type = request.GET.get('transaction_type', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    finances = BusinessFinance.objects.select_related('business', 'recorded_by').all()
    
    if search:
        finances = finances.filter(
            Q(description__icontains=search) |
            Q(business__name__icontains=search) |
            Q(category__icontains=search) |
            Q(receipt_number__icontains=search)
        )
    
    if business_id:
        finances = finances.filter(business_id=business_id)
    
    if transaction_type:
        finances = finances.filter(transaction_type=transaction_type)
    
    if start_date:
        finances = finances.filter(transaction_date__gte=parse_date(start_date))
    
    if end_date:
        finances = finances.filter(transaction_date__lte=parse_date(end_date))
    
    finances = finances.order_by('-transaction_date', '-created_at')
    
    paginator = Paginator(finances, per_page)
    page_obj = paginator.get_page(page)
    
    data = []
    for finance in page_obj:
        data.append({
            'id': finance.id,
            'business_name': finance.business.name,
            'transaction_type': finance.transaction_type,
            'transaction_type_display': finance.get_transaction_type_display(),
            'amount': float(finance.amount),
            'description': finance.description,
            'transaction_date': finance.transaction_date.strftime('%Y-%m-%d'),
            'category': finance.category,
            'receipt_number': finance.receipt_number,
            'recorded_by': finance.recorded_by.username if finance.recorded_by else '-',
            'created_at': finance.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    return JsonResponse({
        'results': data,
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })


# API Helper Views
@login_required
def business_categories_dropdown(request):
    """Get business categories for dropdown"""
    categories = BusinessCategory.objects.filter(is_active=True).order_by('name')
    
    data = []
    for category in categories:
        data.append({
            'id': category.id,
            'name': category.name
        })
    
    return JsonResponse({'categories': data})


@login_required
def businesses_dropdown(request):
    """Get businesses for dropdown"""
    businesses = Business.objects.filter(status='aktif').order_by('name')
    
    data = []
    for business in businesses:
        data.append({
            'id': business.id,
            'name': business.name,
            'business_type': business.get_business_type_display()
        })
    
    return JsonResponse({'businesses': data})


@login_required
def residents_dropdown(request):
    """Get residents for dropdown"""
    residents = Penduduk.objects.filter(status_kependudukan='tetap').order_by('nama')
    
    data = []
    for resident in residents:
        data.append({
            'id': resident.id,
            'name': resident.nama,
            'nik': resident.nik
        })
    
    return JsonResponse({'residents': data})


@login_required
def business_statistics(request):
    """Get business statistics"""
    total_businesses = Business.objects.count()
    active_businesses = Business.objects.filter(status='aktif').count()
    total_categories = BusinessCategory.objects.filter(is_active=True).count()
    total_products = BusinessProduct.objects.filter(is_available=True).count()
    
    # Business by type
    business_by_type = Business.objects.values('business_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Business by category
    business_by_category = Business.objects.values('category__name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Financial summary
    total_income = BusinessFinance.objects.filter(
        transaction_type='income'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    total_expense = BusinessFinance.objects.filter(
        transaction_type='expense'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    return JsonResponse({
        'total_businesses': total_businesses,
        'active_businesses': active_businesses,
        'total_categories': total_categories,
        'total_products': total_products,
        'business_by_type': list(business_by_type),
        'business_by_category': list(business_by_category),
        'total_income': float(total_income),
        'total_expense': float(total_expense),
        'net_profit': float(total_income - total_expense)
    })


# Export Functions
def create_excel_response(filename, workbook):
    """Helper function to create Excel HTTP response"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def style_header_row(worksheet, row_num, columns):
    """Style header row with blue background and white text"""
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col_num, _ in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column width
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 15


@login_required
def export_koperasi(request):
    """Export Koperasi data to Excel"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data Koperasi'
    
    # Headers
    headers = ['No', 'Nama Koperasi', 'Alamat', 'Telepon', 'Email', 'Jenis Koperasi', 
               'Jumlah Anggota', 'Modal Awal', 'Status', 'Tanggal Dibuat']
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    style_header_row(worksheet, 1, headers)
    
    # Write data
    koperasi_list = Koperasi.objects.all().order_by('nama')
    for row_num, koperasi in enumerate(koperasi_list, 2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1)
        worksheet.cell(row=row_num, column=2, value=koperasi.nama)
        worksheet.cell(row=row_num, column=3, value=koperasi.alamat)
        worksheet.cell(row=row_num, column=4, value=koperasi.telepon)
        worksheet.cell(row=row_num, column=5, value=koperasi.email)
        worksheet.cell(row=row_num, column=6, value=koperasi.jenis_usaha)
        worksheet.cell(row=row_num, column=7, value=koperasi.jumlah_anggota)
        worksheet.cell(row=row_num, column=8, value=float(koperasi.modal_awal))
        worksheet.cell(row=row_num, column=9, value=koperasi.status)
        worksheet.cell(row=row_num, column=10, value=koperasi.created_at.strftime('%Y-%m-%d'))
    
    filename = f'Data_Koperasi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(filename, workbook)


@login_required
def export_bumg(request):
    """Export BUMG data to Excel"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data BUMG'
    
    # Headers
    headers = ['No', 'Nama BUMG', 'Alamat', 'Telepon', 'Email', 'Jenis Usaha', 
               'Modal Awal', 'Jumlah Karyawan', 'Status', 'Tanggal Dibuat']
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    style_header_row(worksheet, 1, headers)
    
    # Write data
    bumg_list = BUMG.objects.all().order_by('nama')
    for row_num, bumg in enumerate(bumg_list, 2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1)
        worksheet.cell(row=row_num, column=2, value=bumg.nama)
        worksheet.cell(row=row_num, column=3, value=bumg.alamat)
        worksheet.cell(row=row_num, column=4, value=bumg.telepon)
        worksheet.cell(row=row_num, column=5, value=bumg.email)
        worksheet.cell(row=row_num, column=6, value=bumg.jenis_usaha)
        worksheet.cell(row=row_num, column=7, value=float(bumg.modal_awal))
        worksheet.cell(row=row_num, column=8, value=bumg.jumlah_karyawan)
        worksheet.cell(row=row_num, column=9, value=bumg.status)
        worksheet.cell(row=row_num, column=10, value=bumg.created_at.strftime('%Y-%m-%d'))
    
    filename = f'Data_BUMG_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(filename, workbook)


@login_required
def export_ukm(request):
    """Export UKM data to Excel"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data UKM'
    
    # Headers
    headers = ['No', 'Nama UKM', 'Pemilik', 'Alamat', 'Telepon', 'Email', 
               'Jenis Usaha', 'Modal Awal', 'Jumlah Karyawan', 'Status', 'Tanggal Dibuat']
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    style_header_row(worksheet, 1, headers)
    
    # Write data
    ukm_list = UKM.objects.all().order_by('nama')
    for row_num, ukm in enumerate(ukm_list, 2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1)
        worksheet.cell(row=row_num, column=2, value=ukm.nama)
        worksheet.cell(row=row_num, column=3, value=ukm.pemilik)
        worksheet.cell(row=row_num, column=4, value=ukm.alamat)
        worksheet.cell(row=row_num, column=5, value=ukm.telepon)
        worksheet.cell(row=row_num, column=6, value=ukm.email)
        worksheet.cell(row=row_num, column=7, value=ukm.jenis_usaha)
        worksheet.cell(row=row_num, column=8, value=float(ukm.modal_awal))
        worksheet.cell(row=row_num, column=9, value=ukm.jumlah_karyawan)
        worksheet.cell(row=row_num, column=10, value=ukm.status)
        worksheet.cell(row=row_num, column=11, value=ukm.created_at.strftime('%Y-%m-%d'))
    
    filename = f'Data_UKM_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(filename, workbook)


@login_required
def export_aset(request):
    """Export Aset data to Excel"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data Aset'
    
    # Headers
    headers = ['No', 'Nama Aset', 'Jenis Aset', 'Lokasi', 'Kondisi', 
               'Nilai Aset', 'Keterangan', 'Status', 'Tanggal Dibuat']
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    style_header_row(worksheet, 1, headers)
    
    # Write data
    aset_list = Aset.objects.all().order_by('nama')
    for row_num, aset in enumerate(aset_list, 2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1)
        worksheet.cell(row=row_num, column=2, value=aset.nama)
        worksheet.cell(row=row_num, column=3, value=aset.jenis_aset)
        worksheet.cell(row=row_num, column=4, value=aset.lokasi)
        worksheet.cell(row=row_num, column=5, value=aset.kondisi)
        worksheet.cell(row=row_num, column=6, value=float(aset.nilai_aset))
        worksheet.cell(row=row_num, column=7, value=aset.keterangan)
        worksheet.cell(row=row_num, column=8, value=aset.status)
        worksheet.cell(row=row_num, column=9, value=aset.created_at.strftime('%Y-%m-%d'))
    
    filename = f'Data_Aset_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(filename, workbook)


@login_required
def export_jasa(request):
    """Export Layanan Jasa data to Excel"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data Layanan Jasa'
    
    # Headers
    headers = ['No', 'Nama Layanan', 'Penyedia Jasa', 'Alamat', 'Telepon', 
               'Email', 'Jenis Layanan', 'Tarif', 'Status', 'Tanggal Dibuat']
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    style_header_row(worksheet, 1, headers)
    
    # Write data
    jasa_list = LayananJasa.objects.all().order_by('nama_layanan')
    for row_num, jasa in enumerate(jasa_list, 2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1)
        worksheet.cell(row=row_num, column=2, value=jasa.nama_layanan)
        worksheet.cell(row=row_num, column=3, value=jasa.penyedia_jasa)
        worksheet.cell(row=row_num, column=4, value=jasa.alamat)
        worksheet.cell(row=row_num, column=5, value=jasa.telepon)
        worksheet.cell(row=row_num, column=6, value=jasa.email)
        worksheet.cell(row=row_num, column=7, value=jasa.jenis_layanan)
        worksheet.cell(row=row_num, column=8, value=float(jasa.tarif))
        worksheet.cell(row=row_num, column=9, value=jasa.status)
        worksheet.cell(row=row_num, column=10, value=jasa.created_at.strftime('%Y-%m-%d'))
    
    filename = f'Data_Layanan_Jasa_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(filename, workbook)


# Public API endpoints for statistics (no authentication required)
def koperasi_count_api(request):
    """Get Koperasi count for dashboard"""
    if request.method == 'GET':
        total = Koperasi.objects.count()
        active = Koperasi.objects.filter(status='aktif').count()
        return JsonResponse({
            'total': total,
            'active': active
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)


def bumg_count_api(request):
    """Get BUMG count for dashboard"""
    if request.method == 'GET':
        total = BUMG.objects.count()
        active = BUMG.objects.filter(status='aktif').count()
        return JsonResponse({
            'total': total,
            'active': active
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)


def ukm_count_api(request):
    """Get UKM count for dashboard"""
    if request.method == 'GET':
        total = UKM.objects.count()
        active = UKM.objects.filter(status='aktif').count()
        return JsonResponse({
            'total': total,
            'active': active
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)


def aset_count_api(request):
    """Get Aset count for dashboard"""
    if request.method == 'GET':
        total = Aset.objects.count()
        good_condition = Aset.objects.filter(kondisi='baik').count()
        return JsonResponse({
            'total': total,
            'good_condition': good_condition
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)


def jasa_count_api(request):
    """Get LayananJasa count for dashboard"""
    if request.method == 'GET':
        total = LayananJasa.objects.count()
        active = LayananJasa.objects.filter(status='aktif').count()
        return JsonResponse({
            'total': total,
            'active': active
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)


def public_business_statistics(request):
    """Get public business statistics for dashboard"""
    if request.method == 'GET':
        koperasi_total = Koperasi.objects.count()
        bumg_total = BUMG.objects.count()
        ukm_total = UKM.objects.count()
        aset_total = Aset.objects.count()
        jasa_total = LayananJasa.objects.count()
        
        return JsonResponse({
            'koperasi': koperasi_total,
            'bumg': bumg_total,
            'ukm': ukm_total,
            'aset': aset_total,
            'jasa': jasa_total,
            'total_all': koperasi_total + bumg_total + ukm_total + aset_total + jasa_total
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# Statistics API for each module
def koperasi_statistics_api(request):
    """Get detailed Koperasi statistics"""
    if request.method == 'GET':
        total = Koperasi.objects.count()
        active = Koperasi.objects.filter(status='aktif').count()
        total_members = Koperasi.objects.aggregate(Sum('jumlah_anggota'))['jumlah_anggota__sum'] or 0
        total_assets = Koperasi.objects.aggregate(Sum('total_aset'))['total_aset__sum'] or 0
        
        return JsonResponse({
            'total_koperasi': total,
            'koperasi_aktif': active,
            'total_anggota': total_members,
            'total_aset': float(total_assets)
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def bumg_statistics_api(request):
    """Get detailed BUMG statistics"""
    if request.method == 'GET':
        total = BUMG.objects.count()
        active = BUMG.objects.filter(status='aktif').count()
        total_employees = BUMG.objects.aggregate(Sum('jumlah_karyawan'))['jumlah_karyawan__sum'] or 0
        total_assets = BUMG.objects.aggregate(Sum('total_aset'))['total_aset__sum'] or 0
        
        return JsonResponse({
            'total_bumg': total,
            'bumg_aktif': active,
            'total_karyawan': total_employees,
            'total_aset': float(total_assets)
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def ukm_statistics_api(request):
    """Get detailed UKM statistics"""
    if request.method == 'GET':
        total = UKM.objects.count()
        active = UKM.objects.filter(status='aktif').count()
        total_workers = UKM.objects.aggregate(Sum('jumlah_pekerja'))['jumlah_pekerja__sum'] or 0
        total_turnover = UKM.objects.aggregate(Sum('omzet_bulanan'))['omzet_bulanan__sum'] or 0
        
        return JsonResponse({
            'total_ukm': total,
            'ukm_aktif': active,
            'total_pekerja': total_workers,
            'omzet_bulanan': float(total_turnover)
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def aset_statistics_api(request):
    """Get detailed Aset statistics"""
    if request.method == 'GET':
        try:
            total = Aset.objects.count()
            good_condition = Aset.objects.filter(kondisi='baik').count()
            needs_repair = Aset.objects.filter(kondisi__in=['rusak_ringan', 'rusak_berat', 'tidak_dapat_digunakan']).count()
            
            # Simple approach without aggregation for now
            total_value = 0
            
            return JsonResponse({
                'total_aset': total,
                'aset_baik': good_condition,
                'perlu_perbaikan': needs_repair,
                'total_nilai': total_value
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def layanan_statistics_api(request):
    """Get detailed Layanan Jasa statistics"""
    if request.method == 'GET':
        total = LayananJasa.objects.count()
        active = LayananJasa.objects.filter(status='aktif').count()
        providers = LayananJasa.objects.values('penyedia_jasa').distinct().count()
        avg_tariff = LayananJasa.objects.aggregate(avg_tariff=Avg('tarif'))['avg_tariff'] or 0
        
        return JsonResponse({
            'total_layanan': total,
            'layanan_aktif': active,
            'penyedia_jasa': providers,
            'rata_rata_tarif': float(avg_tariff)
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def kategori_statistics_api(request):
    """Get detailed Kategori statistics"""
    if request.method == 'GET':
        total = BusinessCategory.objects.count()
        active = BusinessCategory.objects.filter(is_active=True).count()
        parent_categories = BusinessCategory.objects.filter(parent__isnull=True).count()
        sub_categories = BusinessCategory.objects.filter(parent__isnull=False).count()
        
        return JsonResponse({
            'total_kategori': total,
            'kategori_aktif': active,
            'kategori_induk': parent_categories,
            'sub_kategori': sub_categories
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)

# CRUD Operations for Business Categories
@login_required
@csrf_protect
def category_api(request):
    """API endpoint for business category CRUD operations"""
    if request.method == 'GET':
        categories = BusinessCategory.objects.all().order_by('name')
        
        # Search functionality
        search = request.GET.get('search', '')
        if search:
            categories = categories.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Pagination
        paginator = Paginator(categories, 20)
        page = request.GET.get('page', 1)
        categories_page = paginator.get_page(page)
        
        data = []
        for category in categories_page:
            data.append({
                'id': category.id,
                'name': category.name,
                'description': category.description,
                'is_active': category.is_active,
                'created_at': category.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': category.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            })
        
        return JsonResponse({
            'data': data,
            'total': paginator.count,
            'page': categories_page.number,
            'total_pages': paginator.num_pages,
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            category = BusinessCategory.objects.create(
                name=data['name'],
                description=data.get('description', ''),
                is_active=data.get('is_active', True)
            )
            return JsonResponse({'success': True, 'id': category.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            category_id = data.get('id')
            category = get_object_or_404(BusinessCategory, id=category_id)
            
            category.name = data['name']
            category.description = data.get('description', '')
            category.is_active = data.get('is_active', True)
            category.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            category_id = data.get('id')
            category = get_object_or_404(BusinessCategory, id=category_id)
            category.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


@login_required
def category_view(request):
    """Category management view"""
    context = {
        'page_title': 'Kategori Bisnis',
        'page_subtitle': 'Kelola kategori untuk semua jenis bisnis'
    }
    return render(request, 'admin/modules/business/category.html', context)


def category_list_api(request):
    """Public API to get list of active categories for dropdowns"""
    if request.method == 'GET':
        categories = BusinessCategory.objects.filter(is_active=True).order_by('name')
        data = []
        for category in categories:
            data.append({
                'id': category.id,
                'name': category.name,
                'description': category.description
            })
        return JsonResponse({'categories': data})
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# CRUD Operations for Jenis Koperasi
@login_required
@csrf_protect
def jenis_koperasi_api(request):
    """API endpoint for Jenis Koperasi CRUD operations"""
    if request.method == 'GET':
        jenis_koperasi = JenisKoperasi.objects.filter(is_active=True).order_by('nama')
        
        # Search functionality
        search = request.GET.get('search', '')
        if search:
            jenis_koperasi = jenis_koperasi.filter(
                Q(nama__icontains=search) |
                Q(deskripsi__icontains=search)
            )
        
        # Pagination
        paginator = Paginator(jenis_koperasi, 20)
        page = request.GET.get('page', 1)
        jenis_page = paginator.get_page(page)
        
        data = []
        for jenis in jenis_page:
            data.append({
                'id': jenis.id,
                'nama': jenis.nama,
                'deskripsi': jenis.deskripsi,
                'is_active': jenis.is_active,
                'created_at': jenis.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': jenis.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            })
        
        return JsonResponse({
            'data': data,
            'total': paginator.count,
            'page': jenis_page.number,
            'total_pages': paginator.num_pages,
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            jenis = JenisKoperasi.objects.create(
                nama=data['nama'],
                deskripsi=data.get('deskripsi', ''),
                is_active=data.get('is_active', True)
            )
            return JsonResponse({'success': True, 'id': jenis.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            jenis_id = data.get('id')
            jenis = get_object_or_404(JenisKoperasi, id=jenis_id)
            
            jenis.nama = data['nama']
            jenis.deskripsi = data.get('deskripsi', '')
            jenis.is_active = data.get('is_active', True)
            jenis.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            jenis_id = data.get('id')
            jenis = get_object_or_404(JenisKoperasi, id=jenis_id)
            jenis.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

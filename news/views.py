from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from django.conf import settings
from django.core.files.storage import default_storage
from django.template.loader import render_to_string
from django.urls import reverse
import json
import csv
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta

from .models import (
    NewsCategory, NewsTag, News, NewsComment, NewsView, 
    NewsImage, NewsLike, NewsShare, Announcement
)
from .forms import (
    NewsCategoryForm, NewsTagForm, NewsForm, NewsCommentForm, 
    NewsSearchForm, NewsImageForm, NewsImageFormSet, 
    NewsCommentModerationForm, BulkNewsActionForm, NewsImportForm,
    AnnouncementForm, AnnouncementSearchForm
)


@login_required
def news_admin(request):
    """Main admin view for news management"""
    # Get statistics
    total_news = News.objects.count()
    published_news = News.objects.filter(status='published').count()
    draft_news = News.objects.filter(status='draft').count()
    scheduled_news = News.objects.filter(status='scheduled').count()
    total_views = NewsView.objects.count()
    total_likes = NewsLike.objects.count()
    total_comments = NewsComment.objects.filter(status='approved').count()
    
    # Debug print
    print(f"DEBUG - Total news: {total_news}")
    print(f"DEBUG - Published: {published_news}")
    print(f"DEBUG - Draft: {draft_news}")
    print(f"DEBUG - Scheduled: {scheduled_news}")
    print(f"DEBUG - Comments: {total_comments}")
    
    # Recent news
    recent_news = News.objects.select_related('category', 'author').order_by('-created_at')[:5]
    
    # Popular news (by views)
    popular_news = News.objects.filter(status='published').annotate(
        views_count=Count('view_records')
    ).order_by('-views_count')[:5]
    
    # Recent comments
    recent_comments = NewsComment.objects.select_related('news').filter(
        status='pending'
    ).order_by('-created_at')[:5]
    
    context = {
        'page_title': 'Berita',
        'page_subtitle': 'Kelola berita dan artikel desa',
        'total_news': total_news,
        'published_news': published_news,
        'draft_news': draft_news,
        'scheduled_news': scheduled_news,
        'total_views': total_views,
        'total_likes': total_likes,
        'total_comments': total_comments,
        'recent_news': recent_news,
        'popular_news': popular_news,
        'recent_comments': recent_comments,
    }
    
    return render(request, 'admin/modules/news/index.html', context)


@login_required
def news_list(request):
    """List all news with filtering and pagination"""
    news_list = News.objects.select_related('category', 'author').prefetch_related('tags')
    
    # Search and filtering
    search_form = NewsSearchForm(request.GET)
    if search_form.is_valid():
        query = search_form.cleaned_data.get('query')
        category = search_form.cleaned_data.get('category')
        tags = search_form.cleaned_data.get('tags')
        date_from = search_form.cleaned_data.get('date_from')
        date_to = search_form.cleaned_data.get('date_to')
        
        if query:
            news_list = news_list.filter(
                Q(title__icontains=query) | 
                Q(content__icontains=query) |
                Q(excerpt__icontains=query)
            )
        
        if category:
            news_list = news_list.filter(category=category)
        
        if tags:
            news_list = news_list.filter(tags__in=tags).distinct()
        
        if date_from:
            news_list = news_list.filter(created_at__gte=date_from)
        
        if date_to:
            news_list = news_list.filter(created_at__lte=date_to)
    
    # Status filtering
    status = request.GET.get('status')
    if status:
        news_list = news_list.filter(status=status)
    
    # Ordering
    order_by = request.GET.get('order_by', '-created_at')
    news_list = news_list.order_by(order_by)
    
    # Pagination
    paginator = Paginator(news_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Bulk action form
    bulk_form = BulkNewsActionForm()
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
        'bulk_form': bulk_form,
        'current_status': status,
        'current_order': order_by,
    }
    
    return render(request, 'admin/modules/news/list.html', context)


@login_required
def news_create(request):
    """Create new news"""
    if request.method == 'POST':
        form = NewsForm(request.POST, request.FILES)
        image_formset = NewsImageFormSet(request.POST, request.FILES)
        
        if form.is_valid() and image_formset.is_valid():
            with transaction.atomic():
                news = form.save(commit=False)
                news.author = request.user
                news.slug = slugify(news.title)
                
                news.save()
                form.save_m2m()  # Save many-to-many relationships (tags)
                
                # Save images
                image_formset.instance = news
                image_formset.save()
                
                messages.success(request, f'Berita "{news.title}" berhasil dibuat.')
                return redirect('news:news_view_detail', pk=news.pk)
    else:
        form = NewsForm()
        image_formset = NewsImageFormSet()
    
    context = {
        'form': form,
        'image_formset': image_formset,
        'action': 'create',
    }
    
    return render(request, 'admin/modules/news/create.html', context)


@login_required
def news_detail(request, pk):
    """View news detail"""
    news = get_object_or_404(News.objects.select_related('category', 'author').prefetch_related('tags', 'images'), pk=pk)
    
    # Get comments
    comments = news.comments.filter(status='approved').order_by('-created_at')
    pending_comments = news.comments.filter(status='pending').count()
    
    # Get related news
    related_news = News.objects.filter(
        category=news.category,
        status='published'
    ).exclude(pk=news.pk)[:5]
    
    context = {
        'news': news,
        'comments': comments,
        'pending_comments': pending_comments,
        'related_news': related_news,
    }
    
    return render(request, 'admin/modules/news/detail.html', context)


@login_required
def news_edit(request, pk):
    """Edit existing news"""
    news = get_object_or_404(News, pk=pk)
    
    if request.method == 'POST':
        form = NewsForm(request.POST, request.FILES, instance=news)
        image_formset = NewsImageFormSet(request.POST, request.FILES, instance=news)
        
        if form.is_valid() and image_formset.is_valid():
            with transaction.atomic():
                news = form.save(commit=False)
                
                # Update slug if title changed
                if 'title' in form.changed_data:
                    news.slug = slugify(news.title)
                
                # Recalculate reading time if content changed
                if 'content' in form.changed_data:
                    news.reading_time = news.calculate_reading_time()
                
                news.save()
                form.save_m2m()
                
                # Save images
                image_formset.save()
                
                messages.success(request, f'Berita "{news.title}" berhasil diperbarui.')
                return redirect('news:detail', pk=news.pk)
    else:
        form = NewsForm(instance=news)
        image_formset = NewsImageFormSet(instance=news)
    
    context = {
        'form': form,
        'image_formset': image_formset,
        'news': news,
        'action': 'edit',
    }
    
    return render(request, 'admin/modules/news/edit.html', context)


@login_required
@require_POST
def news_delete(request, pk):
    """Delete news"""
    news = get_object_or_404(News, pk=pk)
    title = news.title
    
    # Delete associated images from storage
    for image in news.images.all():
        if image.image:
            default_storage.delete(image.image.name)
        if image.thumbnail:
            default_storage.delete(image.thumbnail.name)
    
    news.delete()
    messages.success(request, f'Berita "{title}" berhasil dihapus.')
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': f'Berita "{title}" berhasil dihapus.'})
    
    return redirect('news:list')


@login_required
@require_POST
def news_bulk_action(request):
    """Handle bulk actions on news"""
    form = BulkNewsActionForm(request.POST)
    
    if form.is_valid():
        action = form.cleaned_data['action']
        selected_ids = form.cleaned_data['selected_news'].split(',')
        
        try:
            selected_ids = [int(id.strip()) for id in selected_ids if id.strip()]
            news_queryset = News.objects.filter(id__in=selected_ids)
            
            if action == 'publish':
                news_queryset.update(status='published', published_date=timezone.now())
                messages.success(request, f'{news_queryset.count()} berita berhasil dipublikasikan.')
            
            elif action == 'draft':
                news_queryset.update(status='draft')
                messages.success(request, f'{news_queryset.count()} berita berhasil dijadikan draft.')
            
            elif action == 'archive':
                news_queryset.update(status='archived')
                messages.success(request, f'{news_queryset.count()} berita berhasil diarsipkan.')
            
            elif action == 'featured':
                news_queryset.update(is_featured=True)
                messages.success(request, f'{news_queryset.count()} berita berhasil dijadikan featured.')
            
            elif action == 'unfeatured':
                news_queryset.update(is_featured=False)
                messages.success(request, f'{news_queryset.count()} berita berhasil dihapus dari featured.')
            
            elif action == 'delete':
                # Delete associated images
                for news in news_queryset:
                    for image in news.images.all():
                        if image.image:
                            default_storage.delete(image.image.name)
                        if image.thumbnail:
                            default_storage.delete(image.thumbnail.name)
                
                count = news_queryset.count()
                news_queryset.delete()
                messages.success(request, f'{count} berita berhasil dihapus.')
            
        except ValueError:
            messages.error(request, 'ID berita tidak valid.')
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
    
    return redirect('news:list')


@csrf_exempt
@require_POST
def news_like(request, pk):
    """Like/unlike news"""
    news = get_object_or_404(News, pk=pk)
    
    # Get user info
    user = request.user if request.user.is_authenticated else None
    ip_address = request.META.get('REMOTE_ADDR')
    session_key = request.session.session_key
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    
    # Check if already liked
    like_filter = {'news': news}
    if user:
        like_filter['user'] = user
    else:
        like_filter['ip_address'] = ip_address
        like_filter['session_key'] = session_key
    
    existing_like = NewsLike.objects.filter(**like_filter).first()
    
    if existing_like:
        # Unlike
        existing_like.delete()
        liked = False
    else:
        # Like
        NewsLike.objects.create(
            news=news,
            user=user,
            ip_address=ip_address,
            session_key=session_key,
            user_agent=user_agent
        )
        liked = True
    
    # Update likes count
    news.update_likes_count()
    
    return JsonResponse({
        'success': True,
        'liked': liked,
        'likes_count': news.likes_count
    })


@csrf_exempt
@require_POST
def news_share(request, pk):
    """Record news share"""
    news = get_object_or_404(News, pk=pk)
    platform = request.POST.get('platform', 'unknown')
    
    # Get user info
    user = request.user if request.user.is_authenticated else None
    ip_address = request.META.get('REMOTE_ADDR')
    session_key = request.session.session_key
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    referrer = request.META.get('HTTP_REFERER', '')
    
    # Record share
    NewsShare.objects.create(
        news=news,
        platform=platform,
        user=user,
        ip_address=ip_address,
        session_key=session_key,
        user_agent=user_agent,
        referrer=referrer
    )
    
    # Update shares count
    news.update_shares_count()
    
    return JsonResponse({
        'success': True,
        'shares_count': news.shares_count
    })


@login_required
def news_comments(request, pk):
    """Manage news comments"""
    news = get_object_or_404(News, pk=pk)
    
    # Get comments with pagination
    comments_list = news.comments.select_related('user').order_by('-created_at')
    
    # Filter by moderation status
    status = request.GET.get('status')
    if status:
        comments_list = comments_list.filter(status=status)
    
    paginator = Paginator(comments_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'news': news,
        'page_obj': page_obj,
        'current_status': status,
    }
    
    return render(request, 'admin/modules/news/comments.html', context)


@login_required
@require_POST
def comment_moderate(request, comment_id):
    """Moderate comment (approve/reject)"""
    comment = get_object_or_404(NewsComment, pk=comment_id)
    action = request.POST.get('action')
    
    if action == 'approve':
        comment.status = 'approved'
        comment.moderated_by = request.user
        comment.moderated_at = timezone.now()
        comment.save()
        
        # Update comments count
        comment.news.update_comments_count()
        
        messages.success(request, 'Komentar berhasil disetujui.')
    
    elif action == 'reject':
        comment.status = 'rejected'
        comment.moderated_by = request.user
        comment.moderated_at = timezone.now()
        comment.save()
        
        messages.success(request, 'Komentar berhasil ditolak.')
    
    elif action == 'delete':
        comment.delete()
        messages.success(request, 'Komentar berhasil dihapus.')
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    return redirect('news:comments', pk=comment.news.pk)


@login_required
def news_import(request):
    """Import news from file"""
    if request.method == 'POST':
        form = NewsImportForm(request.POST, request.FILES)
        
        if form.is_valid():
            file = form.cleaned_data['file']
            
            try:
                imported_count = 0
                errors = []
                
                if file.name.endswith('.csv'):
                    # Handle CSV import
                    decoded_file = file.read().decode('utf-8')
                    csv_reader = csv.DictReader(decoded_file.splitlines())
                    
                    for row_num, row in enumerate(csv_reader, start=2):
                        try:
                            # Create news from CSV row
                            news = News(
                                title=row.get('title', ''),
                                content=row.get('content', ''),
                                excerpt=row.get('excerpt', ''),
                                author=request.user,
                                status=row.get('status', 'draft')
                            )
                            
                            # Set category if provided
                            category_name = row.get('category')
                            if category_name:
                                category, created = NewsCategory.objects.get_or_create(
                                    name=category_name,
                                    defaults={'slug': slugify(category_name)}
                                )
                                news.category = category
                            
                            news.slug = slugify(news.title)
                            news.reading_time = news.calculate_reading_time()
                            news.save()
                            
                            # Add tags if provided
                            tags_str = row.get('tags', '')
                            if tags_str:
                                tag_names = [tag.strip() for tag in tags_str.split(',')]
                                for tag_name in tag_names:
                                    if tag_name:
                                        tag, created = NewsTag.objects.get_or_create(
                                            name=tag_name,
                                            defaults={'slug': slugify(tag_name)}
                                        )
                                        news.tags.add(tag)
                            
                            imported_count += 1
                            
                        except Exception as e:
                            errors.append(f'Baris {row_num}: {str(e)}')
                
                elif file.name.endswith(('.xlsx', '.xls')):
                    # Handle Excel import
                    workbook = openpyxl.load_workbook(file)
                    worksheet = workbook.active
                    
                    # Assume first row is header
                    headers = [cell.value for cell in worksheet[1]]
                    
                    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            row_data = dict(zip(headers, row))
                            
                            news = News(
                                title=row_data.get('title', ''),
                                content=row_data.get('content', ''),
                                excerpt=row_data.get('excerpt', ''),
                                author=request.user,
                                status=row_data.get('status', 'draft')
                            )
                            
                            # Set category if provided
                            category_name = row_data.get('category')
                            if category_name:
                                category, created = NewsCategory.objects.get_or_create(
                                    name=category_name,
                                    defaults={'slug': slugify(category_name)}
                                )
                                news.category = category
                            
                            news.slug = slugify(news.title)
                            news.reading_time = news.calculate_reading_time()
                            news.save()
                            
                            # Add tags if provided
                            tags_str = row_data.get('tags', '')
                            if tags_str:
                                tag_names = [tag.strip() for tag in str(tags_str).split(',')]
                                for tag_name in tag_names:
                                    if tag_name:
                                        tag, created = NewsTag.objects.get_or_create(
                                            name=tag_name,
                                            defaults={'slug': slugify(tag_name)}
                                        )
                                        news.tags.add(tag)
                            
                            imported_count += 1
                            
                        except Exception as e:
                            errors.append(f'Baris {row_num}: {str(e)}')
                
                if imported_count > 0:
                    messages.success(request, f'{imported_count} berita berhasil diimpor.')
                
                if errors:
                    error_msg = 'Beberapa baris gagal diimpor:\n' + '\n'.join(errors[:10])
                    if len(errors) > 10:
                        error_msg += f'\n... dan {len(errors) - 10} error lainnya'
                    messages.warning(request, error_msg)
                
                return redirect('news:list')
                
            except Exception as e:
                messages.error(request, f'Gagal mengimpor file: {str(e)}')
    
    else:
        form = NewsImportForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'admin/modules/news/import.html', context)


@login_required
def news_export(request):
    """Export news to Excel"""
    # Get filtered news
    news_list = News.objects.select_related('category', 'author').prefetch_related('tags')
    
    # Apply filters from request
    status = request.GET.get('status')
    if status:
        news_list = news_list.filter(status=status)
    
    category_id = request.GET.get('category')
    if category_id:
        news_list = news_list.filter(category_id=category_id)
    
    # Create Excel file
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'News Export'
    
    # Headers
    headers = [
        'ID', 'Judul', 'Kategori', 'Tags', 'Status', 'Prioritas',
        'Featured', 'Breaking', 'Views', 'Likes', 'Comments',
        'Penulis', 'Tanggal Dibuat', 'Tanggal Publikasi'
    ]
    
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # Data
    for row, news in enumerate(news_list, 2):
        worksheet.cell(row=row, column=1, value=news.id)
        worksheet.cell(row=row, column=2, value=news.title)
        worksheet.cell(row=row, column=3, value=news.category.name if news.category else '')
        worksheet.cell(row=row, column=4, value=', '.join([tag.name for tag in news.tags.all()]))
        worksheet.cell(row=row, column=5, value=news.get_status_display())
        worksheet.cell(row=row, column=6, value=news.get_priority_display())
        worksheet.cell(row=row, column=7, value='Ya' if news.is_featured else 'Tidak')
        worksheet.cell(row=row, column=8, value='Ya' if news.is_breaking else 'Tidak')
        worksheet.cell(row=row, column=9, value=news.views_count)
        worksheet.cell(row=row, column=10, value=news.likes_count)
        worksheet.cell(row=row, column=11, value=news.comments_count)
        worksheet.cell(row=row, column=12, value=news.author.get_full_name() or news.author.username)
        worksheet.cell(row=row, column=13, value=news.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        worksheet.cell(row=row, column=14, value=news.published_date.strftime('%Y-%m-%d %H:%M:%S') if news.published_date else '')
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="news_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Save workbook to response
    workbook.save(response)
    
    return response


# News Category API Views
@login_required
@require_http_methods(["GET"])
def news_category_list_api(request):
    """API to get list of news categories"""
    try:
        search = request.GET.get('search', '')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        
        queryset = NewsCategory.objects.annotate(
            news_count=Count('news')
        )
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        
        paginator = Paginator(queryset, per_page)
        page_obj = paginator.get_page(page)
        
        data = {
            'results': [
                {
                    'id': item.id,
                    'name': item.name,
                    'description': item.description,
                    'slug': item.slug,
                    'color': item.color,
                    'is_active': item.is_active,
                    'news_count': item.news_count,
                    'created_at': item.created_at.strftime('%d/%m/%Y %H:%M')
                }
                for item in page_obj
            ],
            'pagination': {
                'current_page': page,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def news_bulk_action_api(request):
    """API for bulk news actions"""
    try:
        data = json.loads(request.body)
        action = data.get('action')
        news_ids = data.get('news_ids', [])
        
        if not news_ids:
            return JsonResponse({'error': 'No news selected'}, status=400)
        
        news_queryset = News.objects.filter(id__in=news_ids)
        
        if action == 'publish':
            news_queryset.update(status='published')
            return JsonResponse({
                'success': True,
                'message': f'{news_queryset.count()} berita berhasil dipublikasikan'
            })
        
        elif action == 'draft':
            news_queryset.update(status='draft')
            return JsonResponse({
                'success': True,
                'message': f'{news_queryset.count()} berita berhasil dijadikan draft'
            })
        
        elif action == 'archive':
            news_queryset.update(status='archived')
            return JsonResponse({
                'success': True,
                'message': f'{news_queryset.count()} berita berhasil diarsipkan'
            })
        
        elif action == 'delete':
            deleted_count = news_queryset.count()
            news_queryset.delete()
            return JsonResponse({
                'success': True,
                'message': f'{deleted_count} berita berhasil dihapus'
            })
        
        else:
            return JsonResponse({'error': 'Invalid action'}, status=400)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["GET", "PATCH"])
def news_category_detail_api(request, pk):
    """API to get news category detail or update order"""
    try:
        category = get_object_or_404(NewsCategory, pk=pk)
        
        if request.method == 'GET':
            data = {
                'id': category.id,
                'name': category.name,
                'description': category.description,
                'slug': category.slug,
                'color': category.color,
                'is_active': category.is_active,
                'order': getattr(category, 'order', 0),
                'created_at': category.created_at.strftime('%d/%m/%Y %H:%M'),
                'updated_at': category.updated_at.strftime('%d/%m/%Y %H:%M')
            }
            return JsonResponse(data)
            
        elif request.method == 'PATCH':
            data = json.loads(request.body)
            
            # Update order if provided
            if 'order' in data:
                if hasattr(category, 'order'):
                    category.order = data['order']
                    category.save()
                    return JsonResponse({
                        'success': True,
                        'message': 'Urutan kategori berhasil diperbarui'
                    })
                else:
                    return JsonResponse({
                        'error': 'Model NewsCategory tidak memiliki field order'
                    }, status=400)
            
            return JsonResponse({'error': 'Data tidak valid'}, status=400)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def news_category_create_api(request):
    """API to create new news category"""
    try:
        data = json.loads(request.body)
        
        category = NewsCategory.objects.create(
            name=data['name'],
            description=data.get('description', ''),
            color=data.get('color', '#007bff'),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Kategori berhasil ditambahkan',
            'data': {
                'id': category.id,
                'name': category.name
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def news_category_update_api(request, pk):
    """API to update news category"""
    try:
        category = get_object_or_404(NewsCategory, pk=pk)
        data = json.loads(request.body)
        
        category.name = data['name']
        category.description = data.get('description', '')
        category.color = data.get('color', '#007bff')
        category.is_active = data.get('is_active', True)
        category.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Kategori berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def news_category_delete_api(request, pk):
    """API to delete news category"""
    try:
        category = get_object_or_404(NewsCategory, pk=pk)
        
        # Check if category has news
        if category.news.exists():
            return JsonResponse({
                'error': 'Kategori tidak dapat dihapus karena masih memiliki berita'
            }, status=400)
        
        category.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Kategori berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def news_category_statistics_api(request):
    """API to get news category statistics"""
    try:
        total_categories = NewsCategory.objects.count()
        active_categories = NewsCategory.objects.filter(is_active=True).count()
        categories_with_news = NewsCategory.objects.annotate(
            news_count=Count('news')
        ).filter(news_count__gt=0).count()
        
        # Category distribution with news count
        category_distribution = NewsCategory.objects.annotate(
            news_count=Count('news')
        ).values('id', 'name', 'color', 'news_count', 'is_active')
        
        data = {
            'total_categories': total_categories,
            'active_categories': active_categories,
            'categories_with_news': categories_with_news,
            'category_distribution': list(category_distribution)
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# News Tag API Views
@login_required
@require_http_methods(["GET"])
def news_tag_list_api(request):
    """API to get list of news tags"""
    try:
        search = request.GET.get('search', '')
        
        queryset = NewsTag.objects.annotate(
            news_count=Count('news')
        )
        
        if search:
            queryset = queryset.filter(name__icontains=search)
        
        data = {
            'results': [
                {
                    'id': item.id,
                    'name': item.name,
                    'slug': item.slug,
                    'news_count': item.news_count,
                    'created_at': item.created_at.strftime('%d/%m/%Y %H:%M')
                }
                for item in queryset
            ]
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def news_tag_search_api(request):
    """API to search news tags"""
    try:
        query = request.GET.get('q', '')
        
        if not query:
            return JsonResponse({'results': []})
        
        tags = NewsTag.objects.filter(name__icontains=query)[:10]
        
        data = {
            'results': [
                {
                    'id': tag.id,
                    'name': tag.name,
                    'slug': tag.slug
                }
                for tag in tags
            ]
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def news_tag_create_api(request):
    """API to create new news tag"""
    try:
        data = json.loads(request.body)
        
        tag = NewsTag.objects.create(
            name=data['name']
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Tag berhasil ditambahkan',
            'data': {
                'id': tag.id,
                'name': tag.name
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def news_tag_delete_api(request, pk):
    """API to delete news tag"""
    try:
        tag = get_object_or_404(NewsTag, pk=pk)
        tag.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Tag berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# News API Views
@login_required
@require_http_methods(["GET"])
def news_list_api(request):
    """API to get list of news"""
    try:
        search = request.GET.get('search', '')
        category_id = request.GET.get('category_id', '')
        status = request.GET.get('status', '')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        
        queryset = News.objects.select_related('category', 'author').prefetch_related('tags')
        
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(content__icontains=search) |
                Q(excerpt__icontains=search)
            )
        
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        
        if status:
            queryset = queryset.filter(status=status)
        
        paginator = Paginator(queryset, per_page)
        page_obj = paginator.get_page(page)
        
        data = {
            'results': [
                {
                    'id': item.id,
                    'title': item.title,
                    'slug': item.slug,
                    'category': item.category.name,
                    'category_color': item.category.color,
                    'author': item.author.get_full_name() or item.author.username,
                    'status': item.status,
                    'priority': item.priority,
                    'is_featured': item.is_featured,
                    'is_breaking': item.is_breaking,
                    'published_date': item.published_date.strftime('%d/%m/%Y %H:%M') if item.published_date else None,
                    'views_count': item.views_count,
                    'likes_count': item.likes_count,
                    'created_at': item.created_at.strftime('%d/%m/%Y %H:%M'),
                    'excerpt': item.excerpt,
                    'featured_image': item.featured_image.url if item.featured_image else None,
                    'tags': [tag.name for tag in item.tags.all()]
                }
                for item in page_obj
            ],
            'pagination': {
                'current_page': page,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def news_detail_api(request, pk):
    """API to get news detail"""
    try:
        news = get_object_or_404(News.objects.select_related('category', 'author').prefetch_related('tags', 'images'), pk=pk)
        
        data = {
            'id': news.id,
            'title': news.title,
            'slug': news.slug,
            'category_id': news.category.id,
            'category': news.category.name,
            'content': news.content,
            'excerpt': news.excerpt,
            'featured_image': news.featured_image.url if news.featured_image else None,
            'status': news.status,
            'priority': news.priority,
            'is_featured': news.is_featured,
            'is_breaking': news.is_breaking,
            'published_date': news.published_date.isoformat() if news.published_date else None,
            'publish_date': news.published_date.isoformat() if news.published_date else None,
            'author': news.author.get_full_name() or news.author.username,
            'views_count': news.views_count,
            'likes_count': news.likes_count,
            'meta_title': news.meta_title,
            'meta_description': news.meta_description,
            'youtube_url': news.youtube_url,
            'video_file': news.video_file.url if news.video_file else None,
            'tags': [{'id': tag.id, 'name': tag.name} for tag in news.tags.all()],
            'images': [{
                'id': img.id,
                'image': img.image.url if img.image else None,
                'thumbnail': img.thumbnail.url if img.thumbnail else None,
                'caption': img.caption,
                'alt_text': img.alt_text,
                'is_featured': img.is_featured,
                'order': img.order
            } for img in news.images.all().order_by('order', 'created_at')],
            'created_at': news.created_at.isoformat(),
            'updated_at': news.updated_at.isoformat()
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def news_create_api(request):
    """API to create new news"""
    import base64
    import traceback
    from django.core.files.base import ContentFile
    
    try:
        # Handle both JSON and form data
        if request.content_type and 'application/json' in request.content_type:
            data = json.loads(request.body)
        else:
            # Handle form data from modal
            data = {
                'title': request.POST.get('title', ''),
                'category': request.POST.get('category', ''),
                'content': request.POST.get('content', ''),
                'excerpt': request.POST.get('excerpt', ''),
                'status': request.POST.get('status', 'draft'),
                'priority': request.POST.get('priority', 'normal'),
                'is_featured': request.POST.get('is_featured') == 'on',
                'is_breaking': request.POST.get('is_breaking') == 'on',
            }
            
        print(f"Parsed data keys: {list(data.keys())}")
        print(f"Title: {data.get('title')}")
        print(f"Category: {data.get('category')}")
        
        # Parse category from frontend format
        category_id = data.get('category') or data.get('category_id')
        if not category_id:
            print("Error: No category provided")
            return JsonResponse({'error': 'Kategori harus dipilih'}, status=400)
        
        # Parse tags from frontend format
        if 'tag_names' in data:
            tag_names = data['tag_names']
        else:
            tags_str = data.get('tags', '')
            tag_names = [tag.strip() for tag in tags_str.split(',') if tag.strip()] if tags_str else []
        print(f"Processed tag names: {tag_names}")
        
        # Create news with unique slug
        from django.utils.text import slugify
        title = data.get('title', '')
        if not title:
            return JsonResponse({'error': 'Judul harus diisi'}, status=400)
        
        # Generate unique slug
        base_slug = slugify(title)
        slug = base_slug
        counter = 1
        while News.objects.filter(slug=slug).exists():
            slug = f"{base_slug}-{counter}"
            counter += 1
        print(f"Generated unique slug: {slug}")
            
        news = News.objects.create(
            title=title,
            slug=slug,
            category_id=category_id,
            content=data.get('content', ''),
            excerpt=data.get('excerpt', ''),
            status=data.get('status', 'draft'),
            priority=data.get('priority', 'normal'),
            is_featured=data.get('is_featured', False),
            is_breaking=data.get('is_breaking', False),
            published_date=None,  # Will be set after creation with proper timezone handling
            author=request.user,
            meta_title=data.get('meta_title', title),
            meta_description=data.get('meta_description', ''),
            youtube_url=data.get('youtube_url', ''),
        )
        
        # Handle published_date properly with timezone awareness
        if data.get('publish_date'):
            from django.utils.dateparse import parse_datetime
            from django.utils import timezone
            parsed_date = parse_datetime(data['publish_date'])
            if parsed_date:
                # Make timezone aware if it's naive
                if timezone.is_naive(parsed_date):
                    news.published_date = timezone.make_aware(parsed_date)
                else:
                    news.published_date = parsed_date
                news.save(update_fields=['published_date'])
        
        # Handle tags - create new tags if they don't exist
        if tag_names:
            tag_objects = []
            for tag_name in tag_names:
                if tag_name.strip():
                    # First try to get existing tag by name
                    try:
                        tag = NewsTag.objects.get(name=tag_name.strip())
                        tag_objects.append(tag)
                        print(f"Found existing tag: {tag_name}")
                    except NewsTag.DoesNotExist:
                        # Create new tag with unique slug
                        from django.utils.text import slugify
                        base_slug = slugify(tag_name.strip())
                        slug = base_slug
                        counter = 1
                        
                        # Ensure slug is unique
                        while NewsTag.objects.filter(slug=slug).exists():
                            slug = f"{base_slug}-{counter}"
                            counter += 1
                        
                        tag = NewsTag.objects.create(
                            name=tag_name.strip(),
                            slug=slug
                        )
                        tag_objects.append(tag)
            news.tags.set(tag_objects)
            print(f"Added {len(tag_objects)} tags to news")
        
        # Handle multiple images
        if 'images' in data and data['images']:
            featured_image_id = data.get('featured_image_id')
            
            for index, image_data in enumerate(data['images']):
                # Process base64 image data
                if 'data_url' in image_data:
                    try:
                        # Parse data URL
                        format, imgstr = image_data['data_url'].split(';base64,')
                        ext = format.split('/')[-1]
                        
                        # Generate filename
                        filename = f"news_{news.id}_{image_data['id']}.{ext}"
                        
                        # Decode and save image
                        image_file = ContentFile(
                            base64.b64decode(imgstr),
                            name=filename
                        )
                        
                        # Create NewsImage instance
                        news_image = NewsImage.objects.create(
                            news=news,
                            image=image_file,
                            alt_text=image_data.get('alt_text', ''),
                            caption=image_data.get('name', ''),
                            is_featured=(image_data['id'] == featured_image_id),
                            order=index
                        )
                        
                    except Exception as img_error:
                        print(f"Error processing image {image_data.get('name', '')}: {img_error}")
                        continue
        
        return JsonResponse({
            'success': True,
            'message': 'Berita berhasil ditambahkan',
            'id': news.id,
            'title': news.title
        })
        
    except Exception as e:
        print(f"Error in news_create_api: {e}")
        print(f"Full traceback: {traceback.format_exc()}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["PUT", "POST"])
def news_update_api(request, pk):
    """API to update news"""
    try:
        news = get_object_or_404(News, pk=pk)
        
        # Handle both JSON and form data
        if request.content_type and 'application/json' in request.content_type:
            data = json.loads(request.body)
        else:
            # Handle form data from modal
            data = {
                'title': request.POST.get('title', ''),
                'category': request.POST.get('category', ''),
                'content': request.POST.get('content', ''),
                'excerpt': request.POST.get('excerpt', ''),
                'status': request.POST.get('status', 'draft'),
                'priority': request.POST.get('priority', 'normal'),
                'is_featured': request.POST.get('is_featured') == 'on',
                'is_breaking': request.POST.get('is_breaking') == 'on',
            }
        
        print(f"DEBUG: Received data: {data}")
        
        # Validate required fields
        if not data.get('title'):
            return JsonResponse({'error': 'Title is required'}, status=400)
        category_id = data.get('category') or data.get('category_id')
        if not category_id:
            return JsonResponse({'error': 'Category is required'}, status=400)
        if not data.get('content'):
            return JsonResponse({'error': 'Content is required'}, status=400)
        
        news.title = data['title']
        news.category_id = category_id
        news.content = data['content']
        news.excerpt = data.get('excerpt', '')
        news.status = data.get('status', 'draft')
        news.priority = data.get('priority', 'normal')
        news.is_featured = data.get('is_featured', False)
        news.is_breaking = data.get('is_breaking', False)
        
        # Handle featured image upload
        if 'featured_image' in request.FILES:
            news.featured_image = request.FILES['featured_image']
        
        # Handle published_date properly with timezone awareness
        if data.get('published_date'):
            from django.utils.dateparse import parse_datetime
            from django.utils import timezone
            parsed_date = parse_datetime(data['published_date'])
            if parsed_date:
                # Make timezone aware if it's naive
                if timezone.is_naive(parsed_date):
                    news.published_date = timezone.make_aware(parsed_date)
                else:
                    news.published_date = parsed_date
            else:
                news.published_date = None
        
        news.meta_title = data.get('meta_title', '')
        news.meta_description = data.get('meta_description', '')
        news.youtube_url = data.get('youtube_url', '')
        news.save()
        
        # Handle tags - create new tags if they don't exist
        if 'tag_names' in data:
            tag_names = data['tag_names']
            tag_objects = []
            for tag_name in tag_names:
                if tag_name.strip():
                    # First try to get existing tag by name
                    try:
                        tag = NewsTag.objects.get(name=tag_name.strip())
                        tag_objects.append(tag)
                    except NewsTag.DoesNotExist:
                        # Create new tag with unique slug
                        from django.utils.text import slugify
                        base_slug = slugify(tag_name.strip())
                        slug = base_slug
                        counter = 1
                        
                        # Ensure slug is unique
                        while NewsTag.objects.filter(slug=slug).exists():
                            slug = f"{base_slug}-{counter}"
                            counter += 1
                        
                        tag = NewsTag.objects.create(
                            name=tag_name.strip(),
                            slug=slug
                        )
                        tag_objects.append(tag)
            news.tags.set(tag_objects)
        elif 'tag_ids' in data:
            news.tags.set(data['tag_ids'])
        
        return JsonResponse({
            'success': True,
            'message': 'Berita berhasil diperbarui'
        })
        
    except Exception as e:
        print(f"DEBUG: Error in news_update_api: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def news_delete_api(request, pk):
    """API to delete news"""
    try:
        news = get_object_or_404(News, pk=pk)
        news.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Berita berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def news_duplicate_api(request, pk):
    """API to duplicate news"""
    try:
        original_news = get_object_or_404(News, pk=pk)
        
        # Create a copy of the news
        duplicated_news = News.objects.create(
            title=f"{original_news.title} (Copy)",
            slug=slugify(f"{original_news.title} (Copy)"),
            category=original_news.category,
            content=original_news.content,
            excerpt=original_news.excerpt,
            status='draft',
            priority=original_news.priority,
            is_featured=False,
            is_breaking=False,
            published_date=None,
            author=request.user,
            meta_title=original_news.meta_title,
            meta_description=original_news.meta_description,
            youtube_url=original_news.youtube_url,
        )
        
        # Copy tags
        duplicated_news.tags.set(original_news.tags.all())
        
        return JsonResponse({
            'success': True,
            'message': 'Berita berhasil diduplikasi',
            'id': duplicated_news.id
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# News Comment API Views
@login_required
@require_http_methods(["GET"])
def news_comment_list_api(request):
    """API to get list of news comments"""
    try:
        search = request.GET.get('search', '')
        news_id = request.GET.get('news_id', '')
        status = request.GET.get('status', '')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        
        queryset = NewsComment.objects.select_related('news')
        
        if search:
            queryset = queryset.filter(
                Q(author_name__icontains=search) |
                Q(content__icontains=search) |
                Q(news__title__icontains=search)
            )
        
        if news_id:
            queryset = queryset.filter(news_id=news_id)
        
        if status:
            queryset = queryset.filter(status=status)
        
        paginator = Paginator(queryset, per_page)
        page_obj = paginator.get_page(page)
        
        data = {
            'results': [
                {
                    'id': item.id,
                    'news_title': item.news.title,
                    'author_name': item.author_name,
                    'author_email': item.author_email,
                    'content': item.content[:100] + '...' if len(item.content) > 100 else item.content,
                    'status': item.status,
                    'is_author_verified': item.is_author_verified,
                    'created_at': item.created_at.strftime('%d/%m/%Y %H:%M'),
                    'parent_id': item.parent.id if item.parent else None
                }
                for item in page_obj
            ],
            'pagination': {
                'current_page': page,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def news_comment_update_status_api(request, pk):
    """API to update comment status"""
    try:
        comment = get_object_or_404(NewsComment, pk=pk)
        data = json.loads(request.body)
        
        comment.status = data['status']
        comment.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Status komentar berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def news_comment_delete_api(request, pk):
    """API to delete comment"""
    try:
        comment = get_object_or_404(NewsComment, pk=pk)
        comment.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Komentar berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Statistics API
@login_required
@require_http_methods(["GET"])
def news_statistics_api(request):
    """API to get news statistics"""
    try:
        total_news = News.objects.count()
        published_news = News.objects.filter(status='published').count()
        draft_news = News.objects.filter(status='draft').count()
        total_categories = NewsCategory.objects.filter(is_active=True).count()
        total_tags = NewsTag.objects.count()
        total_comments = NewsComment.objects.count()
        pending_comments = NewsComment.objects.filter(status='pending').count()
        total_views = NewsView.objects.count()
        
        data = {
            'total_news': total_news,
            'published_news': published_news,
            'draft_news': draft_news,
            'total_categories': total_categories,
            'total_tags': total_tags,
            'total_comments': total_comments,
            'pending_comments': pending_comments,
            'total_views': total_views
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Helper APIs for dropdowns
@login_required
@require_http_methods(["GET"])
def news_categories_dropdown_api(request):
    """API to get categories for dropdown"""
    try:
        categories = NewsCategory.objects.filter(is_active=True).values('id', 'name', 'color')
        return JsonResponse({'results': list(categories)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def news_tags_dropdown_api(request):
    """API to get tags for dropdown"""
    try:
        tags = NewsTag.objects.values('id', 'name')
        return JsonResponse({'results': list(tags)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def news_dropdown_api(request):
    """API to get news for dropdown"""
    try:
        news = News.objects.filter(status='published').values('id', 'title')
        return JsonResponse({'results': list(news)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Template Views
@login_required
def news_create_view(request):
    """View for create news template"""
    print(f"=== DEBUG: news_create_view called with method: {request.method} ===")
    print(f"User: {request.user}")
    print(f"User authenticated: {request.user.is_authenticated}")
    
    if request.method == 'POST':
        print("=== DEBUG: POST request received ===")
        print(f"POST data: {request.POST}")
        print(f"FILES data: {request.FILES}")
        
        form = NewsForm(request.POST, request.FILES)
        # Skip image formset validation since template doesn't use it
        # image_formset = NewsImageFormSet(request.POST, request.FILES)
        
        print(f"Form is valid: {form.is_valid()}")
        # print(f"Image formset is valid: {image_formset.is_valid()}")
        
        if not form.is_valid():
            print(f"Form errors: {form.errors}")
        # if not image_formset.is_valid():
        #     print(f"Image formset errors: {image_formset.errors}")
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    news = form.save(commit=False)
                    news.author = request.user
                    news.slug = slugify(news.title)
                    
                    print(f"About to save news: {news.title}")
                    news.save()
                    print(f"News saved with ID: {news.pk}")
                    
                    form.save_m2m()  # Save many-to-many relationships (tags)
                    print("Many-to-many relationships saved")
                    
                    # Skip image formset saving since template doesn't use it
                    # image_formset.instance = news
                    # image_formset.save()
                    # print("Images saved")
                    
                    messages.success(request, f'Berita "{news.title}" berhasil dibuat.')
                    return redirect('news:news_view_detail', pk=news.pk)
            except Exception as e:
                print(f"Error saving news: {e}")
                import traceback
                print(f"Full traceback: {traceback.format_exc()}")
                messages.error(request, f'Terjadi kesalahan saat menyimpan berita: {str(e)}')
    else:
        form = NewsForm()
        # image_formset = NewsImageFormSet()
    
    context = {
        'page_title': 'Tambah Berita',
        'page_subtitle': 'Buat berita baru',
        'form': form,
        # 'image_formset': image_formset,
        'action': 'create',
    }
    return render(request, 'admin/modules/news/create.html', context)


@login_required
def news_edit_view(request, pk):
    """View for edit news template"""
    news = get_object_or_404(News, pk=pk)
    context = {
        'page_title': 'Edit Berita',
        'page_subtitle': f'Edit berita: {news.title}',
        'news': news,
    }
    return render(request, 'admin/modules/news/edit.html', context)


@login_required
def news_view_detail(request, pk):
    """View for news detail template"""
    news = get_object_or_404(News, pk=pk)
    context = {
        'page_title': 'Detail Berita',
        'page_subtitle': news.title,
        'news': news,
    }
    return render(request, 'admin/modules/news/view.html', context)


@login_required
def news_list_view(request):
    """View for news list template with data"""
    # Get all news with related data
    news_list = News.objects.select_related('category', 'author').prefetch_related('tags', 'likes', 'comments').order_by('-created_at')
    
    # Get all categories for filter dropdown
    categories = NewsCategory.objects.all().order_by('name')
    
    # Apply filters
    search = request.GET.get('search')
    category_id = request.GET.get('category')
    status = request.GET.get('status')
    
    if search:
        news_list = news_list.filter(
            Q(title__icontains=search) | 
            Q(content__icontains=search) |
            Q(excerpt__icontains=search)
        )
    
    if category_id:
        news_list = news_list.filter(category_id=category_id)
    
    if status:
        news_list = news_list.filter(status=status)
    
    # Pagination
    paginator = Paginator(news_list, 10)  # Show 10 news per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': 'Daftar Berita',
        'page_subtitle': 'Kelola semua berita',
        'news_list': page_obj,
        'categories': categories,
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
    }
    return render(request, 'admin/modules/news/list.html', context)


@login_required
def news_category_view(request):
    """View for news category template"""
    context = {
        'page_title': 'Kategori Berita',
        'page_subtitle': 'Kelola kategori berita',
    }
    return render(request, 'admin/modules/news/category.html', context)


@login_required
def news_category_create_view(request):
    """View for news category create template"""
    context = {
        'page_title': 'Tambah Kategori Berita',
        'page_subtitle': 'Buat kategori berita baru',
    }
    return render(request, 'admin/modules/news/category.html', context)


@login_required
def news_media_view(request):
    """View for news media template"""
    context = {
        'page_title': 'Galeri Media',
        'page_subtitle': 'Kelola media berita',
    }
    return render(request, 'admin/modules/news/media.html', context)


# Media API Views
@login_required
@require_http_methods(["GET"])
def news_media_list_api(request):
    """API to get list of news media"""
    try:
        search = request.GET.get('search', '')
        news_id = request.GET.get('news_id', '')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        
        queryset = NewsImage.objects.select_related('news')
        
        if search:
            queryset = queryset.filter(
                Q(caption__icontains=search) |
                Q(alt_text__icontains=search) |
                Q(news__title__icontains=search)
            )
        
        if news_id:
            queryset = queryset.filter(news_id=news_id)
        
        paginator = Paginator(queryset, per_page)
        page_obj = paginator.get_page(page)
        
        data = {
            'results': [
                {
                    'id': item.id,
                    'news_id': item.news.id,
                    'news_title': item.news.title,
                    'image_url': item.image.url if item.image else None,
                    'thumbnail_url': item.thumbnail.url if item.thumbnail else None,
                    'caption': item.caption,
                    'alt_text': item.alt_text,
                    'is_featured': item.is_featured,
                    'order': item.order,
                    'file_size': item.image.size if item.image else 0,
                    'created_at': item.created_at.strftime('%d/%m/%Y %H:%M')
                }
                for item in page_obj
            ],
            'pagination': {
                'current_page': page,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def news_media_upload_api(request):
    """API to upload news media"""
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)
        
        file = request.FILES['file']
        news_id = request.POST.get('news_id')
        caption = request.POST.get('caption', '')
        alt_text = request.POST.get('alt_text', '')
        is_featured = request.POST.get('is_featured', 'false').lower() == 'true'
        
        # Check if this is a video upload request
        if 'video_file' in request.FILES:
            return news_video_upload_api(request)
        
        # Validate file type for images
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp']
        if file.content_type not in allowed_types:
            return JsonResponse({'error': 'Invalid file type. Only images are allowed.'}, status=400)
        
        # Validate file size (10MB max)
        if file.size > 10 * 1024 * 1024:
            return JsonResponse({'error': 'File too large. Maximum size is 10MB.'}, status=400)
        
        # Create NewsImage instance
        news_image = NewsImage(
            image=file,
            caption=caption,
            alt_text=alt_text,
            is_featured=is_featured
        )
        
        if news_id:
            news = get_object_or_404(News, pk=news_id)
            news_image.news = news
        
        news_image.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Media berhasil diupload',
            'media': {
                'id': news_image.id,
                'image_url': news_image.image.url,
                'thumbnail_url': news_image.thumbnail.url if news_image.thumbnail else None,
                'caption': news_image.caption,
                'alt_text': news_image.alt_text,
                'is_featured': news_image.is_featured,
                'file_size': news_image.image.size,
                'created_at': news_image.created_at.strftime('%d/%m/%Y %H:%M')
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def news_gallery_upload_api(request):
    """API to upload multiple gallery images for news"""
    try:
        import json
        import base64
        from django.core.files.base import ContentFile
        
        data = json.loads(request.body)
        news_id = data.get('news_id')
        gallery_images = data.get('gallery_images', [])
        
        if not news_id:
            return JsonResponse({'error': 'News ID is required'}, status=400)
        
        if not gallery_images:
            return JsonResponse({'error': 'No images provided'}, status=400)
        
        news = get_object_or_404(News, pk=news_id)
        uploaded_images = []
        
        for index, image_data in enumerate(gallery_images):
            try:
                # Process base64 image data
                if 'dataUrl' in image_data:
                    # Parse data URL
                    format, imgstr = image_data['dataUrl'].split(';base64,')
                    ext = format.split('/')[-1]
                    
                    # Generate filename
                    filename = f"news_{news_id}_{image_data['id']}.{ext}"
                    
                    # Decode and save image
                    image_file = ContentFile(
                        base64.b64decode(imgstr),
                        name=filename
                    )
                    
                    # Create NewsImage instance
                    news_image = NewsImage.objects.create(
                        news=news,
                        image=image_file,
                        alt_text=image_data.get('altText', ''),
                        caption=image_data.get('name', ''),
                        is_featured=False,  # Will be set separately if needed
                        order=index
                    )
                    
                    uploaded_images.append({
                        'id': news_image.id,
                        'image_url': news_image.image.url,
                        'thumbnail_url': news_image.thumbnail.url if news_image.thumbnail else None,
                        'caption': news_image.caption,
                        'alt_text': news_image.alt_text,
                        'order': news_image.order
                    })
                    
            except Exception as img_error:
                print(f"Error processing image {image_data.get('name', '')}: {img_error}")
                continue
        
        return JsonResponse({
            'success': True,
            'message': f'{len(uploaded_images)} gambar berhasil diupload',
            'images': uploaded_images
        })
        
    except Exception as e:
        print(f"Error in news_gallery_upload_api: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def news_video_upload_api(request):
    """API to upload news video"""
    try:
        if 'video_file' not in request.FILES:
            return JsonResponse({'error': 'No video file provided'}, status=400)
        
        video_file = request.FILES['video_file']
        news_id = request.POST.get('news_id')
        
        # Validate file type
        allowed_video_types = ['video/mp4', 'video/webm', 'video/ogg', 'video/avi', 'video/mov', 'video/quicktime']
        if video_file.content_type not in allowed_video_types:
            return JsonResponse({'error': 'Invalid file type. Only video files are allowed.'}, status=400)
        
        # Validate file size (50MB max)
        if video_file.size > 50 * 1024 * 1024:
            return JsonResponse({'error': 'File too large. Maximum size is 50MB.'}, status=400)
        
        # Update news with video file
        if news_id:
            news = get_object_or_404(News, pk=news_id)
            news.video_file = video_file
            news.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Video berhasil diupload',
                'video': {
                    'url': news.video_file.url,
                    'name': video_file.name,
                    'size': video_file.size,
                    'type': video_file.content_type
                }
            })
        else:
            return JsonResponse({'error': 'News ID is required'}, status=400)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def news_media_detail_api(request, pk):
    """API to get media detail"""
    try:
        media = get_object_or_404(NewsImage, pk=pk)
        
        data = {
            'id': media.id,
            'news_id': media.news.id if media.news else None,
            'news_title': media.news.title if media.news else None,
            'image_url': media.image.url if media.image else None,
            'thumbnail_url': media.thumbnail.url if media.thumbnail else None,
            'caption': media.caption,
            'alt_text': media.alt_text,
            'is_featured': media.is_featured,
            'order': media.order,
            'file_size': media.image.size if media.image else 0,
            'created_at': media.created_at.strftime('%d/%m/%Y %H:%M')
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def news_media_update_api(request, pk):
    """API to update media"""
    try:
        media = get_object_or_404(NewsImage, pk=pk)
        data = json.loads(request.body)
        
        media.caption = data.get('caption', media.caption)
        media.alt_text = data.get('alt_text', media.alt_text)
        media.is_featured = data.get('is_featured', media.is_featured)
        media.order = data.get('order', media.order)
        
        if 'news_id' in data and data['news_id']:
            news = get_object_or_404(News, pk=data['news_id'])
            media.news = news
        
        media.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Media berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def news_media_delete_api(request, pk):
    """API to delete media"""
    try:
        media = get_object_or_404(NewsImage, pk=pk)
        
        # Delete the actual file
        if media.image:
            default_storage.delete(media.image.name)
        if media.thumbnail:
            default_storage.delete(media.thumbnail.name)
        
        media.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Media berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def news_media_bulk_action_api(request):
    """API for bulk media actions"""
    try:
        data = json.loads(request.body)
        action = data.get('action')
        media_ids = data.get('media_ids', [])
        
        if not media_ids:
            return JsonResponse({'error': 'No media selected'}, status=400)
        
        media_queryset = NewsImage.objects.filter(id__in=media_ids)
        
        if action == 'delete':
            # Delete files and records
            for media in media_queryset:
                if media.image:
                    default_storage.delete(media.image.name)
                if media.thumbnail:
                    default_storage.delete(media.thumbnail.name)
            
            deleted_count = media_queryset.count()
            media_queryset.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'{deleted_count} media berhasil dihapus'
            })
        
        elif action == 'set_featured':
            media_queryset.update(is_featured=True)
            return JsonResponse({
                'success': True,
                'message': f'{media_queryset.count()} media berhasil dijadikan unggulan'
            })
        
        elif action == 'unset_featured':
            media_queryset.update(is_featured=False)
            return JsonResponse({
                'success': True,
                'message': f'{media_queryset.count()} media berhasil dihapus dari unggulan'
            })
        
        else:
            return JsonResponse({'error': 'Invalid action'}, status=400)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def news_media_statistics_api(request):
    """API to get media statistics"""
    try:
        total_media = NewsImage.objects.count()
        featured_media = NewsImage.objects.filter(is_featured=True).count()
        
        # Calculate total file size
        total_size = 0
        for media in NewsImage.objects.all():
            if media.image:
                total_size += media.image.size
        
        # Recent uploads (last 7 days)
        from django.utils import timezone
        from datetime import timedelta
        recent_uploads = NewsImage.objects.filter(
            created_at__gte=timezone.now() - timedelta(days=7)
        ).count()
        
        data = {
            'total_media': total_media,
            'featured_media': featured_media,
            'total_size': total_size,
            'recent_uploads': recent_uploads
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Admin-specific Views (Non-API)
@login_required
def news_duplicate_view(request, pk):
    """Duplicate existing news"""
    news = get_object_or_404(News, pk=pk)
    
    if request.method == 'POST':
        # Create a copy of the news
        news.pk = None
        news.title = f"{news.title} (Copy)"
        news.slug = slugify(f"{news.title} (Copy)")
        news.status = 'draft'
        news.is_featured = False
        news.is_breaking = False
        news.published_date = None
        news.scheduled_date = None
        news.views_count = 0
        news.likes_count = 0
        news.comments_count = 0
        news.shares_count = 0
        news.save()
        
        # Copy tags
        news.tags.set(news.tags.all())
        
        messages.success(request, f'Berita "{news.title}" berhasil diduplikasi.')
        return redirect('news:admin_edit', pk=news.pk)
    
    context = {
        'news': news,
        'page_title': 'Duplikasi Berita',
        'page_subtitle': f'Duplikasi: {news.title}',
    }
    return render(request, 'admin/modules/news/duplicate.html', context)


@login_required
def news_delete_view(request, pk):
    """Delete news view"""
    news = get_object_or_404(News, pk=pk)
    
    if request.method == 'POST':
        title = news.title
        news.delete()
        messages.success(request, f'Berita "{title}" berhasil dihapus.')
        return redirect('news:admin_index')
    
    context = {
        'news': news,
        'page_title': 'Hapus Berita',
        'page_subtitle': f'Hapus: {news.title}',
    }
    return render(request, 'admin/modules/news/delete.html', context)


@login_required
@require_POST
def news_publish_view(request, pk):
    """Publish news"""
    news = get_object_or_404(News, pk=pk)
    
    # Check if it's an AJAX request
    if request.headers.get('Content-Type') == 'application/json':
        try:
            news.status = 'published'
            news.published_date = timezone.now()
            news.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Berita "{news.title}" berhasil dipublikasikan.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Gagal mempublikasikan berita: {str(e)}'
            }, status=500)
    else:
        # Handle regular form submission
        news.status = 'published'
        news.published_date = timezone.now()
        news.save()
        
        messages.success(request, f'Berita "{news.title}" berhasil dipublikasikan.')
        return redirect('news:admin_view', pk=news.pk)


@login_required
@require_POST
def news_unpublish_view(request, pk):
    """Unpublish news"""
    news = get_object_or_404(News, pk=pk)
    
    # Check if it's an AJAX request
    if request.headers.get('Content-Type') == 'application/json':
        try:
            news.status = 'draft'
            news.published_date = None
            news.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Berita "{news.title}" berhasil ditarik dari publikasi.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Gagal menarik publikasi berita: {str(e)}'
            }, status=500)
    else:
        # Handle regular form submission
        news.status = 'draft'
        news.published_date = None
        news.save()
        
        messages.success(request, f'Berita "{news.title}" berhasil ditarik dari publikasi.')
        return redirect('news:admin_view', pk=news.pk)


@login_required
def news_bulk_action_view(request):
    """Handle bulk actions on news"""
    if request.method == 'POST':
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_news')
        
        if not selected_ids:
            messages.error(request, 'Tidak ada berita yang dipilih.')
            return redirect('news:admin_index')
        
        try:
            news_queryset = News.objects.filter(id__in=selected_ids)
            
            if action == 'publish':
                news_queryset.update(status='published', published_date=timezone.now())
                messages.success(request, f'{news_queryset.count()} berita berhasil dipublikasikan.')
            
            elif action == 'draft':
                news_queryset.update(status='draft')
                messages.success(request, f'{news_queryset.count()} berita berhasil dijadikan draft.')
            
            elif action == 'archive':
                news_queryset.update(status='archived')
                messages.success(request, f'{news_queryset.count()} berita berhasil diarsipkan.')
            
            elif action == 'delete':
                count = news_queryset.count()
                news_queryset.delete()
                messages.success(request, f'{count} berita berhasil dihapus.')
            
            else:
                messages.error(request, 'Aksi tidak valid.')
        
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
    
    return redirect('news:admin_index')


# Media Management Views (Non-API)
@login_required
def news_media_upload_view(request):
    """Upload media files"""
    if request.method == 'POST':
        files = request.FILES.getlist('files')
        
        if not files:
            messages.error(request, 'Tidak ada file yang dipilih.')
            return redirect('news:news_media_view')
        
        uploaded_count = 0
        for file in files:
            try:
                # Create NewsImage instance
                news_image = NewsImage(
                    image=file,
                    caption=file.name,
                    alt_text=file.name
                )
                news_image.save()
                uploaded_count += 1
            except Exception as e:
                messages.error(request, f'Gagal mengupload {file.name}: {str(e)}')
        
        if uploaded_count > 0:
            messages.success(request, f'{uploaded_count} file berhasil diupload.')
        
        return redirect('news:news_media_view')
    
    context = {
        'page_title': 'Upload Media',
        'page_subtitle': 'Upload file media baru',
    }
    return render(request, 'admin/modules/news/media_upload.html', context)


@login_required
def news_media_detail_view(request, pk):
    """View media detail"""
    media = get_object_or_404(NewsImage, pk=pk)
    
    context = {
        'media': media,
        'page_title': 'Detail Media',
        'page_subtitle': media.caption or 'Media Detail',
    }
    return render(request, 'admin/modules/news/media_detail.html', context)


@login_required
def news_media_update_view(request, pk):
    """Update media"""
    media = get_object_or_404(NewsImage, pk=pk)
    
    if request.method == 'POST':
        media.caption = request.POST.get('title', '')
        media.alt_text = request.POST.get('alt_text', '')
        media.save()
        
        messages.success(request, 'Media berhasil diupdate.')
        return redirect('news:news_media_view')
    
    context = {
        'media': media,
        'page_title': 'Edit Media',
        'page_subtitle': f'Edit: {media.caption or "Media"}',
    }
    return render(request, 'admin/modules/news/media_edit.html', context)


@login_required
@require_POST
def news_media_delete_view(request, pk):
    """Delete media"""
    media = get_object_or_404(NewsImage, pk=pk)
    
    # Delete the actual file
    if media.image:
        default_storage.delete(media.image.name)
    if media.thumbnail:
        default_storage.delete(media.thumbnail.name)
    
    media.delete()
    messages.success(request, 'Media berhasil dihapus.')
    
    return redirect('news:news_media_view')


@login_required
@require_POST
def news_media_bulk_delete_view(request):
    """Bulk delete media"""
    media_ids = request.POST.getlist('media_ids')
    
    if not media_ids:
        messages.error(request, 'Tidak ada media yang dipilih.')
        return redirect('news:news_media_view')
    
    try:
        media_queryset = NewsImage.objects.filter(id__in=media_ids)
        
        # Delete files
        for media in media_queryset:
            if media.image:
                default_storage.delete(media.image.name)
            if media.thumbnail:
                default_storage.delete(media.thumbnail.name)
        
        deleted_count = media_queryset.count()
        media_queryset.delete()
        
        messages.success(request, f'{deleted_count} media berhasil dihapus.')
    
    except Exception as e:
        messages.error(request, f'Terjadi kesalahan: {str(e)}')
    
    return redirect('news:news_media_view')


@login_required
@require_POST
def news_media_bulk_download_view(request):
    """Bulk download media"""
    media_ids = request.POST.getlist('media_ids')
    
    if not media_ids:
        messages.error(request, 'Tidak ada media yang dipilih.')
        return redirect('news:news_media_view')
    
    # This would typically create a zip file for download
    # For now, just redirect back
    messages.info(request, 'Fitur bulk download akan segera tersedia.')
    return redirect('news:news_media_view')


# Category Management Views (Non-API)
@login_required
def news_category_detail_view(request, pk):
    """View category detail"""
    category = get_object_or_404(NewsCategory, pk=pk)
    
    # Get recent news in this category
    recent_news = category.news.all()[:5]
    
    context = {
        'category': category,
        'recent_news': recent_news,
        'page_title': 'Detail Kategori',
        'page_subtitle': category.name,
    }
    return render(request, 'admin/modules/news/category_detail.html', context)


@login_required
def news_category_update_view(request, pk):
    """Update category"""
    category = get_object_or_404(NewsCategory, pk=pk)
    
    if request.method == 'POST':
        form = NewsCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, f'Kategori "{category.name}" berhasil diupdate.')
            return redirect('news:admin_categories')
    else:
        form = NewsCategoryForm(instance=category)
    
    context = {
        'form': form,
        'category': category,
        'page_title': 'Edit Kategori',
        'page_subtitle': f'Edit: {category.name}',
    }
    return render(request, 'admin/modules/news/category_edit.html', context)


@login_required
@require_POST
def news_category_delete_view(request, pk):
    """Delete category"""
    category = get_object_or_404(NewsCategory, pk=pk)
    
    # Check if category has news
    if category.news.exists():
        messages.error(request, 'Kategori tidak dapat dihapus karena masih memiliki berita.')
        return redirect('news:admin_categories')
    
    name = category.name
    category.delete()
    messages.success(request, f'Kategori "{name}" berhasil dihapus.')
    
    return redirect('news:admin_categories')


# ============ ANNOUNCEMENT VIEWS ============

@login_required
def announcement_list(request):
    """List all announcements"""
    search_form = AnnouncementSearchForm(request.GET or None)
    announcements = Announcement.objects.all().order_by('-created_at')
    
    # Apply search filters
    if search_form.is_valid():
        search_query = search_form.cleaned_data.get('search')
        announcement_type = search_form.cleaned_data.get('type')
        status = search_form.cleaned_data.get('status')
        priority = search_form.cleaned_data.get('priority')
        
        if search_query:
            announcements = announcements.filter(
                Q(title__icontains=search_query) |
                Q(content__icontains=search_query)
            )
        
        if announcement_type:
            announcements = announcements.filter(type=announcement_type)
            
        if status:
            announcements = announcements.filter(status=status)
            
        if priority:
            announcements = announcements.filter(priority=priority)
    
    # Pagination
    paginator = Paginator(announcements, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'announcements': page_obj,
        'search_form': search_form,
        'total_announcements': announcements.count(),
    }
    
    return render(request, 'admin/modules/news/announcements.html', context)


@login_required
def announcement_create(request):
    """Create new announcement"""
    if request.method == 'POST':
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            announcement = form.save(commit=False)
            announcement.author = request.user
            announcement.save()
            messages.success(request, 'Pengumuman berhasil dibuat.')
            return JsonResponse({
                'success': True,
                'message': 'Pengumuman berhasil dibuat.',
                'redirect_url': reverse('news:announcement_list')
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    else:
        form = AnnouncementForm()
    
    context = {
        'form': form,
        'title': 'Tambah Pengumuman'
    }
    
    return render(request, 'admin/modules/news/announcement_form.html', context)


@login_required
def announcement_edit(request, pk):
    """Edit announcement"""
    announcement = get_object_or_404(Announcement, pk=pk)
    
    if request.method == 'POST':
        form = AnnouncementForm(request.POST, instance=announcement)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pengumuman berhasil diperbarui.')
            return JsonResponse({
                'success': True,
                'message': 'Pengumuman berhasil diperbarui.',
                'redirect_url': reverse('news:announcement_list')
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    else:
        form = AnnouncementForm(instance=announcement)
    
    context = {
        'form': form,
        'announcement': announcement,
        'title': 'Edit Pengumuman'
    }
    
    return render(request, 'admin/modules/news/announcement_form.html', context)


@login_required
@require_POST
def announcement_delete(request, pk):
    """Delete announcement"""
    announcement = get_object_or_404(Announcement, pk=pk)
    title = announcement.title
    announcement.delete()
    messages.success(request, f'Pengumuman "{title}" berhasil dihapus.')
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': f'Pengumuman "{title}" berhasil dihapus.'
        })
    
    return redirect('news:announcement_list')


@login_required
def announcement_detail(request, pk):
    """View announcement detail"""
    announcement = get_object_or_404(Announcement, pk=pk)
    
    context = {
        'announcement': announcement
    }
    
    return render(request, 'admin/modules/news/announcement_detail.html', context)


# Public announcement views
def public_announcement_list(request):
    """Public announcement list"""
    announcements = Announcement.objects.filter(
        status='published',
        start_date__lte=timezone.now()
    ).filter(
        Q(end_date__isnull=True) | Q(end_date__gte=timezone.now())
    ).order_by('-priority', '-created_at')
    
    # Filter by type if specified
    announcement_type = request.GET.get('type')
    if announcement_type:
        announcements = announcements.filter(type=announcement_type)
    
    # Pagination
    paginator = Paginator(announcements, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'announcements': page_obj,
        'announcement_types': Announcement.TYPE_CHOICES,
        'current_type': announcement_type
    }
    
    return render(request, 'public/announcements.html', context)


def public_announcement_detail(request, slug):
    """Public announcement detail"""
    announcement = get_object_or_404(
        Announcement,
        slug=slug,
        status='published',
        start_date__lte=timezone.now()
    )
    
    # Check if announcement is still active
    if announcement.end_date and announcement.end_date < timezone.now():
        messages.warning(request, 'Pengumuman ini sudah tidak aktif.')
    
    context = {
        'announcement': announcement
    }
    
    return render(request, 'public/announcement_detail.html', context)


# Public News Views (tanpa autentikasi)
@csrf_exempt
@require_http_methods(["GET"])
def public_news_list(request):
    """Public API endpoint untuk daftar berita tanpa autentikasi"""
    try:
        # Get query parameters
        search = request.GET.get('search', '')
        category_id = request.GET.get('category_id', '')
        limit = int(request.GET.get('limit', 10))
        page = int(request.GET.get('page', 1))
        
        # Base queryset - hanya berita yang published
        queryset = News.objects.filter(
            status='published',
            published_date__lte=timezone.now()
        ).select_related('category', 'author').prefetch_related('tags')
        
        # Apply search filter
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(content__icontains=search) |
                Q(excerpt__icontains=search)
            )
        
        # Apply category filter
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        
        # Order by priority and published date
        queryset = queryset.order_by('-is_featured', '-priority', '-published_date')
        
        # Pagination
        paginator = Paginator(queryset, limit)
        page_obj = paginator.get_page(page)
        
        # Serialize data
        results = []
        for news in page_obj:
            results.append({
                'id': news.id,
                'title': news.title,
                'slug': news.slug,
                'excerpt': news.excerpt,
                'content': news.content[:200] + '...' if len(news.content) > 200 else news.content,
                'featured_image': news.featured_image.url if news.featured_image else None,
                'category': {
                    'id': news.category.id,
                    'name': news.category.name,
                    'color': news.category.color
                } if news.category else None,
                'author': news.author.get_full_name() or news.author.username,
                'published_date': news.published_date.isoformat() if news.published_date else None,
                'is_featured': news.is_featured,
                'is_breaking': news.is_breaking,
                'views_count': news.views_count,
                'likes_count': news.likes_count,
                'tags': [{'id': tag.id, 'name': tag.name} for tag in news.tags.all()],
                'created_at': news.created_at.isoformat(),
                'updated_at': news.updated_at.isoformat()
            })
        
        data = {
            'success': True,
            'results': results,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': page_obj.paginator.num_pages,
                'total_items': page_obj.paginator.count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
                'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
                'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def public_news_detail(request, slug):
    """Public API endpoint untuk detail berita tanpa autentikasi"""
    try:
        news = get_object_or_404(
            News.objects.select_related('category', 'author').prefetch_related('tags', 'images'),
            slug=slug,
            status='published',
            published_date__lte=timezone.now()
        )
        
        # Increment view count
        NewsView.objects.create(
            news=news,
            ip_address=request.META.get('REMOTE_ADDR', ''),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            session_key=request.session.session_key or ''
        )
        
        # Update news views count
        news.views_count = news.view_records.count()
        news.save(update_fields=['views_count'])
        
        data = {
            'success': True,
            'result': {
                'id': news.id,
                'title': news.title,
                'slug': news.slug,
                'excerpt': news.excerpt,
                'content': news.content,
                'featured_image': news.featured_image.url if news.featured_image else None,
                'category': {
                    'id': news.category.id,
                    'name': news.category.name,
                    'color': news.category.color
                } if news.category else None,
                'author': news.author.get_full_name() or news.author.username,
                'published_date': news.published_date.isoformat() if news.published_date else None,
                'is_featured': news.is_featured,
                'is_breaking': news.is_breaking,
                'views_count': news.views_count,
                'likes_count': news.likes_count,
                'meta_title': news.meta_title,
                'meta_description': news.meta_description,
                'youtube_url': news.youtube_url,
                'video_file': news.video_file.url if news.video_file else None,
                'tags': [{'id': tag.id, 'name': tag.name} for tag in news.tags.all()],
                'images': [{
                    'id': img.id,
                    'image': img.image.url if img.image else None,
                    'caption': img.caption,
                    'alt_text': img.alt_text,
                    'is_featured': img.is_featured,
                    'order': img.order
                } for img in news.images.all().order_by('order')],
                'created_at': news.created_at.isoformat(),
                'updated_at': news.updated_at.isoformat()
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def public_featured_news(request):
    """Public API endpoint untuk berita unggulan tanpa autentikasi"""
    try:
        limit = int(request.GET.get('limit', 5))
        
        # Get featured news
        featured_news = News.objects.filter(
            status='published',
            published_date__lte=timezone.now(),
            is_featured=True
        ).select_related('category', 'author').prefetch_related('tags').order_by('-published_date')[:limit]
        
        results = []
        for news in featured_news:
            results.append({
                'id': news.id,
                'title': news.title,
                'slug': news.slug,
                'excerpt': news.excerpt,
                'content': news.content[:200] + '...' if len(news.content) > 200 else news.content,
                'featured_image': news.featured_image.url if news.featured_image else None,
                'category': {
                    'id': news.category.id,
                    'name': news.category.name,
                    'color': news.category.color
                } if news.category else None,
                'author': news.author.get_full_name() or news.author.username,
                'published_date': news.published_date.isoformat() if news.published_date else None,
                'is_featured': news.is_featured,
                'is_breaking': news.is_breaking,
                'views_count': news.views_count,
                'likes_count': news.likes_count,
                'tags': [{'id': tag.id, 'name': tag.name} for tag in news.tags.all()],
                'created_at': news.created_at.isoformat(),
                'updated_at': news.updated_at.isoformat()
            })
        
        data = {
            'success': True,
            'results': results,
            'count': len(results)
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_date
from django.core.files.storage import default_storage
from django.conf import settings
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO

from .models import (
    OrganizationType, Organization, OrganizationMember, 
    OrganizationEvent, OrganizationDocument,
    Jabatan, PeriodeKepengurusan, AnggotaOrganisasi,
    GaleriKegiatan, StrukturOrganisasi
)
from references.models import Penduduk


def organization_admin_view(request):
    """Main organization admin view"""
    context = {
        'page_title': 'Manajemen Organisasi',
        'page_subtitle': 'Kelola data organisasi, anggota, dan kegiatan'
    }
    return render(request, 'admin/modules/organization/index.html', context)


def struktur_organisasi_view(request):
    """Struktur Organisasi view"""
    context = {
        'page_title': 'Struktur Organisasi',
        'page_subtitle': 'Kelola struktur organisasi dan diagram'
    }
    return render(request, 'admin/modules/organization/struktur_organisasi.html', context)


def data_anggota_view(request):
    """Data Anggota view"""
    context = {
        'page_title': 'Data Anggota',
        'page_subtitle': 'Kelola data anggota organisasi'
    }
    return render(request, 'admin/modules/organization/data_anggota.html', context)


def data_jabatan_view(request):
    """Data Jabatan view"""
    context = {
        'page_title': 'Data Jabatan',
        'page_subtitle': 'Kelola data jabatan dalam organisasi'
    }
    return render(request, 'admin/modules/organization/data_jabatan.html', context)


def periode_kepengurusan_view(request):
    """Periode Kepengurusan view"""
    context = {
        'page_title': 'Periode Kepengurusan',
        'page_subtitle': 'Kelola periode kepengurusan organisasi'
    }
    return render(request, 'admin/modules/organization/periode_kepengurusan.html', context)


def galeri_kegiatan_view(request):
    """Galeri Kegiatan view"""
    context = {
        'page_title': 'Galeri Kegiatan',
        'page_subtitle': 'Kelola galeri foto kegiatan organisasi'
    }
    return render(request, 'admin/modules/organization/galeri_kegiatan.html', context)


def kategori_organisasi_view(request):
    """Kategori Organisasi view"""
    context = {
        'page_title': 'Kategori Organisasi',
        'page_subtitle': 'Kelola kategori dan jenis organisasi'
    }
    return render(request, 'admin/modules/organization/kategori_organisasi.html', context)


# Organization Type API Views
@login_required
@require_http_methods(["GET"])
def organization_types_api(request):
    """API for organization types list"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    
    queryset = OrganizationType.objects.all()
    
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    
    queryset = queryset.order_by('-created_at')
    
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'results': [
            {
                'id': item.id,
                'name': item.name,
                'description': item.description,
                'status': item.status,
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M'),
                'organizations_count': item.organization_set.count()
            }
            for item in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }
    }
    
    return JsonResponse(data)


# Jabatan API Views
@require_http_methods(["GET"])
def jabatan_api(request):
    """API for jabatan list"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    
    queryset = Jabatan.objects.all()
    
    if search:
        queryset = queryset.filter(
            Q(nama_jabatan__icontains=search) |
            Q(deskripsi__icontains=search)
        )
    
    queryset = queryset.order_by('level_hierarki', 'nama_jabatan')
    
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'results': [
            {
                'id': item.id,
                'nama': item.nama_jabatan,
                'deskripsi': item.deskripsi,
                'level': item.level_hierarki,
                'is_active': item.is_active,
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M')
            }
            for item in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }
    }
    
    return JsonResponse(data)


@require_http_methods(["GET"])
def jabatan_detail_api(request, jabatan_id):
    """API for jabatan detail"""
    try:
        jabatan = get_object_or_404(Jabatan, id=jabatan_id)
        
        data = {
            'id': jabatan.id,
            'nama': jabatan.nama_jabatan,
            'deskripsi': jabatan.deskripsi or '',
            'level': jabatan.level_hierarki,
            'is_active': jabatan.is_active
        }
        
        return JsonResponse(data)
        
    except Jabatan.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Jabatan tidak ditemukan'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading jabatan: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def jabatan_create_api(request):
    """API for creating jabatan"""
    try:
        data = json.loads(request.body)
        
        jabatan = Jabatan.objects.create(
            nama_jabatan=data['nama'],
            deskripsi=data.get('deskripsi', ''),
            level_hierarki=data.get('level', 1),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Jabatan berhasil ditambahkan',
            'data': {
                'id': jabatan.id,
                'nama': jabatan.nama_jabatan
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan jabatan: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def jabatan_update_api(request, jabatan_id):
    """API for updating jabatan"""
    try:
        jabatan = get_object_or_404(Jabatan, id=jabatan_id)
        data = json.loads(request.body)
        
        jabatan.nama_jabatan = data.get('nama', jabatan.nama_jabatan)
        jabatan.deskripsi = data.get('deskripsi', jabatan.deskripsi)
        jabatan.level_hierarki = data.get('level', jabatan.level_hierarki)
        jabatan.is_active = data.get('is_active', jabatan.is_active)
        jabatan.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Jabatan berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui jabatan: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def jabatan_delete_api(request, jabatan_id):
    """API for deleting jabatan"""
    try:
        jabatan = get_object_or_404(Jabatan, id=jabatan_id)
        jabatan.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Jabatan berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus jabatan: {str(e)}'
        }, status=400)


@require_http_methods(["GET"])
def jabatan_statistics_api(request):
    """API for jabatan statistics"""
    try:
        total_jabatan = Jabatan.objects.count()
        jabatan_eksekutif = Jabatan.objects.filter(level_hierarki__lte=2).count()
        jabatan_struktural = Jabatan.objects.filter(level_hierarki__range=(3, 5)).count()
        jabatan_anggota = Jabatan.objects.filter(level_hierarki__gte=6).count()
        
        return JsonResponse({
            'total_jabatan': total_jabatan,
            'jabatan_eksekutif': jabatan_eksekutif,
            'jabatan_struktural': jabatan_struktural,
            'jabatan_anggota': jabatan_anggota
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memuat statistik jabatan: {str(e)}'
        }, status=400)


# Periode Kepengurusan API Views
@require_http_methods(["GET"])
def periode_kepengurusan_api(request):
    """API for periode kepengurusan list"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    
    queryset = PeriodeKepengurusan.objects.all()
    
    if search:
        queryset = queryset.filter(
            Q(nama_periode__icontains=search) |
            Q(deskripsi__icontains=search)
        )
    
    queryset = queryset.order_by('-tanggal_mulai')
    
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'results': [
            {
                'id': item.id,
                'nama': item.nama_periode,
                'deskripsi': item.deskripsi,
                'tanggal_mulai': item.tanggal_mulai.strftime('%Y-%m-%d'),
                'tanggal_selesai': item.tanggal_selesai.strftime('%Y-%m-%d'),
                'is_active': item.is_active,
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M')
            }
            for item in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }
    }
    
    return JsonResponse(data)


@require_http_methods(["GET"])
def periode_kepengurusan_statistics_api(request):
    """API for periode kepengurusan statistics"""
    try:
        total_periode = PeriodeKepengurusan.objects.count()
        periode_aktif = PeriodeKepengurusan.objects.filter(is_active=True).count()
        periode_selesai = PeriodeKepengurusan.objects.filter(is_active=False).count()
        total_organisasi = Organization.objects.count()
        
        data = {
            'total_periode': total_periode,
            'periode_aktif': periode_aktif,
            'periode_selesai': periode_selesai,
            'total_organisasi': total_organisasi
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading statistics: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def periode_kepengurusan_create_api(request):
    """API for creating periode kepengurusan"""
    try:
        data = json.loads(request.body)
        
        periode = PeriodeKepengurusan.objects.create(
            nama=data['nama'],
            deskripsi=data.get('deskripsi', ''),
            tanggal_mulai=parse_date(data['tanggal_mulai']),
            tanggal_selesai=parse_date(data['tanggal_selesai']),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Periode kepengurusan berhasil ditambahkan',
            'data': {
                'id': periode.id,
                'nama': periode.nama_periode
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan periode kepengurusan: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def periode_kepengurusan_update_api(request, periode_id):
    """API for updating periode kepengurusan"""
    try:
        periode = get_object_or_404(PeriodeKepengurusan, id=periode_id)
        data = json.loads(request.body)
        
        periode.nama_periode = data.get('nama', periode.nama_periode)
        periode.deskripsi = data.get('deskripsi', periode.deskripsi)
        periode.tanggal_mulai = parse_date(data['tanggal_mulai']) if data.get('tanggal_mulai') else periode.tanggal_mulai
        periode.tanggal_selesai = parse_date(data['tanggal_selesai']) if data.get('tanggal_selesai') else periode.tanggal_selesai
        periode.is_active = data.get('is_active', periode.is_active)
        periode.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Periode kepengurusan berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui periode kepengurusan: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def periode_kepengurusan_delete_api(request, periode_id):
    """API for deleting periode kepengurusan"""
    try:
        periode = get_object_or_404(PeriodeKepengurusan, id=periode_id)
        periode.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Periode kepengurusan berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus periode kepengurusan: {str(e)}'
        }, status=400)


# Anggota Organisasi API Views
@require_http_methods(["GET"])
def anggota_organisasi_api(request):
    """API for anggota organisasi list"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    organization_id = request.GET.get('organization_id', '')
    jabatan_id = request.GET.get('jabatan_id', '')
    periode_id = request.GET.get('periode_id', '')
    
    queryset = AnggotaOrganisasi.objects.select_related('penduduk', 'organization', 'jabatan', 'periode')
    
    if search:
        queryset = queryset.filter(
            Q(penduduk__name__icontains=search) |
            Q(penduduk__nik__icontains=search) |
            Q(organization__name__icontains=search)
        )
    
    if organization_id:
        queryset = queryset.filter(organization_id=organization_id)
    
    if jabatan_id:
        queryset = queryset.filter(jabatan_id=jabatan_id)
    
    if periode_id:
        queryset = queryset.filter(periode_id=periode_id)
    
    queryset = queryset.order_by('-created_at')
    
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'results': [
            {
                'id': item.id,
                'penduduk_nama': item.penduduk.name,
                'penduduk_nik': item.penduduk.nik,
                'organization_name': item.organization.name,
                'jabatan_nama': item.jabatan.nama_jabatan,
                'periode_nama': item.periode.nama_periode,
                'tanggal_bergabung': item.tanggal_bergabung.strftime('%Y-%m-%d'),
                'tanggal_keluar': item.tanggal_keluar.strftime('%Y-%m-%d') if item.tanggal_keluar else None,
                'foto_profil': item.foto_profil.url if item.foto_profil else None,
                'status': item.status,
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M')
            }
            for item in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }
    }
    
    return JsonResponse(data)


@require_http_methods(["GET"])
def anggota_organisasi_statistics_api(request):
    """API for anggota organisasi statistics"""
    try:
        total_anggota = AnggotaOrganisasi.objects.count()
        anggota_aktif = AnggotaOrganisasi.objects.filter(status='aktif').count()
        total_organisasi = Organization.objects.count()
        total_jabatan = Jabatan.objects.count()
        
        data = {
            'total_anggota': total_anggota,
            'anggota_aktif': anggota_aktif,
            'total_organisasi': total_organisasi,
            'total_jabatan': total_jabatan
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading statistics: {str(e)}'
        }, status=400)


@require_http_methods(["GET"])
def anggota_organisasi_detail_api(request, anggota_id):
    """API for anggota organisasi detail"""
    try:
        anggota = get_object_or_404(AnggotaOrganisasi.objects.select_related('penduduk', 'organization', 'jabatan', 'periode'), id=anggota_id)
        
        data = {
            'id': anggota.id,
            'penduduk_nama': anggota.penduduk.name,
            'penduduk_nik': anggota.penduduk.nik,
            'penduduk_id': anggota.penduduk.id,
            'organization_name': anggota.organization.name,
            'organization_id': anggota.organization.id,
            'jabatan_nama': anggota.jabatan.nama_jabatan,
            'jabatan_id': anggota.jabatan.id,
            'periode_nama': anggota.periode.nama_periode,
            'periode_id': anggota.periode.id,
            'tanggal_bergabung': anggota.tanggal_bergabung.strftime('%Y-%m-%d') if anggota.tanggal_bergabung else None,
            'tanggal_keluar': anggota.tanggal_keluar.strftime('%Y-%m-%d') if anggota.tanggal_keluar else None,
            'foto_profil': anggota.foto_profil.url if anggota.foto_profil else None,
            'status': anggota.status,
            'created_at': anggota.created_at.strftime('%Y-%m-%d %H:%M'),
            'updated_at': anggota.updated_at.strftime('%Y-%m-%d %H:%M')
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading anggota organisasi: {str(e)}'
        }, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def anggota_organisasi_create_api(request):
    """API for creating anggota organisasi"""
    try:
        penduduk_id = request.POST.get('penduduk_id')
        organization_id = request.POST.get('organization_id')
        jabatan_id = request.POST.get('jabatan_id')
        periode_id = request.POST.get('periode_id')
        tanggal_bergabung = request.POST.get('tanggal_bergabung')
        tanggal_keluar = request.POST.get('tanggal_keluar')
        foto_profil = request.FILES.get('foto_profil')
        status = request.POST.get('status', 'aktif')
        
        anggota = AnggotaOrganisasi.objects.create(
            penduduk_id=penduduk_id,
            organization_id=organization_id,
            jabatan_id=jabatan_id,
            periode_id=periode_id,
            tanggal_bergabung=parse_date(tanggal_bergabung),
            tanggal_keluar=parse_date(tanggal_keluar) if tanggal_keluar else None,
            foto_profil=foto_profil,
            status=status
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Anggota organisasi berhasil ditambahkan',
            'data': {
                'id': anggota.id,
                'nama': anggota.penduduk.name
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan anggota organisasi: {str(e)}'
        }, status=400)


@csrf_exempt
@require_http_methods(["PUT"])
def anggota_organisasi_update_api(request, anggota_id):
    """API for updating anggota organisasi"""
    try:
        anggota = get_object_or_404(AnggotaOrganisasi, id=anggota_id)
        
        penduduk_id = request.POST.get('penduduk_id')
        organization_id = request.POST.get('organization_id')
        jabatan_id = request.POST.get('jabatan_id')
        periode_id = request.POST.get('periode_id')
        tanggal_bergabung = request.POST.get('tanggal_bergabung')
        tanggal_keluar = request.POST.get('tanggal_keluar')
        foto_profil = request.FILES.get('foto_profil')
        status = request.POST.get('status', 'aktif')
        
        if penduduk_id:
            anggota.penduduk_id = penduduk_id
        if organization_id:
            anggota.organization_id = organization_id
        if jabatan_id:
            anggota.jabatan_id = jabatan_id
        if periode_id:
            anggota.periode_id = periode_id
        if tanggal_bergabung:
            anggota.tanggal_bergabung = parse_date(tanggal_bergabung)
        if tanggal_keluar:
            anggota.tanggal_keluar = parse_date(tanggal_keluar)
        elif tanggal_keluar == '':
            anggota.tanggal_keluar = None
        if foto_profil:
            anggota.foto_profil = foto_profil
        
        anggota.status = status
        anggota.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Anggota organisasi berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui anggota organisasi: {str(e)}'
        }, status=400)


@csrf_exempt
@require_http_methods(["DELETE"])
def anggota_organisasi_delete_api(request, anggota_id):
    """API for deleting anggota organisasi"""
    try:
        anggota = get_object_or_404(AnggotaOrganisasi, id=anggota_id)
        anggota.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Anggota organisasi berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus anggota organisasi: {str(e)}'
        }, status=400)


# Galeri Kegiatan API Views
@require_http_methods(["GET"])
def galeri_kegiatan_api(request):
    """API for galeri kegiatan list"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    
    queryset = GaleriKegiatan.objects.all()
    
    if search:
        queryset = queryset.filter(
            Q(judul__icontains=search) |
            Q(deskripsi__icontains=search)
        )
    
    queryset = queryset.order_by('-tanggal_kegiatan')
    
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'results': [
            {
                'id': item.id,
                'judul': item.judul,
                'deskripsi': item.deskripsi,
                'tanggal_kegiatan': item.tanggal_kegiatan.strftime('%Y-%m-%d'),
                'foto': item.foto.url if item.foto else None,
                'is_featured': item.is_featured,
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M')
            }
            for item in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }
    }
    
    return JsonResponse(data)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def galeri_kegiatan_create_api(request):
    """API for creating galeri kegiatan"""
    try:
        judul = request.POST.get('judul')
        deskripsi = request.POST.get('deskripsi', '')
        tanggal_kegiatan = request.POST.get('tanggal_kegiatan')
        foto = request.FILES.get('foto')
        is_featured = request.POST.get('is_featured', 'false').lower() == 'true'
        
        galeri = GaleriKegiatan.objects.create(
            judul=judul,
            deskripsi=deskripsi,
            tanggal_kegiatan=parse_date(tanggal_kegiatan),
            foto=foto,
            is_featured=is_featured
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Galeri kegiatan berhasil ditambahkan',
            'data': {
                'id': galeri.id,
                'judul': galeri.judul
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan galeri kegiatan: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def galeri_kegiatan_update_api(request, galeri_id):
    """API for updating galeri kegiatan"""
    try:
        galeri = get_object_or_404(GaleriKegiatan, id=galeri_id)
        
        galeri.judul = request.POST.get('judul', galeri.judul)
        galeri.deskripsi = request.POST.get('deskripsi', galeri.deskripsi)
        
        if request.POST.get('tanggal_kegiatan'):
            galeri.tanggal_kegiatan = parse_date(request.POST.get('tanggal_kegiatan'))
        
        if request.FILES.get('foto'):
            galeri.foto = request.FILES.get('foto')
        
        galeri.is_featured = request.POST.get('is_featured', str(galeri.is_featured)).lower() == 'true'
        galeri.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Galeri kegiatan berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui galeri kegiatan: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def galeri_kegiatan_delete_api(request, galeri_id):
    """API for deleting galeri kegiatan"""
    try:
        galeri = get_object_or_404(GaleriKegiatan, id=galeri_id)
        galeri.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Galeri kegiatan berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus galeri kegiatan: {str(e)}'
        }, status=400)


@require_http_methods(["GET"])
def galeri_kegiatan_detail_api(request, galeri_id):
    """API for galeri kegiatan detail"""
    galeri = get_object_or_404(GaleriKegiatan.objects.select_related('organization', 'uploaded_by'), id=galeri_id)
    
    data = {
        'id': galeri.id,
        'judul': galeri.judul,
        'deskripsi': galeri.deskripsi,
        'organization_name': galeri.organization.name if galeri.organization else None,
        'organization_id': galeri.organization.id if galeri.organization else None,
        'tanggal_kegiatan': galeri.tanggal_kegiatan.strftime('%Y-%m-%d'),
        'lokasi': galeri.lokasi,
        'fotografer': galeri.fotografer,
        'tags': galeri.tags,
        'foto': galeri.foto.url if galeri.foto else None,
        'is_featured': galeri.is_featured,
        'is_featured': galeri.is_featured,
        'view_count': galeri.view_count,
        'uploaded_by': galeri.uploaded_by.username if galeri.uploaded_by else None,
        'created_at': galeri.created_at.strftime('%Y-%m-%d %H:%M'),
        'updated_at': galeri.updated_at.strftime('%Y-%m-%d %H:%M')
    }
    
    return JsonResponse(data)


# Struktur Organisasi API Views
@require_http_methods(["GET"])
def struktur_organisasi_api(request):
    """API for struktur organisasi list"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    
    queryset = StrukturOrganisasi.objects.select_related('organization', 'periode', 'anggota__penduduk', 'anggota__jabatan')
    
    if search:
        queryset = queryset.filter(
            Q(anggota__penduduk__name__icontains=search) |
            Q(organization__name__icontains=search)
        )
    
    queryset = queryset.order_by('urutan', 'anggota__jabatan__level_hierarki')
    
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'results': [
            {
                'id': item.id,
                'anggota_nama': item.anggota.penduduk.name,
                'jabatan_nama': item.anggota.jabatan.nama_jabatan,
                'organization_name': item.organization.name,
                'periode_nama': item.periode.nama_periode,
                'posisi_x': item.posisi_x,
                'posisi_y': item.posisi_y,
                'urutan': item.urutan,
                'is_visible': item.is_visible,
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M')
            }
            for item in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }
    }
    
    return JsonResponse(data)


@require_http_methods(["GET"])
def struktur_organisasi_statistics_api(request):
    """API for struktur organisasi statistics"""
    try:
        total_struktur = StrukturOrganisasi.objects.count()
        total_anggota = AnggotaOrganisasi.objects.count()
        total_jabatan = Jabatan.objects.count()
        periode_aktif = PeriodeKepengurusan.objects.filter(is_active=True).count()
        
        data = {
            'total_struktur': total_struktur,
            'total_anggota': total_anggota,
            'total_jabatan': total_jabatan,
            'periode_aktif': periode_aktif
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading statistics: {str(e)}'
        }, status=400)


@require_http_methods(["GET"])
def struktur_organisasi_detail_api(request, struktur_id):
    """API for struktur organisasi detail"""
    try:
        struktur = StrukturOrganisasi.objects.select_related('organization', 'periode').get(id=struktur_id)
        
        data = {
            'id': struktur.id,
            'nama': struktur.nama,
            'organisasi': struktur.organization.id,
            'periode': struktur.periode.id,
            'deskripsi': struktur.deskripsi or '',
            'is_active': struktur.is_active,
            'diagram': struktur.diagram.url if struktur.diagram else None
        }
        
        return JsonResponse(data)
        
    except StrukturOrganisasi.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Struktur organisasi tidak ditemukan'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading struktur: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def struktur_organisasi_update_api(request, struktur_id):
    """API for updating struktur organisasi"""
    try:
        struktur = StrukturOrganisasi.objects.get(id=struktur_id)
        data = json.loads(request.body)
        
        struktur.nama = data.get('nama', struktur.nama)
        struktur.organization_id = data.get('organisasi', struktur.organization_id)
        struktur.periode_id = data.get('periode', struktur.periode_id)
        struktur.deskripsi = data.get('deskripsi', struktur.deskripsi)
        struktur.is_active = data.get('is_active', struktur.is_active)
        struktur.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Struktur organisasi berhasil diperbarui',
            'data': {
                'id': struktur.id,
                'nama': struktur.nama
            }
        })
        
    except StrukturOrganisasi.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Struktur organisasi tidak ditemukan'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui struktur organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def struktur_organisasi_delete_api(request, struktur_id):
    """API for deleting struktur organisasi"""
    try:
        struktur = StrukturOrganisasi.objects.get(id=struktur_id)
        struktur.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Struktur organisasi berhasil dihapus'
        })
        
    except StrukturOrganisasi.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Struktur organisasi tidak ditemukan'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus struktur organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def struktur_organisasi_create_api(request):
    """API for creating struktur organisasi"""
    try:
        nama = request.POST.get('nama')
        organization_id = request.POST.get('organization_id')
        periode_id = request.POST.get('periode_id')
        diagram = request.FILES.get('diagram')
        is_active = request.POST.get('is_active', 'true').lower() == 'true'
        
        struktur = StrukturOrganisasi.objects.create(
            nama=nama,
            organization_id=organization_id,
            periode_id=periode_id,
            diagram=diagram,
            is_active=is_active
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Struktur organisasi berhasil ditambahkan',
            'data': {
                'id': struktur.id,
                'nama': struktur.nama
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan struktur organisasi: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["GET"])
def organization_types_api(request):
    """API for organization types list"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    
    queryset = OrganizationType.objects.all()
    
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    
    queryset = queryset.order_by('-created_at')
    
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'results': [
            {
                'id': item.id,
                'name': item.name,
                'description': item.description,
                'is_active': item.is_active,
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M'),
                'organizations_count': item.organization_set.count()
            }
            for item in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["GET"])
def organization_type_detail_api(request, type_id):
    """API for organization type detail"""
    org_type = get_object_or_404(OrganizationType, id=type_id)
    
    data = {
        'id': org_type.id,
        'name': org_type.name,
        'description': org_type.description,
        'is_active': org_type.is_active,
        'created_at': org_type.created_at.strftime('%Y-%m-%d %H:%M'),
        'updated_at': org_type.updated_at.strftime('%Y-%m-%d %H:%M'),
        'organizations_count': org_type.organization_set.count()
    }
    
    return JsonResponse(data)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def organization_type_create_api(request):
    """API for creating organization type"""
    try:
        data = json.loads(request.body)
        
        org_type = OrganizationType.objects.create(
            name=data['name'],
            description=data.get('description', ''),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Jenis organisasi berhasil ditambahkan',
            'data': {
                'id': org_type.id,
                'name': org_type.name,
                'description': org_type.description,
                'is_active': org_type.is_active
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan jenis organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def organization_type_update_api(request, type_id):
    """API for updating organization type"""
    try:
        org_type = get_object_or_404(OrganizationType, id=type_id)
        data = json.loads(request.body)
        
        org_type.name = data.get('name', org_type.name)
        org_type.description = data.get('description', org_type.description)
        org_type.is_active = data.get('is_active', org_type.is_active)
        org_type.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Jenis organisasi berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui jenis organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def organization_type_delete_api(request, type_id):
    """API for deleting organization type"""
    try:
        org_type = get_object_or_404(OrganizationType, id=type_id)
        
        # Check if type is being used
        if org_type.organization_set.exists():
            return JsonResponse({
                'success': False,
                'message': 'Jenis organisasi tidak dapat dihapus karena masih digunakan'
            }, status=400)
        
        org_type.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Jenis organisasi berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus jenis organisasi: {str(e)}'
        }, status=400)


# Organization API Views
@require_http_methods(["GET"])
def organizations_api(request):
    """API for organizations list"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    org_type_id = request.GET.get('organization_type_id', '')
    is_active = request.GET.get('is_active', '')
    
    queryset = Organization.objects.select_related('organization_type', 'leader')
    
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(leader__name__icontains=search)
        )
    
    if org_type_id:
        queryset = queryset.filter(organization_type_id=org_type_id)
    
    if is_active:
        queryset = queryset.filter(is_active=is_active.lower() == 'true')
    
    queryset = queryset.order_by('-created_at')
    
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'results': [
            {
                'id': item.id,
                'name': item.name,
                'organization_type': item.organization_type.name,
                'organization_type_id': item.organization_type.id,
                'description': item.description,
                'established_date': item.established_date.strftime('%Y-%m-%d') if item.established_date else None,
                'leader_name': item.leader.name if item.leader else None,
                'leader_id': item.leader.id if item.leader else None,
                'contact_phone': item.contact_phone,
                'contact_email': item.contact_email,
                'address': item.address,
                'is_active': item.is_active,
                'members_count': item.anggota_detail.filter(status='aktif').count(),
                'events_count': item.events.count(),
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M')
            }
            for item in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["GET"])
def organization_detail_api(request, org_id):
    """API for organization detail"""
    organization = get_object_or_404(Organization.objects.select_related('organization_type', 'leader'), id=org_id)
    
    data = {
        'id': organization.id,
        'name': organization.name,
        'organization_type': organization.organization_type.name,
        'organization_type_id': organization.organization_type.id,
        'description': organization.description,
        'established_date': organization.established_date.strftime('%Y-%m-%d') if organization.established_date else None,
        'leader_name': organization.leader.name if organization.leader else None,
        'leader_id': organization.leader.id if organization.leader else None,
        'contact_phone': organization.contact_phone,
        'contact_email': organization.contact_email,
        'address': organization.address,
        'is_active': organization.is_active,
        'members_count': organization.anggota_detail.filter(status='aktif').count(),
        'events_count': organization.events.count(),
        'documents_count': organization.documents.count(),
        'created_at': organization.created_at.strftime('%Y-%m-%d %H:%M'),
        'updated_at': organization.updated_at.strftime('%Y-%m-%d %H:%M')
    }
    
    return JsonResponse(data)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def organization_create_api(request):
    """API for creating organization"""
    try:
        data = json.loads(request.body)
        
        organization = Organization.objects.create(
            name=data['name'],
            organization_type_id=data['organization_type_id'],
            description=data.get('description', ''),
            established_date=parse_date(data['established_date']) if data.get('established_date') else None,
            leader_id=data.get('leader_id') if data.get('leader_id') else None,
            contact_phone=data.get('contact_phone', ''),
            contact_email=data.get('contact_email', ''),
            address=data.get('address', ''),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Organisasi berhasil ditambahkan',
            'data': {
                'id': organization.id,
                'name': organization.name
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def organization_update_api(request, org_id):
    """API for updating organization"""
    try:
        organization = get_object_or_404(Organization, id=org_id)
        data = json.loads(request.body)
        
        organization.name = data.get('name', organization.name)
        organization.organization_type_id = data.get('organization_type_id', organization.organization_type_id)
        organization.description = data.get('description', organization.description)
        organization.established_date = parse_date(data['established_date']) if data.get('established_date') else organization.established_date
        organization.leader_id = data.get('leader_id') if data.get('leader_id') else organization.leader_id
        organization.contact_phone = data.get('contact_phone', organization.contact_phone)
        organization.contact_email = data.get('contact_email', organization.contact_email)
        organization.address = data.get('address', organization.address)
        organization.is_active = data.get('is_active', organization.is_active)
        organization.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Organisasi berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def organization_delete_api(request, org_id):
    """API for deleting organization"""
    try:
        organization = get_object_or_404(Organization, id=org_id)
        organization.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Organisasi berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus organisasi: {str(e)}'
        }, status=400)


# Organization Members API Views
@login_required
@require_http_methods(["GET"])
def organization_members_api(request):
    """API for organization members list"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    organization_id = request.GET.get('organization_id', '')
    position = request.GET.get('position', '')
    is_active = request.GET.get('is_active', '')
    
    queryset = OrganizationMember.objects.select_related('organization', 'member')
    
    if search:
        queryset = queryset.filter(
            Q(member__nama__icontains=search) |
            Q(member__nik__icontains=search) |
            Q(organization__name__icontains=search)
        )
    
    if organization_id:
        queryset = queryset.filter(organization_id=organization_id)
    
    if position:
        queryset = queryset.filter(position=position)
    
    if is_active:
        queryset = queryset.filter(is_active=is_active.lower() == 'true')
    
    queryset = queryset.order_by('-created_at')
    
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'results': [
            {
                'id': item.id,
                'organization_name': item.organization.name,
                'organization_id': item.organization.id,
                'member_name': item.member.name,
                'member_id': item.member.id,
                'member_nik': item.member.nik,
                'position': item.position,
                'position_display': item.get_position_display(),
                'join_date': item.join_date.strftime('%Y-%m-%d'),
                'end_date': item.end_date.strftime('%Y-%m-%d') if item.end_date else None,
                'is_active': item.is_active,
                'notes': item.notes,
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M')
            }
            for item in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["GET"])
def organization_member_detail_api(request, member_id):
    """API for organization member detail"""
    member = get_object_or_404(OrganizationMember.objects.select_related('organization', 'member'), id=member_id)
    
    data = {
        'id': member.id,
        'organization_name': member.organization.name,
        'organization_id': member.organization.id,
        'member_name': member.member.name,
        'member_id': member.member.id,
        'member_nik': member.member.nik,
        'position': member.position,
        'position_display': member.get_position_display(),
        'join_date': member.join_date.strftime('%Y-%m-%d'),
        'end_date': member.end_date.strftime('%Y-%m-%d') if member.end_date else None,
        'is_active': member.is_active,
        'notes': member.notes,
        'created_at': member.created_at.strftime('%Y-%m-%d %H:%M'),
        'updated_at': member.updated_at.strftime('%Y-%m-%d %H:%M')
    }
    
    return JsonResponse(data)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def organization_member_create_api(request):
    """API for creating organization member"""
    try:
        data = json.loads(request.body)
        
        member = OrganizationMember.objects.create(
            organization_id=data['organization_id'],
            member_id=data['member_id'],
            position=data['position'],
            join_date=parse_date(data['join_date']),
            end_date=parse_date(data['end_date']) if data.get('end_date') else None,
            is_active=data.get('is_active', True),
            notes=data.get('notes', '')
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Anggota organisasi berhasil ditambahkan',
            'data': {
                'id': member.id,
                'member_name': member.member.name,
                'organization_name': member.organization.name
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan anggota organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def organization_member_update_api(request, member_id):
    """API for updating organization member"""
    try:
        member = get_object_or_404(OrganizationMember, id=member_id)
        data = json.loads(request.body)
        
        member.organization_id = data.get('organization_id', member.organization_id)
        member.member_id = data.get('member_id', member.member_id)
        member.position = data.get('position', member.position)
        member.join_date = parse_date(data['join_date']) if data.get('join_date') else member.join_date
        member.end_date = parse_date(data['end_date']) if data.get('end_date') else member.end_date
        member.is_active = data.get('is_active', member.is_active)
        member.notes = data.get('notes', member.notes)
        member.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Anggota organisasi berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui anggota organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def organization_member_delete_api(request, member_id):
    """API for deleting organization member"""
    try:
        member = get_object_or_404(OrganizationMember, id=member_id)
        member.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Anggota organisasi berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus anggota organisasi: {str(e)}'
        }, status=400)


# Organization Events API Views
@login_required
@require_http_methods(["GET"])
def organization_events_api(request):
    """API for organization events list"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    organization_id = request.GET.get('organization_id', '')
    event_type = request.GET.get('event_type', '')
    is_completed = request.GET.get('is_completed', '')
    
    queryset = OrganizationEvent.objects.select_related('organization', 'created_by')
    
    if search:
        queryset = queryset.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(location__icontains=search)
        )
    
    if organization_id:
        queryset = queryset.filter(organization_id=organization_id)
    
    if event_type:
        queryset = queryset.filter(event_type=event_type)
    
    if is_completed:
        queryset = queryset.filter(is_completed=is_completed.lower() == 'true')
    
    queryset = queryset.order_by('-event_date')
    
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'results': [
            {
                'id': item.id,
                'title': item.title,
                'description': item.description,
                'organization_name': item.organization.name,
                'organization_id': item.organization.id,
                'event_type': item.event_type,
                'event_type_display': item.get_event_type_display(),
                'event_date': item.event_date.strftime('%Y-%m-%d %H:%M'),
                'location': item.location,
                'participants_count': item.participants_count,
                'budget': float(item.budget) if item.budget else None,
                'is_completed': item.is_completed,
                'created_by': item.created_by.username if item.created_by else None,
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M')
            }
            for item in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["GET"])
def organization_event_detail_api(request, event_id):
    """API for organization event detail"""
    event = get_object_or_404(OrganizationEvent.objects.select_related('organization', 'created_by'), id=event_id)
    
    data = {
        'id': event.id,
        'title': event.title,
        'description': event.description,
        'organization_name': event.organization.name,
        'organization_id': event.organization.id,
        'event_type': event.event_type,
        'event_type_display': event.get_event_type_display(),
        'event_date': event.event_date.strftime('%Y-%m-%d %H:%M'),
        'location': event.location,
        'participants_count': event.participants_count,
        'budget': float(event.budget) if event.budget else None,
        'is_completed': event.is_completed,
        'created_by': event.created_by.username if event.created_by else None,
        'created_at': event.created_at.strftime('%Y-%m-%d %H:%M'),
        'updated_at': event.updated_at.strftime('%Y-%m-%d %H:%M')
    }
    
    return JsonResponse(data)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def organization_event_create_api(request):
    """API for creating organization event"""
    try:
        data = json.loads(request.body)
        
        event = OrganizationEvent.objects.create(
            organization_id=data['organization_id'],
            title=data['title'],
            description=data.get('description', ''),
            event_type=data['event_type'],
            event_date=datetime.strptime(data['event_date'], '%Y-%m-%d %H:%M'),
            location=data.get('location', ''),
            participants_count=data.get('participants_count', 0),
            budget=data.get('budget') if data.get('budget') else None,
            is_completed=data.get('is_completed', False),
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Acara organisasi berhasil ditambahkan',
            'data': {
                'id': event.id,
                'title': event.title
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan acara organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def organization_event_update_api(request, event_id):
    """API for updating organization event"""
    try:
        event = get_object_or_404(OrganizationEvent, id=event_id)
        data = json.loads(request.body)
        
        event.organization_id = data.get('organization_id', event.organization_id)
        event.title = data.get('title', event.title)
        event.description = data.get('description', event.description)
        event.event_type = data.get('event_type', event.event_type)
        event.event_date = datetime.strptime(data['event_date'], '%Y-%m-%d %H:%M') if data.get('event_date') else event.event_date
        event.location = data.get('location', event.location)
        event.participants_count = data.get('participants_count', event.participants_count)
        event.budget = data.get('budget') if data.get('budget') else event.budget
        event.is_completed = data.get('is_completed', event.is_completed)
        event.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Acara organisasi berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui acara organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def organization_event_delete_api(request, event_id):
    """API for deleting organization event"""
    try:
        event = get_object_or_404(OrganizationEvent, id=event_id)
        event.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Acara organisasi berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus acara organisasi: {str(e)}'
        }, status=400)


# Helper API Views
@login_required
@require_http_methods(["GET"])
def organization_types_dropdown_api(request):
    """API for organization types dropdown"""
    types = OrganizationType.objects.filter(is_active=True).order_by('name')
    
    data = {
        'types': [
            {
                'id': org_type.id,
                'name': org_type.name
            }
            for org_type in types
        ]
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["GET"])
def organizations_dropdown_api(request):
    """API for organizations dropdown"""
    organizations = Organization.objects.filter(is_active=True).order_by('name')
    
    data = {
        'organizations': [
            {
                'id': org.id,
                'name': org.name
            }
            for org in organizations
        ]
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["GET"])
def residents_dropdown_api(request):
    """API for residents dropdown"""
    residents = Penduduk.objects.filter(status_kependudukan='tetap').order_by('nama')
    
    data = {
        'residents': [
            {
                'id': resident.id,
                'name': resident.name,
                'nik': resident.nik
            }
            for resident in residents
        ]
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["GET"])
def organization_statistics_api(request):
    """API for organization statistics"""
    total_organizations = Organization.objects.count()
    active_organizations = Organization.objects.filter(is_active=True).count()
    total_members = AnggotaOrganisasi.objects.filter(status='aktif').count()
    total_events = OrganizationEvent.objects.count()
    completed_events = OrganizationEvent.objects.filter(is_completed=True).count()
    total_galeri = GaleriKegiatan.objects.count()
    featured_galeri = GaleriKegiatan.objects.filter(is_featured=True).count()
    
    # Organization types distribution
    org_types_stats = OrganizationType.objects.annotate(
        org_count=Count('organization')
    ).values('name', 'org_count')
    
    data = {
        'total_organizations': total_organizations,
        'active_organizations': active_organizations,
        'total_members': total_members,
        'total_events': total_events,
        'completed_events': completed_events,
        'total_galeri': total_galeri,
        'featured_galeri': featured_galeri,
        'organization_types': list(org_types_stats)
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["GET"])
def galeri_kegiatan_statistics_api(request):
    """API for galeri kegiatan statistics"""
    total_kegiatan = GaleriKegiatan.objects.count()
    total_foto = GaleriKegiatan.objects.exclude(foto='').count()
    kegiatan_featured = GaleriKegiatan.objects.filter(is_featured=True).count()
    total_views = 0  # This would need a views tracking system
    
    data = {
        'total_kegiatan': total_kegiatan,
        'total_foto': total_foto,
        'kegiatan_featured': kegiatan_featured,
        'total_views': total_views
    }
    
    return JsonResponse(data)


# Kategori Organisasi API Views
@login_required
@require_http_methods(["GET"])
def kategori_organisasi_api(request):
    """API for kategori organisasi list"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    search = request.GET.get('search', '')
    
    queryset = OrganizationType.objects.annotate(
        organization_count=Count('organization')
    )
    
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    
    queryset = queryset.order_by('-created_at')
    
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'results': [
            {
                'id': item.id,
                'name': item.name,
                'description': item.description,
                'organization_count': item.organization_count,
                'is_active': item.is_active,
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M')
            }
            for item in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        }
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["GET"])
def kategori_organisasi_statistics_api(request):
    """API for kategori organisasi statistics"""
    try:
        total_kategori = OrganizationType.objects.count()
        kategori_aktif = OrganizationType.objects.filter(is_active=True).count()
        total_organisasi = Organization.objects.count()
        kategori_dengan_organisasi = OrganizationType.objects.annotate(
            org_count=Count('organization')
        ).filter(org_count__gt=0).count()
        
        data = {
            'total_kategori': total_kategori,
            'kategori_aktif': kategori_aktif,
            'total_organisasi': total_organisasi,
            'kategori_dengan_organisasi': kategori_dengan_organisasi
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading statistics: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def kategori_organisasi_create_api(request):
    """API for creating kategori organisasi"""
    try:
        data = json.loads(request.body)
        
        kategori = OrganizationType.objects.create(
            name=data['name'],
            description=data.get('description', ''),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Kategori organisasi berhasil ditambahkan',
            'data': {
                'id': kategori.id,
                'name': kategori.name
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menambahkan kategori organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["PUT"])
def kategori_organisasi_update_api(request, kategori_id):
    """API for updating kategori organisasi"""
    try:
        kategori = get_object_or_404(OrganizationType, id=kategori_id)
        data = json.loads(request.body)
        
        kategori.name = data.get('name', kategori.name)
        kategori.description = data.get('description', kategori.description)
        kategori.is_active = data.get('is_active', kategori.is_active)
        kategori.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Kategori organisasi berhasil diperbarui'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal memperbarui kategori organisasi: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def kategori_organisasi_delete_api(request, kategori_id):
    """API for deleting kategori organisasi"""
    try:
        kategori = get_object_or_404(OrganizationType, id=kategori_id)
        
        # Check if kategori has organizations
        if kategori.organization_set.exists():
            return JsonResponse({
                'success': False,
                'message': 'Tidak dapat menghapus kategori yang masih memiliki organisasi'
            }, status=400)
        
        kategori.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Kategori organisasi berhasil dihapus'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal menghapus kategori organisasi: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["GET"])
def galeri_kegiatan_export_api(request):
    """API for exporting galeri kegiatan to Excel"""
    try:
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Galeri Kegiatan"
        
        # Headers
        headers = [
            'No', 'Judul', 'Deskripsi', 'Tanggal Kegiatan', 
            'Status', 'Tanggal Dibuat'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get data
        galeri_list = GaleriKegiatan.objects.all().order_by('-created_at')
        
        # Write data
        for row, galeri in enumerate(galeri_list, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=galeri.judul)
            ws.cell(row=row, column=3, value=galeri.deskripsi)
            ws.cell(row=row, column=4, value=galeri.tanggal_kegiatan.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=5, value='Aktif' if galeri.is_active else 'Tidak Aktif')
            ws.cell(row=row, column=6, value=galeri.created_at.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="galeri_kegiatan.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal mengekspor data: {str(e)}'
        }, status=400)


@require_http_methods(["GET"])
def organizations_chart_data_api(request):
    """API for organization chart data"""
    try:
        # Get organization types with counts
        org_types = OrganizationType.objects.annotate(
            count=Count('organization')
        ).filter(count__gt=0)
        
        chart_data = []
        for org_type in org_types:
            chart_data.append({
                'name': org_type.name,
                'value': org_type.count,
                'color': f'hsl({hash(org_type.name) % 360}, 70%, 50%)'
            })
        
        return JsonResponse(chart_data, safe=False)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading chart data: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["GET"])
def jabatan_export_api(request):
    """API for exporting jabatan to Excel"""
    try:
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jabatan"
        
        # Headers
        headers = [
            'No', 'Nama Jabatan', 'Deskripsi', 'Level', 
            'Status', 'Tanggal Dibuat'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get data
        jabatan_list = Jabatan.objects.all().order_by('level_hierarki', 'nama_jabatan')
        
        # Write data
        for row, jabatan in enumerate(jabatan_list, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=jabatan.nama_jabatan)
            ws.cell(row=row, column=3, value=jabatan.deskripsi or '-')
            ws.cell(row=row, column=4, value=jabatan.level_hierarki)
            ws.cell(row=row, column=5, value='Aktif' if jabatan.is_active else 'Tidak Aktif')
            ws.cell(row=row, column=6, value=jabatan.created_at.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="jabatan.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal mengekspor data: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["GET"])
def periode_kepengurusan_export_api(request):
    """API for exporting periode kepengurusan to Excel"""
    try:
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Periode Kepengurusan"
        
        # Headers
        headers = [
            'No', 'Nama Periode', 'Tanggal Mulai', 'Tanggal Selesai', 
            'Status', 'Deskripsi', 'Tanggal Dibuat'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get data
        periode_list = PeriodeKepengurusan.objects.all().order_by('-tanggal_mulai')
        
        # Write data
        for row, periode in enumerate(periode_list, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=periode.nama_periode)
            ws.cell(row=row, column=3, value=periode.tanggal_mulai.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=4, value=periode.tanggal_selesai.strftime('%Y-%m-%d') if periode.tanggal_selesai else '-')
            ws.cell(row=row, column=5, value='Aktif' if periode.is_active else 'Tidak Aktif')
            ws.cell(row=row, column=6, value=periode.deskripsi or '-')
            ws.cell(row=row, column=7, value=periode.created_at.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="periode_kepengurusan.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal mengekspor data: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["GET"])
def anggota_organisasi_export_api(request):
    """API for exporting anggota organisasi to Excel"""
    try:
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Anggota Organisasi"
        
        # Headers
        headers = [
            'No', 'Nama', 'NIK', 'Organisasi', 'Jabatan', 
            'Periode', 'Tanggal Bergabung', 'Tanggal Keluar', 'Status'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get data
        anggota_list = AnggotaOrganisasi.objects.select_related(
            'penduduk', 'organization', 'jabatan', 'periode'
        ).all().order_by('organization__name', 'jabatan__level')
        
        # Write data
        for row, anggota in enumerate(anggota_list, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=anggota.penduduk.name)
            ws.cell(row=row, column=3, value=anggota.penduduk.nik)
            ws.cell(row=row, column=4, value=anggota.organization.name)
            ws.cell(row=row, column=5, value=anggota.jabatan.nama_jabatan)
            ws.cell(row=row, column=6, value=anggota.periode.nama_periode)
            ws.cell(row=row, column=7, value=anggota.tanggal_bergabung.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=8, value=anggota.tanggal_keluar.strftime('%Y-%m-%d') if anggota.tanggal_keluar else '-')
            ws.cell(row=row, column=9, value='Aktif' if anggota.is_active else 'Tidak Aktif')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="anggota_organisasi.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal mengekspor data: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["GET"])
def struktur_organisasi_export_api(request):
    """API for exporting struktur organisasi to Excel"""
    try:
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Struktur Organisasi"
        
        # Headers
        headers = [
            'No', 'Organisasi', 'Nama', 'NIK', 'Jabatan', 
            'Level', 'Periode', 'Status'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get data
        struktur_list = StrukturOrganisasi.objects.select_related(
            'organization', 'anggota__penduduk', 'anggota__jabatan', 'anggota__periode'
        ).all().order_by('organization__name', 'anggota__jabatan__level')
        
        # Write data
        for row, struktur in enumerate(struktur_list, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=struktur.organization.name)
            ws.cell(row=row, column=3, value=struktur.anggota.penduduk.name)
            ws.cell(row=row, column=4, value=struktur.anggota.penduduk.nik)
            ws.cell(row=row, column=5, value=struktur.anggota.jabatan.nama_jabatan)
            ws.cell(row=row, column=6, value=struktur.anggota.jabatan.level_hierarki)
            ws.cell(row=row, column=7, value=struktur.anggota.periode.nama_periode)
            ws.cell(row=row, column=8, value='Aktif' if struktur.is_active else 'Tidak Aktif')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="struktur_organisasi.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal mengekspor data: {str(e)}'
        }, status=400)
